"""Generate the bundled demo workbook used by Loadscope's first-launch demo mode.

Produces `app/resources/Loadscope - Demo Workbook.xlsx` — a fully-populated
6.5 Creedmoor 140gr ELD-M load development workbook with H4350 powder ladder
+ seating depth ladder, a Ballistics DOPE table, and a small Load Library.

The data is REALISTIC but FABRICATED — Chad should review and refine these
numbers before v0.14 ships to make sure they pass the "this could be real"
test from a precision-rifle shooter's eye. All values are within plausible
ranges for the 6.5 Creedmoor 140gr ELD-M with H4350 across mid-velocity nodes.

Usage:
    python3 tools/generate_demo_workbook.py

Output: app/resources/Loadscope - Demo Workbook.xlsx
"""

import os
import shutil
import sys
from copy import copy
from datetime import datetime

# Make sure the project root is importable for import_data + parsers
HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)
sys.path.insert(0, PROJECT)
sys.path.insert(0, os.path.join(PROJECT, "app"))

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

import import_data


TEMPLATE = os.path.join(PROJECT, "Rifle Loads Template (do not edit).xltx")
OUTPUT_DIR = os.path.join(PROJECT, "app", "resources")
OUTPUT_NAME = "Loadscope - Demo Workbook.xlsx"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, OUTPUT_NAME)


# -----------------------------------------------------------------------------
# Realistic 6.5 Creedmoor 140gr ELD-M data
# -----------------------------------------------------------------------------
# Powder Ladder: H4350, 41.5gr - 42.7gr in 0.3 increments.
# Winner: 42.4gr (lowest SD + smallest group + tightest MR).
# All velocities are 10-shot averages. Group sizes / mean radii are in inches.
# These numbers are tuned to look like real PRS-style ladder data; final winner
# is clearly differentiated without being so good it looks fake.

POWDER_LADDER = [
    {  # row 1
        "Tag": "P1",
        "ChargeOrJump": 41.5, "Powder": "H4350",
        "Shots": [2710, 2718, 2705, 2720, 2715, 2712, 2722, 2708, 2716, 2724],
        "AvgVel": 2715.0, "SD": 6.7, "ES": 19.0,
        "BulletWt": 140, "AvgKE": 2293,
        "SessionTitle": "6.5 CM H4350 41.5gr",
        "SessionNote": "Hornady brass, CCI BR-2 primer, 2.875 OAL",
        "Date": "2026-04-12",
    },
    {  # row 2
        "Tag": "P2",
        "ChargeOrJump": 41.8, "Powder": "H4350",
        "Shots": [2732, 2748, 2725, 2740, 2738, 2734, 2742, 2728, 2736, 2747],
        "AvgVel": 2737.0, "SD": 7.8, "ES": 23.0,
        "BulletWt": 140, "AvgKE": 2330,
        "SessionTitle": "6.5 CM H4350 41.8gr",
        "SessionNote": "Hornady brass, CCI BR-2 primer, 2.875 OAL",
        "Date": "2026-04-12",
    },
    {  # row 3
        "Tag": "P3",
        "ChargeOrJump": 42.1, "Powder": "H4350",
        "Shots": [2755, 2762, 2750, 2768, 2760, 2758, 2765, 2752, 2760, 2770],
        "AvgVel": 2760.0, "SD": 6.4, "ES": 20.0,
        "BulletWt": 140, "AvgKE": 2369,
        "SessionTitle": "6.5 CM H4350 42.1gr",
        "SessionNote": "Hornady brass, CCI BR-2 primer, 2.875 OAL",
        "Date": "2026-04-12",
    },
    {  # row 4 — THE WINNER
        "Tag": "P4",
        "ChargeOrJump": 42.4, "Powder": "H4350",
        "Shots": [2778, 2782, 2776, 2784, 2780, 2779, 2782, 2778, 2780, 2781],
        "AvgVel": 2780.0, "SD": 2.5, "ES": 8.0,
        "BulletWt": 140, "AvgKE": 2402,
        "SessionTitle": "6.5 CM H4350 42.4gr (WINNER)",
        "SessionNote": "Hornady brass, CCI BR-2 primer, 2.875 OAL",
        "Date": "2026-04-12",
    },
    {  # row 5
        "Tag": "P5",
        "ChargeOrJump": 42.7, "Powder": "H4350",
        "Shots": [2798, 2805, 2795, 2812, 2800, 2802, 2806, 2796, 2802, 2810],
        "AvgVel": 2803.0, "SD": 5.6, "ES": 17.0,
        "BulletWt": 140, "AvgKE": 2443,
        "SessionTitle": "6.5 CM H4350 42.7gr",
        "SessionNote": "Hornady brass, CCI BR-2 primer, 2.875 OAL",
        "Date": "2026-04-12",
    },
]

# Seating Depth ladder at the winning charge (42.4gr). Jump is bullet-to-rifling.
SEATING_LADDER = [
    {
        "Tag": "S1",
        "ChargeOrJump": 0.005, "Powder": "H4350",
        "Shots": [2782, 2778, 2780, 2784, 2780, 2778, 2782, 2780, 2784, 2782],
        "AvgVel": 2781.0, "SD": 2.2, "ES": 6.0,
        "BulletWt": 140, "AvgKE": 2404,
        "SessionTitle": "Seating depth +0.005 jump",
        "SessionNote": "42.4gr H4350, same brass/primer as PL",
        "Date": "2026-04-26",
    },
    {
        "Tag": "S2",
        "ChargeOrJump": 0.020, "Powder": "H4350",
        "Shots": [2780, 2778, 2782, 2776, 2780, 2778, 2780, 2784, 2780, 2782],
        "AvgVel": 2780.0, "SD": 2.4, "ES": 8.0,
        "BulletWt": 140, "AvgKE": 2402,
        "SessionTitle": "Seating depth +0.020 jump",
        "SessionNote": "42.4gr H4350, same brass/primer as PL",
        "Date": "2026-04-26",
    },
    {  # WINNER for SD ladder
        "Tag": "S3",
        "ChargeOrJump": 0.035, "Powder": "H4350",
        "Shots": [2780, 2781, 2780, 2782, 2780, 2780, 2782, 2780, 2780, 2781],
        "AvgVel": 2780.6, "SD": 0.8, "ES": 2.0,
        "BulletWt": 140, "AvgKE": 2403,
        "SessionTitle": "Seating depth +0.035 jump (WINNER)",
        "SessionNote": "42.4gr H4350, same brass/primer as PL",
        "Date": "2026-04-26",
    },
    {
        "Tag": "S4",
        "ChargeOrJump": 0.050, "Powder": "H4350",
        "Shots": [2780, 2784, 2776, 2782, 2780, 2778, 2784, 2780, 2782, 2774],
        "AvgVel": 2780.0, "SD": 3.4, "ES": 10.0,
        "BulletWt": 140, "AvgKE": 2402,
        "SessionTitle": "Seating depth +0.050 jump",
        "SessionNote": "42.4gr H4350, same brass/primer as PL",
        "Date": "2026-04-26",
    },
    {
        "Tag": "S5",
        "ChargeOrJump": 0.065, "Powder": "H4350",
        "Shots": [2780, 2786, 2774, 2784, 2776, 2782, 2772, 2786, 2776, 2784],
        "AvgVel": 2780.0, "SD": 5.1, "ES": 14.0,
        "BulletWt": 140, "AvgKE": 2402,
        "SessionTitle": "Seating depth +0.065 jump",
        "SessionNote": "42.4gr H4350, same brass/primer as PL",
        "Date": "2026-04-26",
    },
]

# Target-group data corresponding to the above (one group per charge/jump)
# Group sizes / mean radii are typical 100yd 5-shot data for 6.5 CM PRS-grade.
GROUP_RECORDS = [
    # Powder ladder groups (5 charges)
    {"Tag": "P1", "ChargeOrJump": 41.5, "Powder": "H4350", "Date": "2026-04-12",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.62, "WidthIn": 0.58, "HeightIn": 0.55,
     "MRIn": 0.31, "CEPIn": 0.28, "SDRadIn": 0.18,
     "SDVertIn": 0.20, "SDHorizIn": 0.22,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "41.5gr"},
    {"Tag": "P2", "ChargeOrJump": 41.8, "Powder": "H4350", "Date": "2026-04-12",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.81, "WidthIn": 0.76, "HeightIn": 0.72,
     "MRIn": 0.42, "CEPIn": 0.38, "SDRadIn": 0.24,
     "SDVertIn": 0.26, "SDHorizIn": 0.28,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "41.8gr"},
    {"Tag": "P3", "ChargeOrJump": 42.1, "Powder": "H4350", "Date": "2026-04-12",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.45, "WidthIn": 0.42, "HeightIn": 0.40,
     "MRIn": 0.22, "CEPIn": 0.20, "SDRadIn": 0.13,
     "SDVertIn": 0.14, "SDHorizIn": 0.15,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "42.1gr"},
    {"Tag": "P4", "ChargeOrJump": 42.4, "Powder": "H4350", "Date": "2026-04-12",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.38, "WidthIn": 0.34, "HeightIn": 0.32,
     "MRIn": 0.18, "CEPIn": 0.16, "SDRadIn": 0.11,
     "SDVertIn": 0.12, "SDHorizIn": 0.13,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "42.4gr (WINNER)"},
    {"Tag": "P5", "ChargeOrJump": 42.7, "Powder": "H4350", "Date": "2026-04-12",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.58, "WidthIn": 0.55, "HeightIn": 0.52,
     "MRIn": 0.29, "CEPIn": 0.26, "SDRadIn": 0.17,
     "SDVertIn": 0.18, "SDHorizIn": 0.19,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "42.7gr"},

    # Seating depth groups (5 jumps)
    {"Tag": "S1", "ChargeOrJump": 0.005, "Powder": "H4350", "Date": "2026-04-26",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.55, "WidthIn": 0.52, "HeightIn": 0.48,
     "MRIn": 0.27, "CEPIn": 0.24, "SDRadIn": 0.16,
     "SDVertIn": 0.17, "SDHorizIn": 0.18,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "+0.005"},
    {"Tag": "S2", "ChargeOrJump": 0.020, "Powder": "H4350", "Date": "2026-04-26",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.41, "WidthIn": 0.38, "HeightIn": 0.36,
     "MRIn": 0.21, "CEPIn": 0.19, "SDRadIn": 0.12,
     "SDVertIn": 0.13, "SDHorizIn": 0.14,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "+0.020"},
    {"Tag": "S3", "ChargeOrJump": 0.035, "Powder": "H4350", "Date": "2026-04-26",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.32, "WidthIn": 0.29, "HeightIn": 0.27,
     "MRIn": 0.16, "CEPIn": 0.14, "SDRadIn": 0.10,
     "SDVertIn": 0.10, "SDHorizIn": 0.11,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "+0.035 (WINNER)"},
    {"Tag": "S4", "ChargeOrJump": 0.050, "Powder": "H4350", "Date": "2026-04-26",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.48, "WidthIn": 0.45, "HeightIn": 0.42,
     "MRIn": 0.24, "CEPIn": 0.22, "SDRadIn": 0.14,
     "SDVertIn": 0.15, "SDHorizIn": 0.16,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "+0.050"},
    {"Tag": "S5", "ChargeOrJump": 0.065, "Powder": "H4350", "Date": "2026-04-26",
     "Distance": 100, "Caliber": "6.5mm",
     "GroupIn": 0.62, "WidthIn": 0.58, "HeightIn": 0.55,
     "MRIn": 0.31, "CEPIn": 0.28, "SDRadIn": 0.18,
     "SDVertIn": 0.19, "SDHorizIn": 0.21,
     "ElevOffsetIn": 0.0, "WindOffsetIn": 0.0, "Label": "+0.065"},
]


def main():
    if not os.path.isfile(TEMPLATE):
        raise SystemExit(f"Template not found: {TEMPLATE}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Use Python's tempfile copy of the template (loading the .xltx template
    # directly via openpyxl works — it'll save as .xlsx).
    print(f"[1/4] Loading template: {TEMPLATE}")
    wb = load_workbook(TEMPLATE)

    print(f"[2/4] Writing chronograph records ({len(POWDER_LADDER) + len(SEATING_LADDER)} sessions)")
    all_chrono = POWDER_LADDER + SEATING_LADDER
    n = import_data.write_chronograph_records(wb, all_chrono)
    print(f"      wrote {n} chronograph rows")

    print(f"[3/4] Writing group records ({len(GROUP_RECORDS)} groups)")
    n = import_data.write_group_records(wb, GROUP_RECORDS)
    print(f"      wrote {n} group rows")

    print(f"[4/4] Applying workbook repairs (chart refs + composite scoring)")
    import_data.apply_workbook_repairs(wb, GROUP_RECORDS, all_chrono)

    # Populate the user-fill header cells on Load Log + Seating Depth with
    # realistic 6.5 Creedmoor demo values so the workbook looks COMPLETE when
    # a prospect opens it. apply_workbook_repairs only fills date / cartridge /
    # bullet weight — everything else needs to be added by hand.
    print("[5/7] Populating Rifle & Shooter / Load Components / Test Session headers")
    _populate_demo_headers(wb)

    # Write STATIC composite scores + suggested-winner cells. Excel for Mac
    # doesn't reliably recompute the AGGREGATE-based composite formulas after
    # an openpyxl save (a v0.13.3 lesson) — so we precompute the values in
    # Python and write them as static cell values. Excel still shows them
    # correctly; the user just can't change weights without re-running this.
    print("[6/7] Writing static composite scores + suggested-winner row 2")
    _write_static_composite_scores(wb)
    _write_seating_depth_static_values(wb)
    _populate_ballistics_dope(wb)
    _populate_load_library(wb)
    # Remove legacy Range Card / Load Card tabs from prior builds.
    # Chad picked Option A (2026-05-13): drop the in-Excel preview and let
    # the HTML "Print Pocket Range Card" output be the single source of
    # truth for the card. Cleaner design — one card, polished typography.
    for legacy_name in ("Range Card", "Load Card"):
        if legacy_name in wb.sheetnames:
            del wb[legacy_name]

    # Force Excel to do a full recalc when opening this workbook. Without this,
    # Excel may trust the (stale or empty) cached values from openpyxl's save
    # and leave the suggested-winner / composite-score / heat-map cells blank.
    print("[7/7] Setting forceFullCalcOnLoad so Excel recomputes everything on open")
    wb.calculation.fullCalcOnLoad = True

    # Convert from template to regular workbook. Without this, openpyxl
    # preserves the .xltx content-type in [Content_Types].xml as
    # "spreadsheetml.template.main+xml", which Excel for Mac rejects with
    # "file format or file extension is not valid".
    wb.template = False

    print(f"\nSaving: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print(f"OK — {os.path.getsize(OUTPUT_PATH)} bytes")

    # Post-save: resize VML comment boxes (must run after save since VML
    # files only exist inside the saved .xlsx zip)
    n_resized = import_data.resize_comment_boxes(OUTPUT_PATH)
    print(f"Resized {n_resized} comment box(es) to 360x220 px")

    print(f"\nNext: open in Excel + visually verify the data looks realistic.")


def _write_static_composite_scores(wb):
    """Write static (non-formula) composite scores for the powder ladder.

    Excel-Mac doesn't reliably recompute the AGGREGATE-based composite chain
    after an openpyxl save (v0.13.3 lesson — see [[loadscope-v014-inflight]]).
    The workbook ships pre-computed values so a prospect sees the suggested
    winner immediately, no Cmd+= required.

    Targets (Load Log):
      - O16:O20 — composite score per row (lower = better)
      - B2 — suggested-charge value
      - G2/J2/L2 — winning row's AvgVel / SD / MR

    Targets (Charts):
      - L18:L22 — composite scores (mirror of Load Log O column)
      - R18:R22 — per-row "Best in" labels (Vert ✓ / Vel ✓ SD ✓ / MR ✓)
    """
    ll = wb["Load Log"]
    ch = wb["Charts"]

    # Default weights from Charts B11/D11/F11/H11 = 0.3 / 0.2 / 0.2 / 0.3
    w_group, w_sd, w_mr, w_vert = 0.3, 0.2, 0.2, 0.3

    # Pair each powder-ladder chronograph record with its group record (matched by Tag)
    by_tag = {g["Tag"]: g for g in GROUP_RECORDS}
    candidates = []
    for chrono in POWDER_LADDER:
        grp = by_tag.get(chrono["Tag"])
        if not grp:
            continue
        candidates.append({
            "charge": chrono["ChargeOrJump"],
            "vel": chrono["AvgVel"],
            "sd": chrono["SD"],
            "group": grp["GroupIn"],
            "mr": grp["MRIn"],
            "vert": grp["SDVertIn"],
        })

    if not candidates:
        return

    # Normalize each metric — 0 (best) to 1 (worst). Lower SD/group/MR/vert is better.
    def normalize(values):
        lo, hi = min(values), max(values)
        rng = hi - lo
        return [0.0 if rng == 0 else (v - lo) / rng for v in values]

    norm_group = normalize([c["group"] for c in candidates])
    norm_sd = normalize([c["sd"] for c in candidates])
    norm_mr = normalize([c["mr"] for c in candidates])
    norm_vert = normalize([c["vert"] for c in candidates])

    composites = [
        w_group * ng + w_sd * nsd + w_mr * nmr + w_vert * nvert
        for ng, nsd, nmr, nvert in zip(norm_group, norm_sd, norm_mr, norm_vert)
    ]
    # Floor at 0.001 so the winner's score actually displays as a number rather
    # than getting eaten by Excel's `[=NA()]"";0.000` format (which appears to
    # hide exactly-zero values in some Excel-Mac builds).
    composites = [max(0.001, c) for c in composites]

    # Winner = lowest composite. Used for the suggested-charge summary row.
    winner_idx = composites.index(min(composites))
    winner = candidates[winner_idx]

    # Determine the per-row "Best in" label by concatenating EVERY category
    # the row is best in. Matches the spirit of the existing Charts!R formula
    # which builds a multi-tag string ("Vel ✓ SD ✓ MR ✓ Vert ✓") for each row.
    sds = [c["sd"] for c in candidates]
    groups = [c["group"] for c in candidates]
    mrs = [c["mr"] for c in candidates]
    verts = [c["vert"] for c in candidates]

    def best_idx(values):
        return values.index(min(values))

    bests = {
        "Vel ✓": best_idx(groups),   # D column in workbook = group/velocity composite
        "SD ✓": best_idx(sds),
        "MR ✓": best_idx(mrs),
        "Vert ✓": best_idx(verts),
    }

    best_in_labels = []
    for i in range(len(candidates)):
        tags = [tag for tag, best_i in bests.items() if best_i == i]
        # When a row wins all four metrics, the short "All metrics ✓" label
        # reads better than concatenating four "X ✓" tags (avoids awkward
        # wrap in the merged "Best in:" cell).
        if len(tags) == 4:
            best_in_labels.append("All metrics ✓")
        else:
            best_in_labels.append(" ".join(tags))

    # --- Write Load Log O16:O20 — composite per row ---
    for i, comp in enumerate(composites):
        ll.cell(row=16 + i, column=15).value = round(comp, 3)

    # --- Write Load Log row 2 — suggested-charge summary ---
    # D2 is the top-left of the D2:E2 merge that holds the suggested-charge
    # value (A2:C2 is the merged "SUGGESTED CHARGE →" label). G2/J2/L2 are
    # single cells holding Avg Vel / SD / MR — overwriting the cross-sheet
    # ArrayFormula refs with the precomputed static values.
    ll["D2"].value = winner["charge"]
    ll["G2"].value = winner["vel"]
    ll["J2"].value = winner["sd"]
    ll["L2"].value = winner["mr"]
    # M2 — the rich-text "Best in: <label>" is set by apply_workbook_repairs

    # --- Charts T18:U25 — sorted-charge + sorted-velocity helper columns ---
    # Used by chart1 (Velocity vs Powder Charge) and chart4 to plot a clean
    # ascending line instead of a source-order zigzag. Chad hand-wrote these
    # in v0.13.3 with his 7 SAUM data; the template inherited those values.
    # Overwrite with our 6.5 CM data and pad the rest with #N/A error cells
    # so Excel keeps the range as uniformly-numeric (numRef stays numRef on
    # save; mixing types causes Excel to convert to strRef, breaking the chart).
    sorted_pairs = sorted([(c["charge"], c["vel"]) for c in candidates])
    for i, (charge, vel) in enumerate(sorted_pairs):
        ch.cell(row=18 + i, column=20).value = charge   # T
        ch.cell(row=18 + i, column=21).value = vel      # U
    # Pad remaining rows with #N/A so the chart range stays numeric.
    # openpyxl writes the literal string "#N/A" which Excel interprets as
    # the error value when the cell type is "e". Setting via formula triggers
    # that: `=NA()` produces an #N/A error when evaluated.
    for r in range(18 + len(sorted_pairs), 26):
        ch.cell(row=r, column=20).value = "=NA()"
        ch.cell(row=r, column=21).value = "=NA()"

    # --- Force test mode on Charts so all 5 candidates align with our data ---
    # Charts!B100 normally auto-toggles: "window" if 5+ candidates, "test"
    # otherwise. Window mode skips the first Load Log row (B16 = P1), shifting
    # candidates by 1. Our precomputed L18:L22 assume sequential alignment
    # with B16:B20 (test mode order). Pin B100 to "test" so the winner
    # selection matches our data.
    ch["B100"].value = "test"

    # --- Write Charts D18:K22 — metric + normalized blocks for each candidate ---
    # In test mode (which we force above):
    #   D = Avg Vel (fps)         H = Norm Spread (normalized D)
    #   E = SD (fps)              I = Norm SD
    #   F = Mean Radius (MOA)     J = Norm MR
    #   G = SD-Vert (MOA)         K = Norm SD-Vert
    # At 100 yd, 1 MOA ≈ 1 inch, so MR/Vert inches ≈ MOA. Fine for demo realism.
    # Floor normalized values at 0.001 too — same Excel-Mac format issue as
    # composites (the winner's 0 gets hidden by `[=NA()]"";0.000`).
    for i, c in enumerate(candidates):
        r = 18 + i
        ch.cell(row=r, column=4).value = c["vel"]    # D = Avg Vel
        ch.cell(row=r, column=5).value = c["sd"]     # E = SD
        ch.cell(row=r, column=6).value = c["mr"]     # F = MR (≈MOA at 100yd)
        ch.cell(row=r, column=7).value = c["vert"]   # G = SD-Vert (≈MOA at 100yd)
        ch.cell(row=r, column=8).value = round(max(0.001, norm_group[i]), 3)
        ch.cell(row=r, column=9).value = round(max(0.001, norm_sd[i]), 3)
        ch.cell(row=r, column=10).value = round(max(0.001, norm_mr[i]), 3)
        ch.cell(row=r, column=11).value = round(max(0.001, norm_vert[i]), 3)

    # --- Charts B18:C22 (Low/High) intentionally LEFT EMPTY in test mode ---
    # Chad's clarification 2026-05-13: those columns only make sense in
    # window mode (low/middle/high charges of a 3-step rolling window). In
    # test mode they're meaningless — leave blank per template formula intent.
    for i in range(len(candidates)):
        ch.cell(row=18 + i, column=2).value = None
        ch.cell(row=18 + i, column=3).value = None
    # Also fix the column headers in test mode (template uses
    # IF($B$100="window","Middle Charge",...) which Excel-Mac doesn't
    # always recompute).
    ch["A17"].value = "Charge"
    ch["D17"].value = "Avg Vel (fps)"

    # --- Write Charts L18:L22 + R18:R22 (composite + per-row best-in tags) ---
    for i, (comp, label) in enumerate(zip(composites, best_in_labels)):
        ch.cell(row=18 + i, column=12).value = round(comp, 3)  # L column
        if label:
            ch.cell(row=18 + i, column=18).value = label  # R column

    # --- Fix D4/E4 — clear odd dark-blue + white fills that show empty in test mode ---
    # Template styles these for the window-mode "Charge: Spread:" label area.
    # In test mode the formulas return "" but the fills stay, looking unfinished.
    no_fill = PatternFill(fill_type=None)
    ch["D4"].fill = no_fill
    ch["E4"].fill = no_fill

    # --- Write Charts row 3-5 winner-summary cells as STATIC values ---
    # These cells normally hold ArrayFormulas (E5/G3/G4) or LOOKUP formulas
    # (B3/E3) that Excel-Mac doesn't reliably recompute. Overwrite with the
    # precomputed winner values. G5 ("Best in: <tag>") is hardcoded "MR ✓"
    # in the template — overwrite with the full multi-tag winner label.
    winner_for_charts = candidates[winner_idx]
    ch["B3"].value = winner_for_charts["charge"]    # Suggested charge
    ch["E3"].value = round(composites[winner_idx], 3)  # Composite
    ch["G3"].value = winner_for_charts["sd"]         # SD
    ch["G4"].value = winner_for_charts["mr"]         # Group/MR (label F4 says "Group:")
    ch["E5"].value = winner_for_charts["vel"]        # Avg Vel
    ch["G5"].value = best_in_labels[winner_idx] or "Composite ✓"  # Best-in tags

    # --- Restore M2 styling (red fill + yellow font) + write static "Best in: <label>" ---
    # apply_workbook_repairs writes a formula to M2 that uses AGGREGATE +
    # INDEX/MATCH to look up the winner's "Best in" label from Charts!R18:R25.
    # Excel-Mac doesn't reliably compute that chain — we precompute and write
    # the static string here. Styling is template-default (white/black) so we
    # also apply Chad's v0.13.3 red/yellow design.
    red_yellow_fill = PatternFill(
        start_color="FFC00000", end_color="FFC00000", fill_type="solid"
    )
    yellow_font = Font(color="FFFFFF00", bold=True)
    # Wrap_text=True so long Best-in labels (4 tags) flow onto a second line
    # instead of getting clipped at the cell edge.
    centered_wrapped = Alignment(horizontal="center", vertical="center", wrap_text=True)
    winner_label = best_in_labels[winner_idx] or "Composite ✓"
    static_best_in = f"Best in:  {winner_label}"
    ll["M2"].value = static_best_in
    ll["M2"].fill = red_yellow_fill
    ll["M2"].font = yellow_font
    ll["M2"].alignment = centered_wrapped
    # Bump LL row 2 height so wrap has room
    current_h = ll.row_dimensions[2].height if 2 in ll.row_dimensions else None
    if current_h is None or current_h < 38:
        ll.row_dimensions[2].height = 38
    # Seating Depth O2 — same styling. The value is written by
    # _write_seating_depth_static_values which runs separately.
    if "Seating Depth" in wb.sheetnames:
        sd_o2 = wb["Seating Depth"]["O2"]
        sd_o2.fill = red_yellow_fill
        sd_o2.font = yellow_font
        sd_o2.alignment = centered_wrapped


def _populate_ballistics_dope(wb):
    """Fill the Ballistics tab DOPE table with realistic 6.5 Creedmoor numbers.

    Values are typical for a 140gr ELD-M at 2780 fps muzzle velocity with a
    100-yard zero, at sea level (G7 BC ≈ 0.323). User-fill cells:
      B = Mils Elev
      D = MOA Elev
      F = Wind Mils / 10 mph
      H = Wind MOA / 10 mph
      J = TOF (sec)
    The C/E/G/I "Clicks" columns are formulas that compute from B/D/F/H.
    """
    if "Ballistics" not in wb.sheetnames:
        return
    b = wb["Ballistics"]

    # (yards, mils_elev, moa_elev, wind_mils_per_10mph, wind_moa_per_10mph, tof_sec)
    dope = [
        (100,  0.0,   0.0,   0.0,  0.0,  0.11),
        (200,  0.1,   0.5,   0.3,  1.0,  0.23),
        (300,  0.6,   2.0,   0.5,  1.7,  0.36),
        (400,  1.2,   4.1,   0.7,  2.4,  0.50),
        (500,  2.0,   6.9,   0.9,  3.1,  0.66),
        (600,  2.8,   9.6,   1.2,  4.1,  0.83),
        (700,  3.7,  12.7,   1.5,  5.1,  1.02),
        (800,  4.8,  16.5,   1.8,  6.2,  1.23),
        (900,  6.0,  20.6,   2.1,  7.2,  1.46),
        (1000, 7.3,  25.0,   2.4,  8.2,  1.71),
    ]
    # Rows 9-18 correspond to 100-1000 yards
    for i, (yd, mils, moa, wind_mils, wind_moa, tof) in enumerate(dope):
        r = 9 + i
        b.cell(row=r, column=2).value = mils       # B = Mils Elev
        b.cell(row=r, column=4).value = moa        # D = MOA Elev
        b.cell(row=r, column=6).value = wind_mils  # F = Wind Mils / 10 mph
        b.cell(row=r, column=8).value = wind_moa   # H = Wind MOA / 10 mph
        b.cell(row=r, column=10).value = tof       # J = TOF (sec)

    # Row 5 / 6 — rifle setup header that the Pocket Range Card reads. These
    # cells are normally formulas pulling from Load Log; openpyxl save doesn't
    # compute formulas, so the cached values are blank → Pocket Card shows
    # missing fps/gr. Write static values directly.
    b["B5"].value = "Tikka T3X CTR 6.5 CM"
    b["E5"].value = "Hornady 140gr ELD-M"
    b["H5"].value = 42.4   # charge gr
    b["K5"].value = 2780   # avg vel fps
    b["B6"].value = "Leupold Mark 5HD 5-25x56"
    # E6 (Zero yd), H6 (Sight Ht), K6 (Twist) are user-fill defaults
    # already set in the template (100, 1.75, 1:8 RH)


def _add_load_card_tab(wb):
    """Insert a 'Load Card' sheet next to Ballistics that visually mirrors
    the printable Pocket Range Card in Excel cells.

    Prospects can see the card preview WITHOUT clicking 'Print Pocket Range
    Card' (which opens an HTML in browser). Cells layout mimics the 6x4
    landscape card: title bar, rifle/bullet/charge/vel header, scope info,
    DOPE table 100-1000yd, footer notes.

    Pulls static values written elsewhere in the demo build so the card
    displays correctly even without Excel recompute.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Position right after Ballistics in the tab bar
    if "Range Card" in wb.sheetnames:
        return  # already exists; nothing to do
    # Clean up legacy "Load Card" name if a previous build created it
    if "Load Card" in wb.sheetnames:
        del wb["Load Card"]
    ballistics_idx = wb.sheetnames.index("Ballistics") if "Ballistics" in wb.sheetnames else 7
    card = wb.create_sheet("Range Card", index=ballistics_idx + 1)

    # Demo data — mirror what _populate_ballistics_dope writes
    rifle = "Tikka T3X CTR 6.5 CM"
    bullet = "Hornady 140gr ELD-M"
    charge_gr = 42.4
    vel_fps = 2780
    scope = "Leupold Mark 5HD 5-25x56"
    click = "0.1 Mil"
    zero = 100
    sight_ht = 1.75
    twist = "1:8 RH"
    dope = [
        (100,  0.0,   0,  0.0,   0,  0.0,  0,  0.0,  0,  0.11),
        (200,  0.1,   1,  0.5,   2,  0.3,  3,  1.0,  4,  0.23),
        (300,  0.6,   6,  2.0,   8,  0.5,  5,  1.7,  7,  0.36),
        (400,  1.2,  12,  4.1,  16,  0.7,  7,  2.4, 10,  0.50),
        (500,  2.0,  20,  6.9,  28,  0.9,  9,  3.1, 12,  0.66),
        (600,  2.8,  28,  9.6,  38,  1.2, 12,  4.1, 16,  0.83),
        (700,  3.7,  37, 12.7,  51,  1.5, 15,  5.1, 20,  1.02),
        (800,  4.8,  48, 16.5,  66,  1.8, 18,  6.2, 25,  1.23),
        (900,  6.0,  60, 20.6,  82,  2.1, 21,  7.2, 29,  1.46),
        (1000, 7.3,  73, 25.0, 100,  2.4, 24,  8.2, 33,  1.71),
    ]

    # Hide Excel gridlines + header row/col so the tab reads as a printed
    # artifact, not a spreadsheet.
    card.sheet_view.showGridLines = False
    card.sheet_view.showRowColHeaders = False
    card.sheet_view.zoomScale = 130

    # Palette matches the HTML pocket card (app/pocket_card.py):
    #   orange title bar: #d97706 / dark border #b45309
    #   DOPE header (spanner): dark navy #1f2229
    #   Sub-header: cream #f0eee9
    #   Range column: warm cream #faf8f4
    #   Clicks columns: muted gray #777
    #   TOF column: muted gray #555
    orange_fill = PatternFill(start_color="FFD97706", end_color="FFD97706", fill_type="solid")
    spanner_fill = PatternFill(start_color="FF1F2229", end_color="FF1F2229", fill_type="solid")
    subheader_fill = PatternFill(start_color="FFF0EEE9", end_color="FFF0EEE9", fill_type="solid")
    range_fill = PatternFill(start_color="FFFAF8F4", end_color="FFFAF8F4", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")

    title_font = Font(color="FFFFFFFF", bold=True, size=14, name="Helvetica Neue")
    subtitle_font = Font(color="FFFFFFFF", bold=True, size=10, name="Helvetica Neue")
    titleline_font = Font(color="FF000000", bold=True, size=12, name="Helvetica Neue")
    setup_label_font = Font(color="FF555555", bold=True, size=10, name="Helvetica Neue")
    setup_value_font = Font(color="FF000000", size=10, name="Helvetica Neue")
    spanner_font = Font(color="FFFFFFFF", bold=True, size=10, name="Helvetica Neue")
    subheader_font = Font(color="FF000000", bold=True, size=10, name="Helvetica Neue")
    data_font = Font(color="FF000000", size=10, name="Helvetica Neue")
    range_data_font = Font(color="FF000000", bold=True, size=11, name="Helvetica Neue")
    clk_font = Font(color="FF777777", italic=True, size=9, name="Helvetica Neue")
    tof_font = Font(color="FF555555", size=9, name="Helvetica Neue")
    footer_font = Font(color="FF888888", italic=True, size=9, name="Helvetica Neue")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center", indent=1)

    thin_gray = Side(style="thin", color="FFDDDDDD")
    bottom_dark = Side(style="thin", color="FFB45309")  # orange-bar separator

    col_widths = [11, 7, 6, 7, 6, 8, 6, 8, 6, 9]
    for i, w in enumerate(col_widths, start=1):
        card.column_dimensions[get_column_letter(i)].width = w

    # White background on the whole used range so the card reads as a
    # clean printed page (after we hid the gridlines above)
    for r in range(1, 18):
        for col_idx in range(1, 11):
            card.cell(row=r, column=col_idx).fill = white_fill

    # Row 1 — Title (large, with bottom border as underline)
    card.merge_cells("A1:J1")
    c = card["A1"]
    c.value = "POCKET RANGE CARD"
    c.font = title_font
    c.alignment = center
    # Underline via bottom border on each cell of the merged row
    for col_idx in range(1, 11):
        card.cell(row=1, column=col_idx).border = Border(bottom=thick_bottom)
    card.row_dimensions[1].height = 36

    # Row 2 — Rifle / Bullet / Charge / Vel (the prominent identifier)
    card.merge_cells("A2:J2")
    c = card["A2"]
    c.value = f"{rifle}  •  {bullet}  •  {charge_gr}gr  •  {vel_fps} fps"
    c.font = sub_font
    c.alignment = center
    card.row_dimensions[2].height = 26

    # Row 3 — Scope info
    card.merge_cells("A3:J3")
    c = card["A3"]
    c.value = (
        f"Scope: {scope}   ·   Click: {click}   ·   "
        f"Zero: {zero} yd   ·   Sight Ht: {sight_ht}\"   ·   Twist: {twist}"
    )
    c.alignment = center
    c.font = scope_font
    card.row_dimensions[3].height = 22

    # Row 4 — visual breathing space
    card.row_dimensions[4].height = 8

    # Row 5 — DOPE table header
    headers = [
        "Range\n(yd)", "Mils\nElev", "Mil\nClicks",
        "MOA\nElev", "MOA\nClicks",
        "Wind\nMils/10", "Wind\nMil Clk",
        "Wind\nMOA/10", "Wind\nMOA Clk", "TOF\n(sec)"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = card.cell(row=5, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = table_border
    card.row_dimensions[5].height = 34

    # Rows 6-15 — DOPE table data with subtle alternating-row stripes
    for i, row in enumerate(dope):
        r = 6 + i
        row_fill = alt_row_fill if i % 2 == 1 else white_fill
        for col_idx, value in enumerate(row, start=1):
            cell = card.cell(row=r, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = center
            cell.border = table_border
        card.row_dimensions[r].height = 20

    # Row 16 — visual breathing space
    card.row_dimensions[16].height = 8

    # Row 17 — Footer note (subtle, italic)
    card.merge_cells("A17:J17")
    c = card["A17"]
    c.value = (
        "Click \"Print Pocket Range Card\" on the Ballistics tab to print "
        "this card as a 4×6 page."
    )
    c.font = Font(color="FF888888", italic=True, size=10, name="Helvetica Neue")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    card.row_dimensions[17].height = 32

    # Page setup for printing — landscape 6x4
    card.page_setup.orientation = card.ORIENTATION_LANDSCAPE
    card.page_setup.paperSize = 1  # Letter; user can change to custom 6x4
    card.page_margins.left = 0.25
    card.page_margins.right = 0.25
    card.page_margins.top = 0.25
    card.page_margins.bottom = 0.25


def _populate_load_library(wb):
    """Populate Load Library with 3 historical winning loads so prospects
    see what a built-up library looks like.

    Entry 1 = current demo's winning load (6.5 CM 140gr ELD-M H4350 42.4gr).
    Entry 2 = older 6.5 CM season with different bullet (Berger Hybrid).
    Entry 3 = different cartridge (6 ARC) showing the library grows over
    rifles and seasons.
    """
    if "Load Library" not in wb.sheetnames:
        return
    ll_lib = wb["Load Library"]
    from datetime import datetime
    entries = [
        # (date_added, load_name, rifle, bullet, bullet_wt, powder, charge,
        #  primer, brass, cbto, jump, avg_vel, sd, group, mr, notes)
        (
            datetime(2026, 5, 1),
            "6.5 CM 140 ELD-M / H4350 42.4 / 0.035 jump",
            "Tikka T3X CTR 6.5 CM",
            "Hornady ELD-M",
            140,
            "Hodgdon H4350",
            42.4,
            "CCI BR-2",
            "Hornady",
            2.224,
            0.035,
            2780,
            2.5,
            0.38,
            0.18,
            "Confirmed at 100 yd, 65°F, 1500 ft DA. 10-shot string."
        ),
        (
            datetime(2025, 11, 10),
            "6.5 CM 140 Berger Hybrid / H4350 42.1 / 0.020 jump",
            "Tikka T3X CTR 6.5 CM",
            "Berger Hybrid Target",
            140,
            "Hodgdon H4350",
            42.1,
            "Fed 210M",
            "Lapua",
            2.218,
            0.020,
            2745,
            5.1,
            0.52,
            0.24,
            "Prior season — switched to ELD-M for lower SD."
        ),
        (
            datetime(2026, 3, 18),
            "6 ARC 108 ELD-M / CFE 223 26.5 / 0.040 jump",
            "Bergara B14 HMR 6 ARC",
            "Hornady ELD-M",
            108,
            "Hodgdon CFE 223",
            26.5,
            "CCI 450",
            "Hornady",
            1.910,
            0.040,
            2680,
            6.8,
            0.55,
            0.28,
            "Truck gun build. Good cold-bore consistency."
        ),
    ]
    # Set column widths sized to fit content WITHOUT bloat, so the whole
    # library prints on one landscape page. Replaces (not just widens) the
    # template widths — some are way oversized (K was 22 chars for a 5-char
    # CBTO number).
    from openpyxl.styles import Alignment
    LL_COL_WIDTHS = {
        "A": 4,   # # (sequential)
        "B": 11,  # Date Added
        "C": 25,  # Load Name
        "D": 18,  # Rifle
        "E": 15,  # Bullet
        "F": 8,   # Bullet Wt (gr)
        "G": 14,  # Powder
        "H": 8,   # Charge (gr)
        "I": 10,  # Primer
        "J": 10,  # Brass
        "K": 8,   # CBTO (in)
        "L": 8,   # Jump (in)
        "M": 9,   # Avg Vel (fps)
        "N": 7,   # SD (fps)
        "O": 8,   # Group (MOA)
        "P": 8,   # MR (MOA)
        "Q": 22,  # Notes / Conditions
    }
    for col_letter, width in LL_COL_WIDTHS.items():
        ll_lib.column_dimensions[col_letter].width = width

    # Print setup — landscape, fit-to-1-page-wide, no fixed page-tall so
    # the library can grow down across pages as the user adds more loads.
    ll_lib.page_setup.orientation = ll_lib.ORIENTATION_LANDSCAPE
    ll_lib.page_setup.fitToWidth = 1
    ll_lib.page_setup.fitToHeight = 0
    ll_lib.sheet_properties.pageSetUpPr.fitToPage = True
    ll_lib.page_margins.left = 0.25
    ll_lib.page_margins.right = 0.25

    # Rows 5-19 hold up to 15 entries. Write our 3 starting at row 5.
    for i, (date, name, rifle, bullet, bwt, powder, charge, primer, brass,
            cbto, jump, vel, sd, group, mr, notes) in enumerate(entries):
        r = 5 + i
        ll_lib.cell(row=r, column=1).value = i + 1
        ll_lib.cell(row=r, column=2).value = date
        ll_lib.cell(row=r, column=2).number_format = "mm/dd/yyyy"
        ll_lib.cell(row=r, column=3).value = name
        ll_lib.cell(row=r, column=4).value = rifle
        ll_lib.cell(row=r, column=5).value = bullet
        ll_lib.cell(row=r, column=6).value = bwt
        ll_lib.cell(row=r, column=7).value = powder
        ll_lib.cell(row=r, column=8).value = charge
        ll_lib.cell(row=r, column=9).value = primer
        ll_lib.cell(row=r, column=10).value = brass
        ll_lib.cell(row=r, column=11).value = cbto
        ll_lib.cell(row=r, column=12).value = jump
        ll_lib.cell(row=r, column=13).value = vel
        ll_lib.cell(row=r, column=14).value = sd
        ll_lib.cell(row=r, column=15).value = group
        ll_lib.cell(row=r, column=16).value = mr
        ll_lib.cell(row=r, column=17).value = notes
        # Wrap text on the Notes column so multi-sentence notes display
        ll_lib.cell(row=r, column=17).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        # Bump row height for wrapped notes
        ll_lib.row_dimensions[r].height = 32


def _write_seating_depth_static_values(wb):
    """Write static composite scores + suggested-winner cells for the Seating Depth tab.

    Same shape as _write_static_composite_scores but targeting Seating Depth's
    analysis block at rows 30-37 (vs Charts' 18-25), with Seating Depth's
    default weights (0.15 Vel / 0.25 SD / 0.25 MR / 0.35 Vert).
    """
    if "Seating Depth" not in wb.sheetnames:
        return
    sd = wb["Seating Depth"]

    # SD-specific weights (Seating Depth!C28/F28/I28/L28 = 0.15/0.25/0.25/0.35)
    w_vel, w_sd, w_mr, w_vert = 0.15, 0.25, 0.25, 0.35

    # Pair each SD chronograph record with its group record (matched by Tag)
    by_tag = {g["Tag"]: g for g in GROUP_RECORDS}
    candidates = []
    for chrono in SEATING_LADDER:
        grp = by_tag.get(chrono["Tag"])
        if not grp:
            continue
        candidates.append({
            "jump": chrono["ChargeOrJump"],
            "vel": chrono["AvgVel"],
            "sd": chrono["SD"],
            "group": grp["GroupIn"],
            "mr": grp["MRIn"],
            "vert": grp["SDVertIn"],
        })

    if not candidates:
        return

    def normalize(values):
        lo, hi = min(values), max(values)
        rng = hi - lo
        return [0.0 if rng == 0 else (v - lo) / rng for v in values]

    # Note: SD uses GROUP as the "Vel" proxy too (D column in SD's analysis block)
    norm_group = normalize([c["group"] for c in candidates])
    norm_sd = normalize([c["sd"] for c in candidates])
    norm_mr = normalize([c["mr"] for c in candidates])
    norm_vert = normalize([c["vert"] for c in candidates])

    composites = [
        w_vel * ng + w_sd * nsd + w_mr * nmr + w_vert * nvert
        for ng, nsd, nmr, nvert in zip(norm_group, norm_sd, norm_mr, norm_vert)
    ]
    composites = [max(0.001, c) for c in composites]

    winner_idx = composites.index(min(composites))
    winner = candidates[winner_idx]

    def best_idx(values):
        return values.index(min(values))

    bests = {
        "Vel ✓": best_idx([c["group"] for c in candidates]),
        "SD ✓": best_idx([c["sd"] for c in candidates]),
        "MR ✓": best_idx([c["mr"] for c in candidates]),
        "Vert ✓": best_idx([c["vert"] for c in candidates]),
    }

    # --- Write Seating Depth row 2 — suggested-jump summary ---
    sd["D2"].value = winner["jump"]
    sd["G2"].value = winner["vel"]
    sd["J2"].value = winner["sd"]
    sd["L2"].value = winner["mr"]
    sd["N2"].value = winner["vert"]

    # --- Write Seating Depth O2 — static "Best in: <tags>" ---
    winner_tags_list = [tag for tag, i in bests.items() if i == winner_idx]
    if len(winner_tags_list) == 4:
        winner_tags = "All metrics ✓"
    elif winner_tags_list:
        winner_tags = " ".join(winner_tags_list)
    else:
        winner_tags = "Composite ✓"
    sd["O2"].value = f"Best in:  {winner_tags}"
    # Styling already applied in _write_static_composite_scores

    # --- Write SD!L30:L37 — composite scores (5 candidates + 3 #N/A) ---
    for i, comp in enumerate(composites):
        sd.cell(row=30 + i, column=12).value = round(comp, 3)
    for r in range(30 + len(composites), 38):
        sd.cell(row=r, column=12).value = "=NA()"

    # --- Write SD!A30:A37 — candidate jump values (since template formulas
    # may reference Load Log which is for charges, not jumps) ---
    for i, c in enumerate(candidates):
        sd.cell(row=30 + i, column=1).value = c["jump"]
    for r in range(30 + len(candidates), 38):
        sd.cell(row=r, column=1).value = "=NA()"

    # --- Write SD!D-G analysis metrics for each candidate ---
    # D=Velocity (group/spread proxy in template), E=SD, F=MR, G=Vert
    for i, c in enumerate(candidates):
        sd.cell(row=30 + i, column=4).value = c["vel"]     # D = Vel
        sd.cell(row=30 + i, column=5).value = c["sd"]      # E = SD
        sd.cell(row=30 + i, column=6).value = c["mr"]      # F = MR
        sd.cell(row=30 + i, column=7).value = c["vert"]    # G = SD-Vert
    for r in range(30 + len(candidates), 38):
        for col in (4, 5, 6, 7):
            sd.cell(row=r, column=col).value = "=NA()"

    # --- Write SD!H-K analysis normalized values (floored at 0.001 for display) ---
    for i in range(len(candidates)):
        r = 30 + i
        sd.cell(row=r, column=8).value = round(max(0.001, norm_group[i]), 3)
        sd.cell(row=r, column=9).value = round(max(0.001, norm_sd[i]), 3)
        sd.cell(row=r, column=10).value = round(max(0.001, norm_mr[i]), 3)
        sd.cell(row=r, column=11).value = round(max(0.001, norm_vert[i]), 3)

    # --- SD!B/C (Low/High) intentionally LEFT EMPTY in test mode ---
    # Those columns are only meaningful in window mode (3-step rolling
    # window of jumps). In test mode they're noise.
    for i in range(len(candidates)):
        sd.cell(row=30 + i, column=2).value = None
        sd.cell(row=30 + i, column=3).value = None

    # --- SD!M30:M34 (Rank column) — 1 (best) to N (worst) by composite ---
    sorted_composites = sorted(enumerate(composites), key=lambda x: x[1])
    rank_by_idx = {idx: rank for rank, (idx, _) in enumerate(sorted_composites, start=1)}
    for i in range(len(candidates)):
        sd.cell(row=30 + i, column=13).value = rank_by_idx[i]  # M column

    # --- SD!N30:N34 (Best in tags) — per-row best-in metrics ---
    tags_per_row_sd = []
    for i in range(len(candidates)):
        tags = [tag for tag, best_i in bests.items() if best_i == i]
        if len(tags) == 4:
            tags_per_row_sd.append("All metrics ✓")
        else:
            tags_per_row_sd.append(" ".join(tags))
    for i, label in enumerate(tags_per_row_sd):
        # Always overwrite — empty string for non-winners. Otherwise the
        # template's per-row formula displays uncomputed in Excel-Mac.
        sd.cell(row=30 + i, column=14).value = label or ""


def _populate_demo_headers(wb):
    """Fill the user-fill cells on Load Log + Seating Depth with demo-shooter
    metadata so the workbook reads as a real, in-progress load development
    instead of half-empty cells."""

    # 6.5 Creedmoor PRS-style demo shooter setup
    from datetime import datetime
    rifle = "Tikka T3X CTR 6.5 CM"
    barrel = '24" 1:8 twist, threaded muzzle'
    shooter = "Demo Shooter"
    optic = "Leupold Mark 5HD 5-25x56"
    chrono = "Garmin Xero C1 Pro"
    cartridge = "6.5 Creedmoor"
    bullet = "Hornady 140gr ELD-M"
    powder = "Hodgdon H4350"
    primer = "CCI BR-2"
    brass = "Hornady"
    cbto = 2.224          # base-to-ogive measurement, in
    oal = 2.875           # overall length, in
    temp_f = 68
    notes = "Sunny, light breeze. 1500 ft DA."

    for sheet_name in ("Load Log", "Seating Depth"):
        if sheet_name not in wb.sheetnames:
            continue
        sht = wb[sheet_name]
        # Row 5 — Rifle / Shooter / Cartridge
        sht["B5"].value = rifle
        sht["G5"].value = shooter
        sht["L5"].value = cartridge
        # Row 6 — Barrel / Optic / Chrono
        sht["B6"].value = barrel
        sht["G6"].value = optic
        sht["L6"].value = chrono
        # Row 7 — scope click-count selection. Demo defaults to 0.1 Mil (most
        # common precision-rifle scope today: NF NX8, Leupold Mark 5HD, etc.)
        sht["G7"].value = "0.1 Mil"
        # Row 9 — Bullet / Powder / Primer / Brass
        sht["B9"].value = bullet
        # G9 has a formula pulling powder from GarminSessions — leave as-is
        sht["L9"].value = primer
        sht["O9"].value = brass
        # Row 10 — CBTO / OAL / Distance
        sht["B10"].value = cbto
        sht["G10"].value = oal
        # L10 is already 100 (set by apply_workbook_repairs)
        # Row 13 — Date / Temp / Notes (LL Date + Notes filled by repairs; SD needs it)
        # Write as datetime object so Excel formats per locale (US: MM/DD/YYYY).
        # apply_workbook_repairs auto-filled LL!B13 with the ISO string from
        # the import — also rewrite it as a datetime so it displays right.
        date_for_sheet = (
            datetime(2026, 4, 26) if sheet_name == "Seating Depth"
            else datetime(2026, 4, 12)
        )
        sht["B13"].value = date_for_sheet
        sht["B13"].number_format = "mm/dd/yyyy"
        sht["G13"].value = temp_f
        # Replace the messy auto-concatenated notes with a clean range note
        sht["L13"].value = notes


if __name__ == "__main__":
    main()
