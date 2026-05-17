#!/usr/bin/env python3
"""Build-time tool — render each demo-tour stop to a high-res PNG.

Path B (pre-rendered demo): the shipped app shows these PNGs in a single
controlled window. Excel is NEVER launched on a customer's machine for
the demo — no popups, no window juggling, no macOS permission prompts.
Run by the developer on a Mac with Microsoft Excel installed; the cost
of needing Excel is paid here once, for pixel-perfect fidelity.

    python3 tools/render_demo_screenshots.py

Writes app/resources/demo_screenshots/:
    01-load-log.png 02-charts.png 03-seating-depth.png
    04-ballistics.png 05-pocket-card.png 06-load-library.png

Pipeline (all built-in macOS tooling, no shipped deps):
  * Excel "save … as PDF" emits the WHOLE workbook as one multi-page
    PDF (one page per visible sheet — verified). We export ONCE, then
    render the page for each demo sheet with Quartz (CoreGraphics).
    The page index per sheet comes from openpyxl's visible-sheet order
    (reliable — Excel's `repeat … visible of sheet` AppleScript throws
    an intermittent -50).
  * Pocket card -> pocket_card HTML -> qlmanage (built-in WebKit) PNG.
  * All outputs autocropped to content with Pillow (dev-only; the
    shipped app never imports PIL).
"""
import glob
import os
import shutil
import subprocess
import sys
import time

import Quartz
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEMO_WB = os.path.join(ROOT, "app", "resources", "Loadscope - Demo Workbook.xlsx")
OUT_DIR = os.path.join(ROOT, "app", "resources", "demo_screenshots")
TMP = "/tmp/loadscope_demo_render"

# (output filename, kind, target). Mirrors TOUR_STOPS in app/demo_tour.py.
STOPS = [
    ("01-load-log.png",      "sheet", "Powder Charge Log"),
    ("02-charts.png",        "sheet", "Charts"),
    ("03-seating-depth.png", "sheet", "Seating Depth Log"),
    ("04-ballistics.png",    "sheet", "Ballistics"),
    ("05-pocket-card.png",   "card",  None),
    ("06-load-library.png",  "sheet", "Load Library"),
]


def _osa(script, timeout=40):
    return subprocess.run(["osascript", "-e", script],
                          capture_output=True, text=True, timeout=timeout)


def _visible_sheet_pages():
    """{sheet_name: 1-based PDF page} from openpyxl's visible-sheet order.
    Each visible sheet is fit-to-one-page, so it's 1 page each."""
    wb = load_workbook(DEMO_WB)
    vis = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]
    return {name: i + 1 for i, name in enumerate(vis)}, len(vis)


def _prepped_workbook():
    """Copy the demo workbook and strip print headers/footers from every
    sheet. Powder Charge Log / Charts / Seating Depth Log carry a footer ("Rifle Load
    Development Log / Page N") that pins the rendered content's bounding
    box to the page bottom — so autocrop keeps a big white band and the
    image displays small. Ballistics / Load Library have NO footer, which
    is exactly why they already look right. Removing header+footer makes
    every demo image crop tight and display large + consistent."""
    from openpyxl import load_workbook
    wb = load_workbook(DEMO_WB)
    for name in wb.sheetnames:
        ws = wb[name]
        for hf in (ws.oddHeader, ws.oddFooter, ws.evenHeader,
                   ws.evenFooter, ws.firstHeader, ws.firstFooter):
            hf.left.text = hf.center.text = hf.right.text = None
        ws.HeaderFooter.differentFirst = False
        ws.HeaderFooter.differentOddEven = False
        # Force every sheet to exactly one print page. Excel happened to
        # render the demo sheets 1-page each; LibreOffice (the headless
        # renderer used when Excel-for-Mac's AppleScript PDF export is
        # broken — confirmed dead in 16.109) honors the xlsx pageSetup,
        # so pin fit-to-1-page to keep the "1 visible sheet = 1 PDF page"
        # mapping in _visible_sheet_pages() bulletproof under either path.
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
    out = os.path.join(TMP, "_demo_noheader.xlsx")
    wb.save(out)
    return out


def _soffice_bin():
    """Locate the LibreOffice headless binary, or None."""
    for p in ("/Applications/LibreOffice.app/Contents/MacOS/soffice",
              shutil.which("soffice") or "",
              shutil.which("libreoffice") or ""):
        if p and os.path.isfile(p):
            return p
    return None


def _export_workbook_pdf_libreoffice(pdf_path):
    """Headless LibreOffice xlsx->PDF. Used because Microsoft broke
    Excel-for-Mac's AppleScript `save … as PDF` in 16.109 (silent no-op /
    parameter error) and its print/save dialogs don't present under
    automation. LibreOffice reads the same xlsx pageSetup, so with
    _prepped_workbook()'s fit-to-1-page it emits one page per visible
    sheet — same contract _visible_sheet_pages() relies on. An isolated
    -env user profile avoids clashing with any running LibreOffice."""
    soffice = _soffice_bin()
    if not soffice:
        raise RuntimeError("LibreOffice not found")
    src = _prepped_workbook()
    outdir = os.path.join(TMP, "_lo_out")
    if os.path.isdir(outdir):
        shutil.rmtree(outdir)
    os.makedirs(outdir)
    profile = os.path.join(TMP, "_lo_profile")
    r = subprocess.run(
        [soffice, "-env:UserInstallation=file://" + profile,
         "--headless", "--norestore", "--convert-to",
         "pdf:calc_pdf_Export", "--outdir", outdir, src],
        capture_output=True, text=True, timeout=180)
    produced = glob.glob(os.path.join(outdir, "*.pdf"))
    if not produced:
        raise RuntimeError(
            f"LibreOffice produced no PDF.\nstdout:{r.stdout}\nstderr:{r.stderr}")
    if os.path.exists(pdf_path):
        os.remove(pdf_path)
    shutil.move(produced[0], pdf_path)


def _export_workbook_pdf(pdf_path):
    src = _prepped_workbook()
    _osa('tell application "Microsoft Excel" to try\n'
         'close every workbook saving no\nend try')
    r = _osa(f'tell application "Microsoft Excel" to open POSIX file "{src}"')
    if r.returncode != 0:
        raise SystemExit(f"Could not open prepped workbook in Excel:\n{r.stderr}")
    time.sleep(2.5)
    if os.path.exists(pdf_path):
        os.remove(pdf_path)
    # Excel emits the WHOLE workbook here (one page per visible sheet) —
    # verified. We rely on that and split by page below.
    #
    # Excel-for-Mac PDF export: the workbook understands `save`, a sheet
    # does NOT (Excel 16.16 raises -1708 "active sheet doesn't understand
    # the save message" for the old `save active sheet ...` form). Use
    # the workbook form (verified working); keep the legacy sheet form as
    # a fallback so this is a strict superset of prior behavior.
    r = _osa('tell application "Microsoft Excel" to save active workbook '
             f'in POSIX file "{pdf_path}" as PDF file format')
    if not os.path.isfile(pdf_path):
        r = _osa(f'tell application "Microsoft Excel" to save active sheet '
                 f'in POSIX file "{pdf_path}" as PDF file format')
    _osa('tell application "Microsoft Excel" to try\n'
         'close every workbook saving no\nend try')
    if not os.path.isfile(pdf_path):
        raise RuntimeError(f"Workbook PDF export failed: {r.stderr}")


def _pdf_page_to_png(pdf_path, page_num, png_path, scale=3.0):
    url = Quartz.CFURLCreateWithFileSystemPathRelativeToBase(
        None, pdf_path, Quartz.kCFURLPOSIXPathStyle, False, None)
    doc = Quartz.CGPDFDocumentCreateWithURL(url)
    if not doc:
        raise RuntimeError(f"Cannot open PDF {pdf_path}")
    npages = Quartz.CGPDFDocumentGetNumberOfPages(doc)
    page = Quartz.CGPDFDocumentGetPage(doc, page_num)
    if not page:
        raise RuntimeError(f"No page {page_num} (PDF has {npages})")
    rect = Quartz.CGPDFPageGetBoxRect(page, Quartz.kCGPDFCropBox)
    w = int(rect.size.width * scale)
    h = int(rect.size.height * scale)
    cs = Quartz.CGColorSpaceCreateDeviceRGB()
    ctx = Quartz.CGBitmapContextCreate(
        None, w, h, 8, 0, cs, Quartz.kCGImageAlphaPremultipliedLast)
    Quartz.CGContextSetRGBFillColor(ctx, 1, 1, 1, 1)
    Quartz.CGContextFillRect(ctx, Quartz.CGRectMake(0, 0, w, h))
    Quartz.CGContextScaleCTM(ctx, scale, scale)
    Quartz.CGContextDrawPDFPage(ctx, page)
    img = Quartz.CGBitmapContextCreateImage(ctx)
    durl = Quartz.CFURLCreateWithFileSystemPathRelativeToBase(
        None, png_path, Quartz.kCFURLPOSIXPathStyle, False, None)
    dest = Quartz.CGImageDestinationCreateWithURL(durl, "public.png", 1, None)
    Quartz.CGImageDestinationAddImage(dest, img, None)
    if not Quartz.CGImageDestinationFinalize(dest):
        raise RuntimeError(f"PNG write failed: {png_path}")


def _render_card_png(png_path):
    sys.path.insert(0, os.path.join(ROOT, "app"))
    import pocket_card
    html = pocket_card.generate_pocket_card(DEMO_WB, open_after=False,
                                            layout="card")
    # Dedicated subdir — must NOT wipe TMP, which holds _workbook.pdf
    # that later sheet stops still need.
    card_tmp = os.path.join(TMP, "_card")
    if os.path.isdir(card_tmp):
        shutil.rmtree(card_tmp)
    os.makedirs(card_tmp)
    subprocess.run(["qlmanage", "-t", "-s", "2000", html, "-o", card_tmp],
                   check=True, capture_output=True, text=True)
    produced = glob.glob(os.path.join(card_tmp, "*.png"))
    if not produced:
        raise RuntimeError("qlmanage produced no PNG for the pocket card")
    shutil.move(produced[0], png_path)


def _autocrop(png_path):
    """Trim surrounding whitespace so each demo image is tight (the PDF
    page is letter-size with content at the top). Pillow is dev-only."""
    try:
        from PIL import Image, ImageChops
    except ImportError:
        print("  (PIL unavailable — skipping autocrop)")
        return
    im = Image.open(png_path).convert("RGB")
    bg = Image.new("RGB", im.size, (255, 255, 255))
    bbox = ImageChops.difference(im, bg).getbbox()
    if bbox:
        pad = 16
        l, t, r, b = bbox
        im.crop((max(0, l - pad), max(0, t - pad),
                 min(im.width, r + pad), min(im.height, b + pad))).save(png_path)


def main():
    if not os.path.isfile(DEMO_WB):
        raise SystemExit(f"Demo workbook missing: {DEMO_WB}")
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(TMP, exist_ok=True)
    pages, nvis = _visible_sheet_pages()
    print(f"Visible sheets ({nvis}): " +
          ", ".join(f"{n}=p{p}" for n, p in pages.items()))
    wb_pdf = os.path.join(TMP, "_workbook.pdf")
    if _soffice_bin():
        print("Exporting whole workbook to PDF via LibreOffice (headless)…")
        _export_workbook_pdf_libreoffice(wb_pdf)
    else:
        print("Exporting whole workbook to PDF via Excel…")
        _export_workbook_pdf(wb_pdf)
    # The contract: one PDF page per visible sheet. Verify it (the demo
    # workbook is small + every sheet is pinned fit-to-1-page) so a
    # renderer pagination change can never silently mis-map sheets.
    _ndoc = Quartz.CGPDFDocumentGetNumberOfPages(
        Quartz.CGPDFDocumentCreateWithURL(
            Quartz.CFURLCreateWithFileSystemPathRelativeToBase(
                None, wb_pdf, Quartz.kCFURLPOSIXPathStyle, False, None)))
    if _ndoc != nvis:
        raise SystemExit(
            f"PDF has {_ndoc} pages but {nvis} visible sheets — page "
            f"mapping unsafe. Check fit-to-page / print areas.")
    for fname, kind, target in STOPS:
        out = os.path.join(OUT_DIR, fname)
        if kind == "sheet":
            if target not in pages:
                raise SystemExit(f"Sheet {target!r} not visible in workbook")
            print(f"-> {fname}  (sheet {target}, page {pages[target]})")
            _pdf_page_to_png(wb_pdf, pages[target], out)
        else:
            print(f"-> {fname}  (pretty pocket card)")
            _render_card_png(out)
        _autocrop(out)
        print(f"   {out}  ({os.path.getsize(out)//1024} KB)")
    print(f"\nDone. {len(STOPS)} images in {OUT_DIR}")


if __name__ == "__main__":
    main()
