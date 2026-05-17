"""Gate-4 predicted-DOPE overlay (pure, no Qt). Workbook is never
written by this module — predictions are a live overlay only."""
import os
import shutil
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import dope_solver as ds  # noqa: E402

DEMO = os.path.join(os.path.dirname(__file__), "..", "app", "resources",
                    "Loadscope - Demo Workbook.xlsx")


@pytest.fixture
def wb(tmp_path):
    dst = tmp_path / "wb.xlsx"
    shutil.copy(DEMO, dst)
    return str(dst)


def test_read_solver_inputs_from_demo(wb):
    inp = ds.read_solver_inputs(wb)
    assert inp["mv"] == 2780
    assert inp["zero"] == 100
    assert inp["sight_h"] == 1.75
    assert inp["unit"] == "mil"
    assert inp["bullet"] == "Hornady 140gr ELD-M"
    # rows 9..18 -> 100..1000 yd
    assert inp["row_ranges"][0] == (9, 100)
    assert inp["row_ranges"][-1] == (18, 1000)


def test_predicted_dope_ok_from_curated_bc(wb):
    # demo bullet "Hornady 140gr ELD-M" is curated g7=0.326
    r = ds.predicted_dope(wb)
    assert r["status"] == ds.OK, r["message"]
    assert r["bc"] == 0.326 and r["bc_model"] == "G7"
    assert "database" in r["bc_source"]
    assert r["unit"] == "mil"
    # every DOPE row predicted
    assert set(r["rows"]) == set(range(9, 19))
    el = [float(r["rows"][rw]["elev"]) for rw in range(9, 19)]
    # ~0 come-up at the 100 yd zero, strictly increasing past it
    assert abs(el[0]) < 0.3
    assert all(b > a for a, b in zip(el[1:], el[2:]))
    # 1000 yd 6.5 CM 140 ELD-M is a few mil — sane band
    assert 7.0 < el[-1] < 12.0
    # wind present and growing
    wd = [float(r["rows"][rw]["wind"]) for rw in range(9, 19)]
    assert wd[-1] > wd[0] >= 0.0


def test_manual_bc_override_wins(wb):
    r = ds.predicted_dope(wb, manual_bc=0.290, manual_model="G7")
    assert r["status"] == ds.OK
    assert r["bc"] == 0.290 and r["bc_source"] == "manual"


def test_no_bc_when_bullet_unknown_and_no_manual(wb):
    w = load_workbook(wb)
    w["Powder Charge Log"]["B9"] = "Acme 999gr Unobtainium"
    w.save(wb)
    r = ds.predicted_dope(wb)
    assert r["status"] == ds.NO_BC
    assert not r["rows"]


def test_no_inputs_when_mv_missing(wb):
    w = load_workbook(wb)
    w["Ballistics"]["K5"] = None
    w.save(wb)
    r = ds.predicted_dope(wb)
    assert r["status"] == ds.NO_INPUTS
    assert not r["rows"]


def test_module_never_writes_workbook(wb):
    before = os.path.getmtime(wb)
    ds.predicted_dope(wb)
    ds.predicted_dope(wb, manual_bc=0.3)
    # predictions are a live overlay — workbook untouched
    assert os.path.getmtime(wb) == before
