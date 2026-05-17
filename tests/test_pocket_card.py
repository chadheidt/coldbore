"""Tests for the Pocket Range Card generator.

Covers:
 - _fmt formatting helper
 - generate_pocket_card raises with no DOPE data
 - generate_pocket_card writes a valid HTML file when data is present
 - URL dispatcher recognizes 'print-pocket-card' action
"""
import os
import sys
import shutil
import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import pocket_card


def test_fmt_blank_glyph_for_none():
    assert pocket_card._fmt(None) == "—"


def test_fmt_blank_glyph_for_empty_string():
    assert pocket_card._fmt("") == "—"
    assert pocket_card._fmt("   ") == "—"


def test_fmt_returns_string_unchanged():
    assert pocket_card._fmt("Tikka") == "Tikka"


def test_fmt_int_value_no_decimal():
    # Float that's actually an int → no decimal
    assert pocket_card._fmt(100.0) == "100"
    assert pocket_card._fmt(7) == "7"


def test_fmt_float_trims_trailing_zeros():
    # Trim trailing zeros, keep meaningful digits
    assert pocket_card._fmt(0.30) == "0.3"
    assert pocket_card._fmt(1.4) == "1.4"


# ---------- Generator end-to-end ----------

@pytest.fixture
def workbook_no_dope(tmp_path):
    """Copy the template — no DOPE data filled in. Used for error-case test."""
    src = os.path.join(
        os.path.dirname(__file__), "..",
        "Rifle Loads Template (do not edit).xltx",
    )
    if not os.path.isfile(src):
        pytest.skip("template not available")
    dst = tmp_path / "no_dope.xlsx"
    shutil.copy(src, dst)
    return str(dst)


@pytest.fixture
def workbook_with_dope(tmp_path):
    """Template with sample DOPE data filled in for the generator test."""
    src = os.path.join(
        os.path.dirname(__file__), "..",
        "Rifle Loads Template (do not edit).xltx",
    )
    if not os.path.isfile(src):
        pytest.skip("template not available")
    dst = tmp_path / "with_dope.xlsx"
    shutil.copy(src, dst)
    from openpyxl import load_workbook
    wb = load_workbook(dst, keep_vba=False)
    wb.template = False
    bal = wb["Ballistics"]
    bal["B5"] = "Tikka T3x 6.5 CM"
    bal["E5"] = "140 ELD-M"
    bal["H5"] = 41.5
    bal["K5"] = 2780
    bal["B6"] = "Vortex Razor"
    bal["E6"] = 100
    bal["H6"] = 1.75
    bal["K6"] = "1:8 RH"
    # Scope click value lives on Powder Charge Log (single source of truth)
    wb["Powder Charge Log"]["G7"] = "1/4 MOA"
    # 3 rows of sample DOPE (enough to satisfy "any data" check)
    sample = [
        (100, 0.0, 0, 0.0, 0, 0.1, 1, 0.3, 1, 0.11),
        (200, 0.3, 3, 1.1, 4, 0.2, 2, 0.6, 2, 0.23),
        (300, 0.8, 8, 2.6, 10, 0.3, 3, 1.0, 4, 0.36),
    ]
    for r, vals in zip(range(9, 12), sample):
        for col, val in zip("ABCDEFGHIJ", vals):
            bal[f"{col}{r}"] = val
    wb.save(dst)
    return str(dst)


def test_generate_raises_without_dope_data(workbook_no_dope):
    with pytest.raises(ValueError, match="no DOPE data"):
        pocket_card.generate_pocket_card(workbook_no_dope, open_after=False)


def test_generate_writes_html_with_data(workbook_with_dope, tmp_path):
    out = pocket_card.generate_pocket_card(workbook_with_dope, open_after=False)
    assert os.path.isfile(out), f"generator should write a file: {out}"
    assert out.endswith(".html")
    with open(out) as f:
        content = f.read()
    # Header content
    assert "Tikka T3x 6.5 CM" in content
    assert "140 ELD-M" in content
    assert "41.5" in content
    # Standard card elements (subtitle uses CSS uppercase, raw text is title-case)
    assert "Pocket Range Card" in content
    assert "Loadscope" in content
    # Trademark mark on the brand
    assert "™" in content
    # Default print layout is now US Letter (prints 1:1 on a normal home
    # printer, unlike a 6x4 page which printers shrink/float) with TWO
    # true-size cards per sheet.
    assert "size: letter" in content
    assert content.count('class="card"') == 2


def test_generate_data_rows_count(workbook_with_dope):
    import re
    # Canonical single card (layout="card" — used for the website /
    # demo-tour screenshot): a 6x4 page with all 10 DOPE rows.
    card = pocket_card.generate_pocket_card(
        workbook_with_dope, open_after=False, layout="card")
    with open(card) as f:
        c1 = f.read()
    assert "size: 6in 4in" in c1
    assert len(re.findall(r"<td class='r'>", c1)) == 10
    # Default "field" layout = US Letter with TWO cards per sheet, so
    # the same 10 rows render twice (20 total).
    fld = pocket_card.generate_pocket_card(
        workbook_with_dope, open_after=False, layout="field")
    with open(fld) as f:
        c2 = f.read()
    assert "size: letter" in c2
    assert len(re.findall(r"<td class='r'>", c2)) == 20


# ---------- URL dispatcher ----------

def test_url_dispatcher_recognizes_print_pocket_card():
    from PyQt5.QtCore import QUrl
    from main import parse_loadscope_action
    assert parse_loadscope_action(QUrl("loadscope://print-pocket-card")) == "print-pocket-card"


# ---------- Gate-4: predicted-DOPE watermark ----------

import shutil as _sh
from openpyxl import load_workbook as _lw

_DEMO = os.path.join(os.path.dirname(__file__), "..", "app", "resources",
                     "Loadscope - Demo Workbook.xlsx")


@pytest.fixture
def fresh_demo(tmp_path):
    """Demo with DOPE cleared = pre-range state -> solver predicts."""
    dst = tmp_path / "fresh.xlsx"
    _sh.copy(_DEMO, dst)
    wb = _lw(dst)
    bal = wb["Ballistics"]
    for r in range(9, 19):
        for c in ("B", "D", "F", "H"):
            bal[f"{c}{r}"] = None
    wb.save(dst)
    return str(dst)


def test_pre_range_card_is_predicted_with_banner(fresh_demo):
    d = pocket_card._gather(fresh_demo)
    assert d["has_predicted"] is True
    assert sum(1 for r in d["dope"] if r.get("predicted")) == 10
    html = pocket_card._build_html(d)
    assert "VERIFY AT THE RANGE" in html          # rendered banner
    assert "tr class='pred'" in html              # predicted rows italic


def test_confirmed_card_has_no_predicted_banner():
    # demo ships fully confirmed -> behaviour unchanged, no banner
    import tempfile
    dst = os.path.join(tempfile.mkdtemp(), "conf.xlsx")
    _sh.copy(_DEMO, dst)
    d = pocket_card._gather(dst)
    assert d["has_predicted"] is False
    assert "VERIFY AT THE RANGE" not in pocket_card._build_html(d)


def test_no_dope_no_bc_raises_with_bc_hint(fresh_demo):
    wb = _lw(fresh_demo)
    wb["Powder Charge Log"]["B9"] = "Acme 999gr Unobtainium"
    wb.save(fresh_demo)
    with pytest.raises(ValueError) as ei:
        pocket_card._gather(fresh_demo)
    assert "ballistic coefficient" in str(ei.value).lower()
