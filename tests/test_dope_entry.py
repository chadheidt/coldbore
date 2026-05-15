"""UI-change Phase 2: pure read/write for the Range Session & DOPE form."""
import os
import shutil
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

from dope_entry_dialog import (  # noqa: E402
    DOPE_ROWS,
    dope_unit,
    read_dope,
    write_dope,
    _unit_cols,
)

DEMO = os.path.join(os.path.dirname(__file__), "..", "app", "resources",
                    "Loadscope - Demo Workbook.xlsx")


@pytest.fixture
def wb(tmp_path):
    dst = tmp_path / "wb.xlsx"
    shutil.copy(DEMO, dst)
    return str(dst)


def test_unit_detection(wb):
    # Demo workbook scope click is "0.1 Mil".
    assert dope_unit(wb) == "mil"
    assert _unit_cols("mil") == ("B", "F")
    assert _unit_cols("moa") == ("D", "H")


def test_read_returns_all_distances(wb):
    d = read_dope(wb)
    assert d["unit"] == "mil"
    assert len(d["rows"]) == len(list(DOPE_ROWS)) == 10
    assert d["rows"][0]["range"] == "100"
    assert d["rows"][-1]["range"] == "1000"
    assert d["date"]  # demo has a session date


def test_write_then_read_roundtrip(wb):
    rows = [{"row": 9, "elev": "0.2", "wind": "0.1", "notes": "calm"}]
    changed = write_dope(wb, "mil", rows)
    assert "B9" in changed and "F9" in changed and "K9" in changed
    back = {r["row"]: r for r in read_dope(wb)["rows"]}
    assert back[9]["elev"] == "0.2"
    assert back[9]["wind"] == "0.1"
    assert back[9]["notes"] == "calm"


def test_numbers_coerced_to_float(wb):
    write_dope(wb, "mil", [{"row": 10, "elev": "1.5", "wind": "0.4",
                            "notes": ""}])
    b = load_workbook(wb)["Ballistics"]
    assert b["B10"].value == 1.5 and isinstance(b["B10"].value, float)
    assert b["F10"].value == 0.4


def test_click_formula_columns_never_touched(wb):
    b0 = load_workbook(wb)["Ballistics"]
    before = {c: b0[f"{c}9"].value for c in ("C", "E", "G", "I")}
    write_dope(wb, "mil", [{"row": 9, "elev": "3.3", "wind": "1.1",
                            "notes": "x"}])
    b1 = load_workbook(wb)["Ballistics"]
    after = {c: b1[f"{c}9"].value for c in ("C", "E", "G", "I")}
    assert before == after  # the auto-click formulas are untouched


def test_blank_clears_cell(wb):
    write_dope(wb, "mil", [{"row": 11, "elev": "2.0", "wind": "",
                            "notes": ""}])
    write_dope(wb, "mil", [{"row": 11, "elev": "", "wind": "",
                            "notes": ""}])
    assert read_dope(wb)["rows"][2]["elev"] == ""  # row 11 is index 2


def test_junk_number_skipped_not_corrupting(wb):
    before = read_dope(wb)["rows"][0]["elev"]  # row 9
    changed = write_dope(wb, "mil", [{"row": 9, "elev": "abc",
                                      "wind": "", "notes": ""}])
    assert "B9" not in changed
    assert read_dope(wb)["rows"][0]["elev"] == before


def test_moa_writes_d_and_h(wb):
    changed = write_dope(wb, "moa", [{"row": 9, "elev": "0.5",
                                      "wind": "1.0", "notes": ""}])
    assert "D9" in changed and "H9" in changed
    assert "B9" not in changed


def test_session_date_parsed_and_written(wb):
    write_dope(wb, "mil", [], date_str="2026-05-20")
    from datetime import datetime
    v = load_workbook(wb)["Load Log"]["B13"].value
    assert isinstance(v, datetime) and v.strftime("%Y-%m-%d") == "2026-05-20"


def test_bad_date_does_not_corrupt(wb):
    before = load_workbook(wb)["Load Log"]["B13"].value
    write_dope(wb, "mil", [], date_str="not a date")
    after = load_workbook(wb)["Load Log"]["B13"].value
    assert before == after
