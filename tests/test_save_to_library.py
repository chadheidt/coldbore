"""Tests for the Save Suggested Load to Library feature.

Covers:
 - _inches_to_moa unit conversion
 - gather_suggested_load reads from cached values correctly
 - save_suggested_load_to_library writes a new row + handles errors
 - URL dispatcher recognizes the new 'save-to-library' action
"""
import os
import sys
import shutil
import tempfile
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import import_data


def test_inches_to_moa_100yd():
    # 1.047 inches at 100yd = 1.0 MOA
    assert import_data._inches_to_moa(1.047, 100) == pytest.approx(1.0, abs=0.01)


def test_inches_to_moa_400yd():
    # 4.188 inches at 400yd = 1.0 MOA
    assert import_data._inches_to_moa(4.188, 400) == pytest.approx(1.0, abs=0.01)


def test_inches_to_moa_handles_missing_inputs():
    assert import_data._inches_to_moa(None, 100) is None
    assert import_data._inches_to_moa(1.0, None) is None
    assert import_data._inches_to_moa(1.0, 0) is None
    assert import_data._inches_to_moa("", 100) is None


def test_inches_to_moa_handles_bad_input():
    assert import_data._inches_to_moa("not a number", 100) is None
    assert import_data._inches_to_moa(1.0, "not a number") is None


# ---------- gather + save round-trip ----------

@pytest.fixture
def workbook_with_winner(tmp_path):
    """Build a minimal but realistic workbook fixture for testing.

    Uses the project's actual template, but writes cached values
    directly so gather_suggested_load can read them without Excel."""
    src = os.path.join(
        os.path.dirname(__file__), "..",
        "Rifle Loads Template (do not edit).xltx",
    )
    if not os.path.isfile(src):
        pytest.skip("template not available in this environment")
    dst = tmp_path / "test_workbook.xlsx"
    shutil.copy(src, dst)

    # Pre-populate the workbook with values that simulate a finished
    # powder ladder + seating depth session.
    from openpyxl import load_workbook
    wb = load_workbook(dst, keep_vba=False)
    wb.template = False
    # Load Log header
    ll = wb["Load Log"]
    ll["B5"].value = "Tikka T3x 6.5 CM"
    ll["B9"].value = "Hornady 140 ELD-M (140 gr)"
    ll["G9"].value = "H4350"
    ll["L9"].value = "CCI BR-2"
    ll["O9"].value = "Lapua"
    ll["B10"].value = 2.165   # CBTO
    ll["L10"].value = 100     # distance yards
    ll["L13"].value = "Calm day, 65°F"
    # Pre-cache values on the Charts winner cells (data_only=True
    # reads the <v> cache directly).
    charts = wb["Charts"]
    charts["B3"].value = 41.5  # winning charge
    # Pre-cache a winning Load Log row at row 16 (charge 41.5, AvgVel etc.)
    ll.cell(16, 2).value = 41.5
    ll.cell(16, 8).value = 2780.0   # H Avg Vel
    ll.cell(16, 10).value = 8.2     # J SD
    ll.cell(16, 11).value = 0.51    # K Group (in)
    ll.cell(16, 14).value = 0.18    # N Mean Radius (in)
    # Seating Depth winner cells
    sd = wb["Seating Depth"]
    sd["D2"].value = 0.014
    sd["G2"].value = 2782.5  # better Avg Vel
    sd["J2"].value = 6.4     # better SD
    sd["L2"].value = 0.142   # MR in MOA
    sd["N2"].value = 0.085   # SD-Vert (MOA)
    wb.save(dst)
    return str(dst)


def test_gather_suggested_load_reads_everything(workbook_with_winner):
    data = import_data.gather_suggested_load(workbook_with_winner)
    assert data["charge"] == 41.5
    assert data["rifle"] == "Tikka T3x 6.5 CM"
    assert data["bullet"] == "Hornady 140 ELD-M (140 gr)"
    assert data["bullet_wt"] == 140.0
    assert data["powder"] == "H4350"
    assert data["primer"] == "CCI BR-2"
    assert data["brass"] == "Lapua"
    assert data["jump"] == 0.014
    # SD winner overrides Load Log values for Avg Vel + SD
    assert data["avg_vel"] == 2782.5
    assert data["sd_fps"] == 6.4
    # Group in MOA: 0.51 inches at 100yd = 0.487 MOA
    assert data["group_moa"] == pytest.approx(0.487, abs=0.01)
    # MR comes from SD analysis (already in MOA)
    assert data["mr_moa"] == pytest.approx(0.142, abs=0.001)


def test_gather_raises_without_winner(tmp_path):
    """If Charts!B3 is empty, gather should refuse."""
    src = os.path.join(
        os.path.dirname(__file__), "..",
        "Rifle Loads Template (do not edit).xltx",
    )
    if not os.path.isfile(src):
        pytest.skip("template not available")
    dst = tmp_path / "no_winner.xlsx"
    shutil.copy(src, dst)
    # Make sure B3 is empty
    from openpyxl import load_workbook
    wb = load_workbook(dst, keep_vba=False)
    wb.template = False
    wb["Charts"]["B3"].value = None
    wb.save(dst)
    with pytest.raises(ValueError, match="No suggested charge"):
        import_data.gather_suggested_load(str(dst))


def test_save_writes_row(workbook_with_winner):
    """End-to-end: gather → save → reload → verify the row landed."""
    row, data = import_data.save_suggested_load_to_library(workbook_with_winner)
    assert row >= 5  # data rows start at row 5
    # Reload and check the Library row
    from openpyxl import load_workbook
    wb = load_workbook(workbook_with_winner, keep_vba=False)
    lib = wb["Load Library"]
    # Col B = Date Added
    assert lib.cell(row, 2).value  # not empty
    # Col D = Rifle
    assert lib.cell(row, 4).value == "Tikka T3x 6.5 CM"
    # Col H = Charge
    assert lib.cell(row, 8).value == 41.5
    # Col L = Jump
    assert lib.cell(row, 12).value == 0.014


def test_save_picks_next_empty_row(workbook_with_winner):
    """Calling save twice should land in two different rows."""
    row1, _ = import_data.save_suggested_load_to_library(workbook_with_winner)
    row2, _ = import_data.save_suggested_load_to_library(workbook_with_winner)
    assert row2 == row1 + 1


def test_save_uses_overridden_data(workbook_with_winner):
    """If data dict is passed explicitly, save should use it verbatim."""
    custom = import_data.gather_suggested_load(workbook_with_winner)
    custom["notes"] = "Manually edited notes"
    custom["bullet_wt"] = 142.0  # user correction
    row, _ = import_data.save_suggested_load_to_library(workbook_with_winner, data=custom)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_with_winner, keep_vba=False)
    lib = wb["Load Library"]
    assert lib.cell(row, 6).value == 142.0  # bullet_wt
    assert lib.cell(row, 17).value == "Manually edited notes"


# ---------- URL dispatcher ----------

def test_url_dispatcher_recognizes_save_action():
    """The parse_loadscope_action helper should return 'save-to-library'
    for our new URL shape."""
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PyQt5.QtCore import QUrl
    from main import parse_loadscope_action
    assert parse_loadscope_action(QUrl("loadscope://save-to-library")) == "save-to-library"
    assert parse_loadscope_action(QUrl("loadscope:///save-to-library")) == "save-to-library"
    assert parse_loadscope_action(QUrl("loadscope://Save-To-Library")) == "save-to-library"
