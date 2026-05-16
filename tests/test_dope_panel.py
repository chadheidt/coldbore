"""Gate-4 Range&DOPE panel: predicted-vs-confirmed behaviour + the
hard guarantee that a live prediction is NEVER written to the workbook.
Headless (offscreen Qt)."""
import os
import shutil
import sys

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

import dope_entry_dialog as ded  # noqa: E402
from dope_entry_dialog import _cell_save_text  # noqa: E402

DEMO = os.path.join(os.path.dirname(__file__), "..", "app", "resources",
                    "Loadscope - Demo Workbook.xlsx")


@pytest.fixture(scope="module")
def qapp():
    from PyQt5.QtWidgets import QApplication
    return QApplication.instance() or QApplication(sys.argv)


@pytest.fixture
def fresh_wb(tmp_path):
    """Demo with the DOPE table cleared = realistic pre-range state
    (rifle/load set, never shot) -> solver overlay should appear."""
    dst = tmp_path / "fresh.xlsx"
    shutil.copy(DEMO, dst)
    wb = load_workbook(dst)
    bal = wb["Ballistics"]
    for r in range(9, 19):
        for col in ("B", "D", "F", "H"):     # mil+moa elev/wind
            bal[f"{col}{r}"] = None
    wb.save(dst)
    return str(dst)


@pytest.fixture
def confirmed_wb(tmp_path):
    dst = tmp_path / "conf.xlsx"
    shutil.copy(DEMO, dst)               # demo ships fully-confirmed
    return str(dst)


# ---- pure save-decision helper ---------------------------------------
def test_cell_save_text_never_persists_a_prediction():
    # confirmed -> the widget text is saved
    assert _cell_save_text(True, "3.6", "") == "3.6"
    # predicted/untouched -> the ORIGINAL workbook value (empty) is
    # saved, so the grey prediction is NOT written
    assert _cell_save_text(False, "3.6", "") == ""
    assert _cell_save_text(False, "9.9", "2.0") == "2.0"


# ---- predicted overlay on a pre-range workbook -----------------------
def test_pre_range_shows_predicted_cells(qapp, fresh_wb):
    p = ded.DopeEntryPanel(fresh_wb, with_save=True)
    assert p._pred["status"] == "ok", p._pred["message"]
    # every elev/wind cell is an unconfirmed prediction with a value
    preds = [(rn, f) for (rn, f), st in p._cells.items()
             if not st["confirmed"] and st["w"].text().strip() != ""]
    assert len(preds) == 20                       # 10 rows x elev+wind
    # predicted cells are styled grey+italic; row tag visible
    st = p._cells[(9, "elev")]
    assert "italic" in st["w"].styleSheet()
    assert p._t.TEXT_TERTIARY in st["w"].styleSheet()
    # offscreen: top-level never shown, so isVisible() is always False;
    # isHidden() reflects the explicit setVisible() we control.
    assert not p._tags[9].isHidden()              # predicted -> shown


def test_editing_a_prediction_confirms_it(qapp, fresh_wb):
    p = ded.DopeEntryPanel(fresh_wb, with_save=True)
    st = p._cells[(13, "elev")]
    assert not st["confirmed"]
    st["w"].setText("4.25")
    p._on_edit(13, "elev")                        # user-edit signal
    assert st["confirmed"]
    assert "normal" in st["w"].styleSheet()
    assert p._t.TEXT_PRIMARY in st["w"].styleSheet()


def test_save_writes_only_confirmed_not_predictions(qapp, fresh_wb):
    p = ded.DopeEntryPanel(fresh_wb, with_save=True)
    # confirm exactly one cell; leave the other 19 as predictions
    p._cells[(13, "elev")]["w"].setText("4.25")
    p._on_edit(13, "elev")
    status, changed = p.save()
    assert status == "ok"
    wb = load_workbook(fresh_wb)
    bal = wb["Ballistics"]
    assert float(bal["B13"].value) == 4.25        # confirmed -> written
    # every other elev/wind cell stayed EMPTY (predictions not written)
    for r in range(9, 19):
        for col in ("B", "F"):
            if (col, r) == ("B", 13):
                continue
            assert bal[f"{col}{r}"].value in (None, ""), (col, r)


def test_confirmed_workbook_shows_no_predictions(qapp, confirmed_wb):
    # demo ships fully-confirmed: nothing should be in predicted state
    p = ded.DopeEntryPanel(confirmed_wb, with_save=True)
    assert all(st["confirmed"] for st in p._cells.values())
    assert all(t.isHidden() for t in p._tags.values())   # none predicted
    # idempotent save (no changes)
    status, changed = p.save()
    assert status == "ok" and changed == []


def test_update_predictions_with_manual_bc(qapp, fresh_wb):
    p = ded.DopeEntryPanel(fresh_wb, with_save=True)
    p._bc_edit.setText("0.290")
    p._model_combo.setCurrentText("G7")
    p._on_update()
    assert p._pred["status"] == "ok"
    assert p._pred["bc"] == 0.290
    # still all predicted (untouched) and not written on save
    s, _ = p.save()
    assert s == "ok"
    bal = load_workbook(fresh_wb)["Ballistics"]
    assert all(bal[f"B{r}"].value in (None, "") for r in range(9, 19))
