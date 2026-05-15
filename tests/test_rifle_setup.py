"""UI-change Phase 1: pure read/write round-trip for the Rifle & Setup form.

No Qt needed — RIFLE_SETUP_FIELDS + read/write are pure openpyxl.
"""
import os
import shutil
import sys

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

from rifle_setup_dialog import (  # noqa: E402
    RIFLE_SETUP_FIELDS,
    read_rifle_setup,
    write_rifle_setup,
)

DEMO = os.path.join(os.path.dirname(__file__), "..", "app", "resources",
                    "Loadscope - Demo Workbook.xlsx")


@pytest.fixture
def wb(tmp_path):
    dst = tmp_path / "wb.xlsx"
    shutil.copy(DEMO, dst)
    return str(dst)


def test_field_map_is_sane():
    cells = [c for _s, _l, c, _k in RIFLE_SETUP_FIELDS]
    assert len(cells) == len(set(cells)), "duplicate target cell"
    # Powder G9 (formula) and Date B13 (typed) are intentionally excluded.
    assert "G9" not in cells
    assert "B13" not in cells
    valid_kinds = {"text", "number", "turret", "bullet", "primer",
                   "brass", "chrono", "shooter", "distance"}
    assert all(k in valid_kinds for *_x, k in RIFLE_SETUP_FIELDS)


def test_read_returns_every_field(wb):
    data = read_rifle_setup(wb)
    for _s, _l, cell, _k in RIFLE_SETUP_FIELDS:
        assert cell in data
    # Demo workbook has a real rifle name in B5.
    assert data["B5"].strip() != ""


def test_write_then_read_roundtrip(wb):
    changed = write_rifle_setup(wb, {"B5": "Test Rifle 6 Dasher",
                                     "L13": "Cold, 20F, 4000ft DA"})
    assert set(changed) == {"B5", "L13"}
    back = read_rifle_setup(wb)
    assert back["B5"] == "Test Rifle 6 Dasher"
    assert back["L13"] == "Cold, 20F, 4000ft DA"


def test_number_field_coerced_to_float(wb):
    write_rifle_setup(wb, {"B10": "2.250"})
    from openpyxl import load_workbook
    v = load_workbook(wb)["Load Log"]["B10"].value
    assert v == 2.25 and isinstance(v, float)


def test_unparseable_number_is_skipped(wb):
    before = read_rifle_setup(wb)["B10"]
    changed = write_rifle_setup(wb, {"B10": "not a number"})
    assert "B10" not in changed
    assert read_rifle_setup(wb)["B10"] == before


def test_no_change_is_noop(wb):
    current = read_rifle_setup(wb)
    changed = write_rifle_setup(wb, {"B5": current["B5"]})
    assert changed == []


def test_formula_cell_g9_untouched(wb):
    # Powder G9 is a formula and not in the field map; writing the form
    # must never disturb it.
    from openpyxl import load_workbook
    before = load_workbook(wb)["Load Log"]["G9"].value
    write_rifle_setup(wb, {"B5": "Another Rifle"})
    after = load_workbook(wb)["Load Log"]["G9"].value
    assert before == after
