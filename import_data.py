#!/usr/bin/env python3
"""
Rifle Loads — Import Data

Walks every parser's import folder (Garmin Imports/, BallisticX Imports/, etc.),
parses each CSV with its registered parser, and writes the data into the
active working .xlsx in this project folder.

Run via: double-click "Import Rifle Data.command" on your Desktop.

CSV parsing is delegated to the parser registry in app/parsers/. Adding support
for a new chronograph or target-analysis app means dropping a new file in
app/parsers/ — see app/parsers/__init__.py for the contract.
"""

import os
import sys
import shutil
import subprocess
from datetime import datetime
from openpyxl import load_workbook

# Make the parser registry importable. import_data.py runs from the project
# root (CLI flow) and from inside the bundled .app, so add app/ to sys.path.
_HERE = os.path.dirname(os.path.abspath(__file__))
_APP_DIR = os.path.join(_HERE, "app")
if os.path.isdir(_APP_DIR) and _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

from parsers import ALL_PARSERS, chronograph_parsers, group_parsers


def _default_project_dir():
    """Resolve the project folder for CLI-only callers. The GUI always passes
    project_dir= explicitly, so this is only used when import_data.py is run
    directly from the command line."""
    try:
        # Lazy import — config lives under app/, which we've already added to sys.path
        import config as _app_config
        p = _app_config.get_project_folder()
        if p and os.path.isdir(p):
            return p
    except Exception:
        pass
    # Last-resort fallback: project folder is the directory this script lives in
    return _HERE


PROJECT = _default_project_dir()
GARMIN_DIR = os.path.join(PROJECT, "Garmin Imports")
BX_DIR = os.path.join(PROJECT, "BallisticX Imports")


# Backwards-compatible re-exports for any old callers that imported these
# helpers directly from import_data (e.g., older test scripts).
from parsers._common import parse_label, extract_inches, extract_signed  # noqa: F401


# Sanity ranges for incoming data — we WARN but don't reject if values fall
# outside these. They're meant to catch obvious garbage (parsing failures,
# Garmin reading a phantom shot, BallisticX CSV from the wrong source).
# Tuned wide enough to accept anything plausible from a rifle cartridge.
SANITY_VEL_FPS = (500, 5000)         # rifle muzzle velocity in fps
SANITY_SD_FPS = (0, 60)              # standard deviation in fps
SANITY_ES_FPS = (0, 200)             # extreme spread in fps
SANITY_GROUP_IN = (0, 10)            # group size in inches at any reasonable distance
SANITY_VERTICAL_IN = (0, 10)         # vertical dispersion in inches
SANITY_MR_IN = (0, 10)               # mean radius in inches
SANITY_CHARGE_GR = (5, 200)          # powder charge weight (grains) — covers everything from .22 to magnum
SANITY_BULLET_WT_GR = (10, 800)      # bullet weight (grains)


def _validate_chronograph_record(d, source_filename, warnings):
    """Sanity-check a chronograph record. Append human-readable warnings to
    the warnings list rather than raising — the data still gets imported,
    but the user sees a flagged note in the activity log."""
    name = source_filename
    tag = d.get("Tag", "?")

    avg_vel = d.get("AvgVel")
    if avg_vel is not None and not (SANITY_VEL_FPS[0] <= avg_vel <= SANITY_VEL_FPS[1]):
        warnings.append(f"  ⚠ {name} ({tag}): AvgVel {avg_vel} fps is outside "
                        f"plausible range ({SANITY_VEL_FPS[0]}-{SANITY_VEL_FPS[1]})")

    sd = d.get("SD")
    if sd is not None and not (SANITY_SD_FPS[0] <= sd <= SANITY_SD_FPS[1]):
        warnings.append(f"  ⚠ {name} ({tag}): SD {sd} fps is outside "
                        f"plausible range ({SANITY_SD_FPS[0]}-{SANITY_SD_FPS[1]})")

    es = d.get("ES")
    if es is not None and not (SANITY_ES_FPS[0] <= es <= SANITY_ES_FPS[1]):
        warnings.append(f"  ⚠ {name} ({tag}): ES {es} fps is outside "
                        f"plausible range ({SANITY_ES_FPS[0]}-{SANITY_ES_FPS[1]})")

    bw = d.get("BulletWt")
    if bw is not None and not (SANITY_BULLET_WT_GR[0] <= bw <= SANITY_BULLET_WT_GR[1]):
        warnings.append(f"  ⚠ {name} ({tag}): BulletWt {bw} gr is outside "
                        f"plausible range ({SANITY_BULLET_WT_GR[0]}-{SANITY_BULLET_WT_GR[1]})")

    charge = d.get("ChargeOrJump")
    # ChargeOrJump can be either a powder charge (5-200 grains) or a jump
    # (typically 0-0.2 inches). Skip range check — too ambiguous to flag.

    # Per-shot velocities
    shots = d.get("Shots") or []
    for i, shot in enumerate(shots):
        if shot is None:
            continue
        if not (SANITY_VEL_FPS[0] <= shot <= SANITY_VEL_FPS[1]):
            warnings.append(f"  ⚠ {name} ({tag}): Shot {i+1} velocity {shot} fps "
                            f"is outside plausible range ({SANITY_VEL_FPS[0]}-{SANITY_VEL_FPS[1]})")


def _validate_group_record(g, source_filename, warnings):
    """Sanity-check a target-group record."""
    name = source_filename
    tag = g.get("Tag", "?")

    for field, label, rng in [
        ("GroupIn", "Group", SANITY_GROUP_IN),
        ("HeightIn", "Vertical", SANITY_VERTICAL_IN),
        ("MRIn", "Mean radius", SANITY_MR_IN),
    ]:
        v = g.get(field)
        if v is not None and not (rng[0] <= v <= rng[1]):
            warnings.append(f"  ⚠ {name} ({tag}): {label} {v}\" is outside "
                            f"plausible range ({rng[0]}-{rng[1]}\")")


def _check_duplicate_tags(records, label):
    """Return human-readable warnings for any tag that appears twice in records."""
    warnings = []
    by_tag = {}
    for r in records:
        tag = r.get("Tag", "")
        if not tag:
            continue
        by_tag.setdefault(tag, []).append(r)
    for tag, group in by_tag.items():
        if len(group) > 1:
            warnings.append(
                f"  ⚠ Duplicate tag '{tag}' in {label}: {len(group)} records share "
                f"this tag. The workbook will only show the last one for that row."
            )
    return warnings


def _read_backup_keep_setting(default=5):
    """Read the backup retention count from the GUI's config file. Falls back
    to `default` if config or app/ folder isn't available (e.g. running as the
    pure CLI flow without the GUI installed). 0 = disable backups."""
    try:
        import config as app_config  # only imports if app/ is on sys.path
        cfg = app_config.load_config()
        return int(cfg.get("backup_keep_count", default))
    except Exception:
        return default


def _rotate_workbook_backups(workbook_path, keep=5):
    """Keep a small rotating set of timestamped backups of the workbook in a
    .backups/ subfolder next to it. Called before each import so the user
    can recover from a botched run.

    Backups are named like:  <Workbook> 2026-05-07 14-32-08.xlsx
    Oldest are deleted once the count exceeds `keep`.
    """
    workbook_path = os.path.abspath(workbook_path)
    project_dir = os.path.dirname(workbook_path)
    backups_dir = os.path.join(project_dir, ".backups")
    os.makedirs(backups_dir, exist_ok=True)

    base = os.path.splitext(os.path.basename(workbook_path))[0]
    ext = os.path.splitext(workbook_path)[1]
    stamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    dest = os.path.join(backups_dir, f"{base} {stamp}{ext}")

    try:
        shutil.copy2(workbook_path, dest)
        print(f"  Backed up to: .backups/{os.path.basename(dest)}")
    except Exception as e:
        print(f"  WARNING: couldn't create backup: {e}")
        return

    # Rotate — delete oldest backups for THIS workbook beyond `keep`
    same_base_backups = sorted(
        [f for f in os.listdir(backups_dir)
         if f.startswith(base + " ") and f.endswith(ext)],
        reverse=True,  # newest first
    )
    for old in same_base_backups[keep:]:
        try:
            os.remove(os.path.join(backups_dir, old))
        except OSError:
            pass


def list_workbooks(project_dir=None):
    """Return a list of working .xlsx paths in the project folder, most-recent-first.

    Excludes .xltx template, Excel lock files (~$), and hidden files. This is the
    library function — both the CLI find_workbook() and the GUI use it.
    """
    if project_dir is None:
        project_dir = PROJECT
    # Match the canonical template filename only — we used to strip any .xlsx
    # whose name contained "template", which silently hid user files like
    # "Template Test 6.5CM.xlsx" or "MyTemplate.xlsx".
    canonical_template = "rifle loads template (do not edit).xlsx"
    candidates = []
    for f in os.listdir(project_dir):
        full = os.path.join(project_dir, f)
        if not os.path.isfile(full):
            continue
        if f.startswith("~$") or f.startswith("."):
            continue
        if not f.lower().endswith(".xlsx"):
            continue
        if f.lower() == canonical_template:
            continue
        candidates.append(full)
    candidates.sort(key=os.path.getmtime, reverse=True)
    return candidates


def find_workbook():
    """Find the working .xlsx in the project folder (CLI flow).

    Behavior:
      - 0 working .xlsx files  →  return None (caller prints setup error)
      - 1 working .xlsx file   →  auto-pick it, no prompt
      - 2+ working .xlsx files →  print a numbered list and ask the user
                                  which one to import into
    """
    candidates = list_workbooks()
    if not candidates:
        return None

    if len(candidates) == 1:
        return candidates[0]

    # Multiple workbooks — prompt the user
    print("\n" + "=" * 60)
    print("Multiple working workbooks found in the project folder.")
    print("=" * 60)
    print("\nWhich workbook do you want to import into?\n")
    for i, path in enumerate(candidates, start=1):
        name = os.path.basename(path)
        mtime = os.path.getmtime(path)
        import datetime
        when = datetime.datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
        marker = "  (most recently edited)" if i == 1 else ""
        print(f"  {i}.  {name}    [{when}]{marker}")
    print()

    while True:
        try:
            choice = input(f"Type a number 1-{len(candidates)} (or press Enter for 1): ").strip()
        except EOFError:
            choice = ""
        if choice == "":
            return candidates[0]
        try:
            idx = int(choice)
            if 1 <= idx <= len(candidates):
                return candidates[idx - 1]
        except ValueError:
            pass
        print(f"  Invalid choice. Type a number from 1 to {len(candidates)}.")


def migrate_schema_to_10_shot(wb):
    """One-time, idempotent migration from the pre-v0.13.3 7-shot schema
    to the v0.13.3 10-shot schema. MUST run BEFORE write_chronograph_records
    on any workbook that might be on the legacy schema — otherwise the
    writer's new column targets collide with legacy data positions.

    Detection signals:
      - GarminSessions!L1 == 'AvgVel'  (legacy hidden-sheet header)
      - 'Garmin Xero Import'!L5 == 'Avg Vel (fps)'  (legacy visible header)

    Returns a list of human-readable fix strings (empty if no migration
    was needed).
    """
    fixes = []

    # --- GarminSessions hidden sheet: shift legacy summary cols L..R → O..U
    if "GarminSessions" in wb.sheetnames:
        gs = wb["GarminSessions"]
        if (gs.cell(1, 12).value or "") == "AvgVel":
            last_row = gs.max_row
            for r in range(2, last_row + 1):
                # Walk right-to-left so dst writes don't clobber unread src
                for src_col, dst_col in [
                    (18, 21), (17, 20), (16, 19), (15, 18),
                    (14, 17), (13, 16), (12, 15),
                ]:
                    gs.cell(r, dst_col).value = gs.cell(r, src_col).value
                    gs.cell(r, src_col).value = None
            for col, header in [
                (12, "Shot8"), (13, "Shot9"), (14, "Shot10"),
                (15, "AvgVel"), (16, "SD"), (17, "ES"),
                (18, "BulletWt"), (19, "AvgKE"),
                (20, "SessionTitle"), (21, "SessionNote"),
            ]:
                gs.cell(1, col).value = header
            fixes.append(
                "GarminSessions: migrated 7-shot → 10-shot schema "
                f"({last_row - 1} data row(s) shifted)"
            )

    # --- Garmin Xero Import visible sheet: shift L..Q → O..T, add new shot cells
    if "Garmin Xero Import" in wb.sheetnames:
        gxi = wb["Garmin Xero Import"]
        if (gxi.cell(5, 12).value or "") == "Avg Vel (fps)":
            # Map: old visible col → new visible col, GS letter the new
            # formula must reference (post-GS-migration positions).
            #   old L (AvgVel, GS!$L) → new O (AvgVel, GS!$O)
            #   old M (SD,     GS!$M) → new P (SD,     GS!$P)
            #   old N (ES,     GS!$N) → new Q (ES,     GS!$Q)
            #   old O (BWt,    GS!$O) → new R (BWt,    GS!$R)
            #   old P (KE,     GS!$P) → new S (KE,     GS!$S)
            #   old Q (Notes,  GS!$R) → new T (Notes,  GS!$U)
            shift_map = [
                (12, 15, "L", "O"),
                (13, 16, "M", "P"),
                (14, 17, "N", "Q"),
                (15, 18, "O", "R"),
                (16, 19, "P", "S"),
                (17, 20, "R", "U"),  # Notes: GS source jumps R→U
            ]
            for row_n in range(6, 27):
                if row_n == 16:
                    continue  # section header
                e_style = gxi.cell(row_n, 5).style
                # Move summary cells right-to-left
                for src_col, dst_col, old_gs, new_gs in shift_map[::-1]:
                    src = gxi.cell(row_n, src_col)
                    dst = gxi.cell(row_n, dst_col)
                    val = src.value
                    if isinstance(val, str) and val.startswith("="):
                        val = val.replace(
                            f"GarminSessions!${old_gs}$2:${old_gs}$200",
                            f"GarminSessions!${new_gs}$2:${new_gs}$200",
                        )
                    dst.value = val
                    dst.style = src.style
                # Plant new Shot 8/9/10 formulas at L/M/N with shot styling.
                # ISNUMBER guard so empty source cells display as "" not 0.
                for col, gs_letter in [(12, "L"), (13, "M"), (14, "N")]:
                    cell = gxi.cell(row_n, col)
                    inner = (
                        f"LOOKUP(2,1/(GarminSessions!$A$2:$A$200=$A{row_n}),"
                        f"GarminSessions!${gs_letter}$2:${gs_letter}$200)"
                    )
                    cell.value = f'=IFERROR(IF({inner}>0,{inner},""),"")'
                    cell.style = e_style
                # Update D's COUNT range
                d_cell = gxi.cell(row_n, 4)
                if isinstance(d_cell.value, str) and d_cell.value.startswith("="):
                    d_cell.value = d_cell.value.replace(
                        f"E{row_n}:K{row_n}", f"E{row_n}:N{row_n}"
                    )
                # Update C's SessionTitle ref (GS!$Q → GS!$T)
                c_cell = gxi.cell(row_n, 3)
                if isinstance(c_cell.value, str) and "GarminSessions!$Q" in c_cell.value:
                    c_cell.value = c_cell.value.replace(
                        "GarminSessions!$Q$2:$Q$200",
                        "GarminSessions!$T$2:$T$200",
                    )
            # Row 5 headers
            for col, label in [
                (12, "Shot 8 (fps)"), (13, "Shot 9 (fps)"), (14, "Shot 10 (fps)"),
                (15, "Avg Vel (fps)"), (16, "SD (fps)"), (17, "ES (fps)"),
                (18, "Bullet Wt (gr)"), (19, "Avg KE (ft·lb)"), (20, "Notes"),
            ]:
                gxi.cell(5, col).value = label
            # Row 2 help text
            r2 = gxi.cell(2, 1)
            if isinstance(r2.value, str) and "Shot 1–5 cells (E:I)" in r2.value:
                r2.value = r2.value.replace(
                    "Shot 1–5 cells (E:I)", "Shot 1–10 cells (E:N)"
                )
            fixes.append(
                "Garmin Xero Import: migrated 7-shot → 10-shot layout"
            )

    # --- Load Log + Seating Depth: cross-sheet refs to 'Garmin Xero Import'
    # The candidate tables on these sheets pull Avg Vel / ES / SD values
    # from 'Garmin Xero Import' cols L/M/N (pre-migration AvgVel/SD/ES).
    # After the visible-sheet migration above, those values now live in
    # cols O/P/Q. Remap each formula. Only fires if the legacy references
    # are still in place.
    for sheet_name, rows in (("Load Log", range(16, 26)), ("Seating Depth", range(16, 27))):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_fixes = 0
        for ll_row in rows:
            for ll_col in range(2, 17):  # B..P
                cell = ws.cell(ll_row, ll_col)
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=") and "Garmin Xero Import" in v):
                    continue
                new_v = v
                # The GXI row referenced is independent of this cell's row;
                # the legacy refs are like 'Garmin Xero Import'!L6, M6, N6.
                # Apply substitutions: !L<n> → !O<n>, !M<n> → !P<n>, !N<n> → !Q<n>.
                # Use word-boundary-style match so we don't double-shift.
                import re as _re
                new_v = _re.sub(r"('Garmin Xero Import'!\$?)L(\$?\d+)", r"\1O\2", new_v)
                new_v = _re.sub(r"('Garmin Xero Import'!\$?)M(\$?\d+)", r"\1P\2", new_v)
                new_v = _re.sub(r"('Garmin Xero Import'!\$?)N(\$?\d+)", r"\1Q\2", new_v)
                if new_v != v:
                    cell.value = new_v
                    sheet_fixes += 1
        if sheet_fixes:
            fixes.append(
                f"{sheet_name}: remapped {sheet_fixes} cross-sheet ref(s) "
                "to new Garmin Xero Import layout"
            )

    if fixes:
        print("  Schema migration applied:")
        for f in fixes:
            print(f"    - {f}")
    return fixes


def write_chronograph_records(wb, records):
    """Write all chronograph session records to the GarminSessions hidden sheet.

    Sheet name is historical (kept as-is to avoid breaking existing workbooks),
    but it now holds chronograph data from any source — Garmin, LabRadar,
    MagnetoSpeed, etc. The Source field on each record indicates which parser
    produced it.
    """
    if "GarminSessions" not in wb.sheetnames:
        print("  ERROR: workbook has no GarminSessions sheet")
        return 0
    sht = wb["GarminSessions"]
    # Clear rows 2+
    for row in sht.iter_rows(min_row=2, max_row=sht.max_row):
        for cell in row:
            cell.value = None

    # v0.13.3 schema:
    #   A=Tag, B=ChargeOrJump, C=Powder, D=Date,
    #   E-N = Shot 1..10,
    #   O=AvgVel, P=SD, Q=ES, R=BulletWt, S=AvgKE,
    #   T=SessionTitle, U=SessionNote.
    out_row = 2
    for d in records:
        sht.cell(row=out_row, column=1).value = d.get("Tag", "")
        sht.cell(row=out_row, column=2).value = d.get("ChargeOrJump")
        sht.cell(row=out_row, column=3).value = d.get("Powder", "")
        sht.cell(row=out_row, column=4).value = d.get("Date", "")
        for i, shot in enumerate((d.get("Shots") or [])[:10]):
            sht.cell(row=out_row, column=5 + i).value = shot
        sht.cell(row=out_row, column=15).value = d.get("AvgVel")
        sht.cell(row=out_row, column=16).value = d.get("SD")
        sht.cell(row=out_row, column=17).value = d.get("ES")
        sht.cell(row=out_row, column=18).value = d.get("BulletWt")
        sht.cell(row=out_row, column=19).value = d.get("AvgKE")
        sht.cell(row=out_row, column=20).value = d.get("SessionTitle", "")
        sht.cell(row=out_row, column=21).value = d.get("SessionNote", "")
        out_row += 1
    return out_row - 2


# Backwards-compat alias for any callers still using the old name
write_garmin = write_chronograph_records


def inherit_rifle_setup(new_wb, project_dir, exclude_path=None):
    """Pre-fill the new workbook's Rifle/Shooter/Components fields from the
    user's most recent prior workbook in the project folder (or the
    Completed Loads subfolder). Users typically run multiple cycles on the
    same rifle and same components — copying these fields saves them from
    re-entering 9 metadata cells every time.

    Only fills cells that are currently empty. Never overwrites.
    `exclude_path` is the new workbook being created — don't read from itself.
    """
    import os
    from openpyxl import load_workbook
    if "Load Log" not in new_wb.sheetnames:
        return []

    # Find candidate prior workbooks: .xlsx files in the project folder + Completed Loads
    candidates = []
    for d in (project_dir, os.path.join(project_dir, "Completed Loads")):
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            if not fn.lower().endswith(".xlsx") or fn.startswith("~"):
                continue
            full = os.path.join(d, fn)
            if exclude_path and os.path.abspath(full) == os.path.abspath(exclude_path):
                continue
            try:
                mtime = os.path.getmtime(full)
            except OSError:
                continue
            candidates.append((mtime, full))
    if not candidates:
        return []
    candidates.sort(reverse=True)  # most recent first

    # Fields to inherit, by (row, col) coordinate on Load Log
    # Layout: A=label, B=value, F=label, G=value, K=label, L=value (and below)
    FIELDS = [
        (5, 2,  "Rifle"),
        (5, 7,  "Shooter"),
        (5, 12, "Cartridge"),
        (6, 2,  "Barrel"),
        (6, 7,  "Optic"),
        (6, 12, "Chrono"),
        (9, 2,  "Bullet"),
        (9, 6,  "Powder"),
        (9, 9,  "Primer"),
        (9, 12, "Brass"),
    ]
    target_ws = new_wb["Load Log"]
    inherited = []
    for _mtime, prior_path in candidates:
        try:
            prior_wb = load_workbook(prior_path, data_only=True, keep_vba=False)
        except Exception:
            continue
        if "Load Log" not in prior_wb.sheetnames:
            continue
        prior_ws = prior_wb["Load Log"]
        any_inherited = False
        for r, c, label in FIELDS:
            tgt = target_ws.cell(r, c)
            if tgt.value not in (None, ""):
                continue  # don't overwrite
            src_val = prior_ws.cell(r, c).value
            # Skip formulas, empty, and openpyxl objects (formula placeholders)
            if src_val in (None, ""):
                continue
            if isinstance(src_val, str) and src_val.startswith("="):
                continue
            tgt.value = src_val
            inherited.append(label)
            any_inherited = True
        if any_inherited:
            # Found a prior workbook with usable data — stop scanning
            break
    return inherited


def stamp_load_name(wb, load_name):
    """Write the user-chosen load name into the top header cell on each
    user-facing sheet so the user knows which load they're viewing at a glance.

    Called once when a new workbook is created (first-load prompt or new-cycle
    dialog). Idempotent — calling again just overwrites the title cells.
    """
    if not load_name:
        return
    stamps = {
        "Load Log": load_name,
        "Seating Depth": f"{load_name}  —  Seating Depth",
        "Charts": f"{load_name}  —  Suggested Best Load",
        # Tab was renamed to "Start Here" in v0.14 — keep old name as
        # fallback for any in-flight workbooks.
        "Start Here": f"Start Here  —  {load_name}",
        "After Range Day": f"After Range Day  —  {load_name}",
    }
    for sheet, value in stamps.items():
        if sheet in wb.sheetnames:
            try:
                wb[sheet]["A1"] = value
            except Exception:
                # Cell may be part of a merged range that needs special handling.
                # The top-left cell of a merge is writable, so this should
                # normally succeed; swallow errors so a stamp failure doesn't
                # block workbook creation.
                pass


def apply_workbook_repairs(wb, group_records, chronograph_records=None):
    """Patch a workbook's template-level bugs and auto-populate user-entry
    fields that we can infer from the import data.

    Idempotent — only fills cells/attributes that are currently missing or
    set to the wrong defaults. Existing user input is never overwritten.

    Fixes:
      - Load Log + Charts "Accuracy Node Finder" LineChart: display_blanks
        was set to 'zero' in early templates, which made the chart line
        drop to a flat zero baseline after the last shot. Changes it to
        'gap' so empty cells are skipped.
      - Load Log!L10 + Seating Depth!L10 (Dist (yd)): if the user hasn't
        filled it in yet AND the imported BallisticX records agree on a
        single distance, set it. The MR display in the suggested-charge
        red row divides by this distance to convert inches to MOA --
        without it, MR + Best-in render blank.
    """
    fixes = []

    # Self-heal the A3 "Click here to jump to Charts" hyperlink on Load Log
    # and Seating Depth. Pre-v0.13.3 templates pointed it at an intra-sheet
    # row (A27) that was just a header label — no charts under it. We
    # retarget the link to the Charts sheet where the actual charts live.
    from openpyxl.worksheet.hyperlink import Hyperlink
    JUMP_DISPLAY = (
        "↓  Click here to jump to the Charts sheet "
        "(Suggested Best Load)  ↓"
    )
    for sheet_name in ("Load Log", "Seating Depth"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        a3 = ws["A3"]
        hl = a3.hyperlink
        # Treat as broken if missing, if pointing at any !A27 target, or if
        # target/display still mentions A27 / 'this sheet'.
        broken = (
            hl is None
            or (hl.location or "").endswith("!A27")
            or (hl.target or "").endswith("!A27")
            or (hl.target or "").endswith("'!A27")
        )
        if broken:
            a3.hyperlink = Hyperlink(
                ref="A3", location="Charts!A1", display=JUMP_DISPLAY
            )
            a3.value = JUMP_DISPLAY
            fixes.append(f"{sheet_name}!A3: retargeted jump link to Charts sheet")

    for sheet_name in ("Load Log", "Charts"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for chart in ws._charts:
            try:
                title = chart.title.tx.rich.paragraphs[0].r[0].t if chart.title else ""
            except Exception:
                title = ""
            if "Accuracy Node" not in title:
                continue
            chart_changed = False
            if getattr(chart, "display_blanks", None) == "zero":
                chart.display_blanks = "gap"
                chart_changed = True
            # Series 1 (Group MOA) + Series 2 (SD) had white markers that
            # sat ON the x-axis line at the bottom of the chart, since both
            # series plot at very small y-values compared to Velocity. The
            # markers visually broke up the gray axis line. Hide them --
            # the lines themselves stay visible (when colored).
            try:
                from openpyxl.chart.marker import Marker
                for series_idx in (1, 2):
                    if series_idx < len(chart.series):
                        s = chart.series[series_idx]
                        if s.marker is None or (s.marker.symbol or "circle") != "none":
                            s.marker = Marker(symbol="none")
                            chart_changed = True
            except Exception:
                pass
            if chart_changed:
                fixes.append(f"{sheet_name}/Accuracy Node Finder: cleaned up empty-cell handling + axis-obscuring markers")

    distances = [g.get("Distance") for g in (group_records or []) if g.get("Distance")]
    common_distance = None
    if distances:
        from collections import Counter
        common_distance = Counter(distances).most_common(1)[0][0]
    if common_distance is not None:
        for sheet_name in ("Load Log", "Seating Depth"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            l10 = ws.cell(10, 12)
            if l10.value is None or l10.value == "":
                l10.value = common_distance
                fixes.append(f"{sheet_name}!L10: set Dist (yd) = {common_distance}")

    # Chart-source cells in Load Log H/J/K (Avg Vel, SD, Group) — these
    # feed the Accuracy Node Finder LineChart. Excel charts plot empty
    # strings as zero (which is why the line dropped to a zero baseline
    # after the last shot). Returning NA() instead makes the chart skip
    # the cell. Custom number format hides #N/A from the user view.
    if "Load Log" in wb.sheetnames:
        ws = wb["Load Log"]
        chart_src_changes = 0
        for r in range(16, 26):
            for col in (8, 10, 11):  # H, J, K
                cell = ws.cell(r, col)
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and ',""' in v and "NA()" not in v:
                    cell.value = v.replace(',""', ',NA()')
                    chart_src_changes += 1
                if '[=NA()]' not in (cell.number_format or ""):
                    cell.number_format = '[=NA()]"";General'
        if chart_src_changes:
            fixes.append(f"Load Log H/J/K16:25: rewired {chart_src_changes} chart-source formula(s) to skip empties")

    # Load Log!M2 "Best in:" indicator. v0.13.2 templates had an inline
    # cross-sheet concat formula that Excel for Mac failed to render
    # (the cell appeared blank). v0.13.3 mirrors the working Seating
    # Depth pattern: pre-compute a per-row Best-in string on Charts!R18:R25,
    # then Load Log!M2 does a simple INDEX/MATCH against the winning row.
    if "Charts" in wb.sheetnames and "Load Log" in wb.sheetnames:
        charts = wb["Charts"]
        loadlog = wb["Load Log"]
        # AGGREGATE(5,6,...) = MIN ignoring errors. Required because the
        # NA() rewiring on Load Log H/J/K (above) propagates #N/A into
        # Charts D-G for empty rows, and plain MIN errors out on any #N/A.
        new_m2 = ('="Best in: "&IFERROR(INDEX(Charts!$R$18:$R$25,'
                  'MATCH(AGGREGATE(5,6,Charts!$L$18:$L$25),'
                  'Charts!$L$18:$L$25,0)),"")')
        if loadlog["M2"].value != new_m2:
            loadlog["M2"] = new_m2
            fixes.append("Load Log!M2: simplified Best-in formula (cross-sheet INDEX/MATCH)")
        if charts["R17"].value != "Best in (per row)":
            charts["R17"] = "Best in (per row)"
            fixes.append("Charts!R17: added Best-in (per row) header")
        for r in range(18, 26):
            want = (f'=IFERROR('
                    f'IF(D{r}=AGGREGATE(5,6,$D$18:$D$25),"Vel ✓ ","")&'
                    f'IF(E{r}=AGGREGATE(5,6,$E$18:$E$25),"SD ✓ ","")&'
                    f'IF(F{r}=AGGREGATE(5,6,$F$18:$F$25),"MR ✓ ","")&'
                    f'IF(G{r}=AGGREGATE(5,6,$G$18:$G$25),"Vert ✓ ","")'
                    f',"")')
            if charts.cell(r, 18).value != want:
                charts.cell(r, 18).value = want
                fixes.append(f"Charts!R{r}: per-row Best-in formula")

    # ---- Auto-fill session metadata from CSV data ----
    # Skip if the workbook isn't a Load Log structure.
    if "Load Log" in wb.sheetnames:
        ws = wb["Load Log"]

        # Auto-fill Date in the Test Session bar (row 13, column B).
        # Use earliest date present in chronograph or group records.
        if (chronograph_records or group_records):
            b13 = ws.cell(13, 2)
            if b13.value in (None, ""):
                # Collect candidate dates from both sources
                date_candidates = []
                for d in (chronograph_records or []):
                    dv = d.get("Date")
                    if dv: date_candidates.append(str(dv).strip())
                for g in (group_records or []):
                    dv = g.get("Date")
                    if dv: date_candidates.append(str(dv).strip())
                if date_candidates:
                    # Pick the first encountered (typically all from same range trip)
                    b13.value = date_candidates[0]
                    fixes.append(f"Load Log!B13 (Date): auto-filled = {date_candidates[0]!r}")

        # Auto-fill Cartridge from BallisticX caliber (row 5, column L)
        if group_records:
            calibers = [g.get("Caliber") for g in group_records if g.get("Caliber")]
            if calibers:
                from collections import Counter
                most_common_cal = Counter(calibers).most_common(1)[0][0]
                l5 = ws.cell(5, 12)  # Cartridge value cell
                if l5.value in (None, ""):
                    l5.value = most_common_cal
                    fixes.append(f"Load Log!L5 (Cartridge): auto-filled = {most_common_cal!r}")

        # Auto-fill Bullet weight in row 9 (append to Bullet field if it has
        # text but no weight, or set to "{weight} gr" if empty)
        if chronograph_records:
            weights = [d.get("BulletWt") for d in chronograph_records if d.get("BulletWt")]
            if weights:
                from collections import Counter
                most_common_wt = Counter(weights).most_common(1)[0][0]
                b9 = ws.cell(9, 2)
                cur = (b9.value or "")
                if not isinstance(cur, str):
                    cur = str(cur)
                # If the current value already mentions the weight, skip
                if not (cur and (f"{most_common_wt}" in cur or "gr" in cur.lower())):
                    if cur:
                        b9.value = f"{cur} ({most_common_wt} gr)"
                    else:
                        b9.value = f"{most_common_wt} gr"
                    fixes.append(f"Load Log!B9 (Bullet): added weight = {most_common_wt} gr")

        # Auto-fill Session note from Garmin into the Notes column for the
        # matching row. Each chronograph record may have a SessionNote.
        if chronograph_records:
            for d in chronograph_records:
                note = (d.get("SessionNote") or "").strip()
                tag = (d.get("Tag") or "").strip()
                if not note or not tag:
                    continue
                # Find which Load Log row this tag will land in. Tags map to
                # rows by order of P1, P2, ... appearing in the data table.
                # Simpler approach: scan B16:B25, find matching charge, write note.
                # The matching is loose — we just want to ENRICH, never overwrite.
                # Since tag→row mapping depends on lookup formulas, just append
                # session notes to the row 13 Test Session Notes field if it
                # doesn't already contain them.
                k13 = ws.cell(13, 12)  # Notes value cell (column L is value-of K13 label)
                # Actually K13 is the "Notes:" label. Value cell is L13.
                # Re-check: row 13 layout is A=Date:, B=date, F=Temp:, G=temp, K=Notes:, L=notes
                l13 = ws.cell(13, 12)
                cur_notes = str(l13.value or "")
                if note not in cur_notes:
                    new_notes = (cur_notes + " " + note).strip() if cur_notes else note
                    if new_notes != cur_notes:
                        l13.value = new_notes
                        fixes.append(f"Load Log!L13 (Notes): appended session note from {tag}")

    # Reset every user-facing sheet's saved scroll/selection state so
    # the workbook opens at the top of each tab next time. Excel persists
    # the scroll position per-sheet — resetting on every import means the
    # saved file always opens fresh, regardless of where the user last
    # scrolled in their previous session.
    from openpyxl.worksheet.views import Selection
    user_facing_sheets = (
        # "Start Here" is the v0.14 rename; keep "After Range Day" as
        # fallback so this loop hits whichever name the sheet has at the
        # moment apply_workbook_repairs runs.
        "Start Here", "After Range Day", "Load Log", "Charts", "Seating Depth",
        "Garmin Xero Import", "BallisticX Import", "Load Library", "Ballistics",
    )
    view_resets = 0
    for sheet_name in user_facing_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        view = wb[sheet_name].sheet_view
        changed = False
        if view.topLeftCell is not None:
            view.topLeftCell = None
            changed = True
        # Force selection to A1
        target_sel = [Selection(activeCell="A1", sqref="A1")]
        # Compare by attribute since Selection __eq__ may not be identity-safe
        cur_sel = list(view.selection or [])
        if (len(cur_sel) != 1 or
                getattr(cur_sel[0], "activeCell", None) != "A1" or
                getattr(cur_sel[0], "sqref", None) != "A1"):
            view.selection = target_sel
            changed = True
        # Only the welcome tab (Start Here / legacy After Range Day) should
        # claim tabSelected; everyone else is unselected. activeTab="0" in
        # workbook.xml picks the initial tab on open.
        want_selected = sheet_name in ("Start Here", "After Range Day")
        if bool(view.tabSelected) != want_selected:
            view.tabSelected = want_selected
            changed = True
        if changed:
            view_resets += 1
    if view_resets:
        fixes.append(
            f"Reset saved scroll/selection state on {view_resets} sheet(s) "
            "so workbook opens at top of every tab"
        )

    if fixes:
        print("  Workbook repairs applied:")
        for f in fixes:
            print(f"    - {f}")

    # v0.14 fix — repair template-level row-height + column-width bugs
    # carried in the .xltx that make important rows nearly invisible. Each
    # is a 1-line height set; idempotent (won't shrink rows the user widened).
    LAYOUT_FIXES = [
        ("Load Log", "row", 7, 22.0,
         "MOA/MIL scope click dropdown row (G7:J7) was 6.0 px — invisible"),
        ("Seating Depth", "row", 26, 30.0,
         "'Save Suggested Load to Library' button row was 7.5 px — invisible"),
    ]
    for sheet_name, kind, idx, target, why in LAYOUT_FIXES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if kind == "row":
            current = ws.row_dimensions[idx].height if idx in ws.row_dimensions else None
            if current is None or current < target:
                ws.row_dimensions[idx].height = target
                fixes.append(f"{sheet_name}!row{idx}: height {current or 'default'} → {target} ({why})")

    # v0.14 fix — Turret-type dropdown on G7 of both Load Log and Seating
    # Depth. The same list is mirrored on both sheets so the user can pick
    # (or see) their click value on whichever tab they're working in.
    # "N/A" added 2026-05-13 for users whose scope has no adjustable turret
    # (e.g., fixed-power optics, BDC reticles only).
    TURRET_LIST = '"0.1 Mil,0.05 Mil,1/4 MOA,1/8 MOA,N/A"'

    def _replace_g7_validation(ws, sheet_name_for_log):
        # Remove any existing G7 validation
        from openpyxl.worksheet.datavalidation import DataValidation
        keep = []
        for dv in ws.data_validations.dataValidation:
            covers_g7 = False
            for rng in dv.sqref.ranges:
                if rng.min_row <= 7 <= rng.max_row and rng.min_col <= 7 <= rng.max_col:
                    covers_g7 = True
                    break
            if not covers_g7:
                keep.append(dv)
        ws.data_validations.dataValidation = keep
        # v0.14 fix: apply to the full merged range G7:J7 (not just G7).
        # On Mac Excel, dropdowns whose sqref covers only the top-left of a
        # merged range sometimes refuse to draw the arrow — Chad couldn't see
        # the dropdown 2026-05-14 even though the validation was present.
        # Adding an input-message prompt also makes it discoverable on click.
        dv = DataValidation(
            type="list",
            formula1=TURRET_LIST,
            allow_blank=True,
            showDropDown=False,  # 0 = SHOW arrow (Excel's inverted flag)
            showInputMessage=True,
            promptTitle="Pick your turret clicks",
            prompt=(
                "Click the arrow on the right to choose your scope's click "
                "value (0.1 Mil, 0.05 Mil, 1/4 MOA, 1/8 MOA, or N/A)."
            ),
        )
        dv.add("G7:J7")
        ws.add_data_validation(dv)
        fixes.append(
            f"{sheet_name_for_log}!G7:J7: turret dropdown {TURRET_LIST} "
            "(applied to full merge + input-message prompt)"
        )

    for sheet_name in ("Load Log", "Seating Depth"):
        if sheet_name in wb.sheetnames:
            _replace_g7_validation(wb[sheet_name], sheet_name)

    # v0.14 fix — F7 label describing what G7's dropdown is for, and match
    # row 7's styling to rows 5/6 (rifle/shooter/cartridge row). Without
    # this, the Turret Type row looks like an orphan vs the rest of the
    # rifle-setup block.
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    yellow_label_fill = PatternFill(
        start_color="FFFFE082", end_color="FFFFE082", fill_type="solid"
    )
    cream_value_fill = PatternFill(
        start_color="FFFFF8E1", end_color="FFFFF8E1", fill_type="solid"
    )
    label_font = Font(color="FF1F4E78", bold=True)
    value_font = Font(color="FF000000")
    # Alignment conventions (matches rows 5/6/9/10/13):
    #   Labels (A/F/K/N column): horizontal=right, vertical=center
    #   Values (B/G/L/O column): horizontal=left,  vertical=center
    label_align = Alignment(horizontal="right", vertical="center")
    value_align = Alignment(horizontal="left", vertical="center")
    # Thin gray border to match rest of rifle-setup block (style="thin",
    # color="FFB0B0B0").
    gray_side = Side(style="thin", color="FFB0B0B0")
    rifle_border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)
    for sheet_name in ("Load Log", "Seating Depth"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if not ws["F7"].value:
            ws["F7"].value = "Turret:"
            fixes.append(f"{sheet_name}!F7: added 'Turret:' label for G7 dropdown")
        # Ensure G7:J7 merge exists (Load Log already has it from the
        # template; Seating Depth doesn't, leaving G/H/I/J as four
        # separate cells which looks broken next to row 6's merge).
        g7_merge_present = False
        for r in ws.merged_cells.ranges:
            if r.min_row == 7 and r.max_row == 7 and r.min_col == 7 and r.max_col == 10:
                g7_merge_present = True
                break
        if not g7_merge_present:
            ws.merge_cells("G7:J7")
            fixes.append(f"{sheet_name}!G7:J7: added missing merge for turret dropdown")
        # Match row 7 height + fills to row 5/6 (rifle-setup block)
        ws.row_dimensions[7].height = 31.5
        # F7 = label cell
        ws["F7"].fill = yellow_label_fill
        ws["F7"].font = label_font
        ws["F7"].alignment = label_align
        ws["F7"].border = rifle_border
        # G7:J7 merge — apply value styling to every cell in the merge so
        # the border draws continuously (Excel uses each cell's own border
        # within a merge for the outline). value font + alignment on the
        # top-left only; fill on top-left propagates visually to the merge.
        ws["G7"].fill = cream_value_fill
        ws["G7"].font = value_font
        ws["G7"].alignment = value_align
        for coord in ("G7", "H7", "I7", "J7"):
            ws[coord].border = rifle_border

    # v0.14 fix — give all action-button cells consistent yellow accent +
    # bold blue text + thin border so they read as clickable. Each cell is
    # already a hyperlink (clicking triggers the loadscope:// URL handler),
    # but template styling left them as plain text — easy to miss visually.
    button_fill = PatternFill(start_color="FFFFE082", end_color="FFFFE082", fill_type="solid")
    button_font = Font(color="FF1F4E78", bold=True)
    BUTTON_CELLS = [
        ("Charts", "A6"),    # Save Suggested Load to Library
        ("Charts", "A12"),   # Reset weights to Loadscope defaults
        ("Seating Depth", "A26"),  # Save Suggested Load to Library
        ("Ballistics", "A2"),       # Print Pocket Range Card
    ]
    for sheet_name, coord in BUTTON_CELLS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        c = ws[coord]
        if c.hyperlink:  # only style cells that are actually clickable
            c.fill = button_fill
            c.font = button_font
            c.border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)
            fixes.append(f"{sheet_name}!{coord}: button styling (yellow fill + bold blue + border)")

    # v0.14.2 — Now that Excel's ribbon is hidden when Loadscope opens a
    # workbook, users can't reach the standard Print button anymore. Add
    # in-workbook "Print This Workbook" buttons on Charts (the analysis
    # hub) and Ballistics (alongside the existing Pocket Range Card
    # button) so printing is always one click away. Routes through the
    # loadscope://print-workbook URL handler.
    PRINT_BUTTON_CELLS = [
        ("Charts", "A14"),       # Below existing Save / Reset buttons
        ("Ballistics", "A3"),    # Below the Pocket Range Card button
    ]
    print_button_label = "→  Print This Workbook  ←"
    print_button_url = "https://loadscope.app/launch?action=print-workbook"
    for sheet_name, coord in PRINT_BUTTON_CELLS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        c = ws[coord]
        # Skip if the cell was already populated (e.g., by a future
        # template version that added a different button there) so we
        # don't clobber unrelated content.
        if c.value not in (None, "") and c.value != print_button_label:
            continue
        c.value = print_button_label
        c.hyperlink = print_button_url
        c.fill = button_fill
        c.font = button_font
        c.border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Match neighboring button row heights for visual consistency.
        ws.row_dimensions[c.row].height = 22
        fixes.append(f"{sheet_name}!{coord}: added 'Print This Workbook' button")

    # v0.14 fix — SD!A28 "Reset weights" cell shares row 28 with the
    # weight VALUE cells (B28-L28), so column A can't merge wider for the
    # long instruction text. Inserting a new dedicated row would shift the
    # $C$28/$F$28/$I$28/$L$28 references downstream formulas use. Pragmatic
    # fix: shorten the text to fit column A's width, move the "hover for
    # rationale" hint into a cell comment.
    from openpyxl.styles import Alignment
    from openpyxl.comments import Comment
    if "Seating Depth" in wb.sheetnames:
        sd_ws = wb["Seating Depth"]
        a28 = sd_ws["A28"]
        current_value = (a28.value or "").strip()
        if current_value and "Reset" in current_value:
            # Widen column A so "Reset weights" fits (was 10; bump to 16).
            current_w = sd_ws.column_dimensions["A"].width if "A" in sd_ws.column_dimensions else None
            if current_w is None or current_w < 16:
                sd_ws.column_dimensions["A"].width = 16
            a28.value = "Reset weights"
            current_h = sd_ws.row_dimensions[28].height if 28 in sd_ws.row_dimensions else None
            if current_h is None or current_h < 24:
                sd_ws.row_dimensions[28].height = 24
            a28.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=False, shrink_to_fit=False,
            )
            # Yellow accent fill makes it look like a button
            a28.fill = PatternFill(
                start_color="FFFFE082", end_color="FFFFE082", fill_type="solid"
            )
            a28.font = Font(color="FF1F4E78", bold=True)
            a28.border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)
            fixes.append("Seating Depth!A28: 'Reset weights' button + widened col A + yellow accent")

    # v0.14 fix — Excel-Mac doesn't reliably recompute the AGGREGATE-based
    # composite-score formula chain after an openpyxl save. Precompute the
    # static values in Python and write them as cell values so the user
    # opens the workbook to a fully-rendered analysis grid, not a sea of
    # empty cells that "would" compute if Excel recalced.
    static_fixes = _write_static_analysis_values(wb, chronograph_records or [], group_records or [])
    if static_fixes:
        fixes.extend(static_fixes)
        print("Wrote static analysis values:")
        for f in static_fixes:
            print(f"    - {f}")

    # v0.14 fix: Seating Depth!B10 charge cell — show a clear placeholder
    # when no powder ladder data has populated Charts!B3 (Chad 2026-05-14:
    # users who only do a seating depth test wouldn't know to type the
    # charge into a blank cell). Smart formula: empty Charts!B3 → display
    # "← Type your charge (gr) here"; populated → display the winning charge.
    if "Seating Depth" in wb.sheetnames:
        sd_sheet = wb["Seating Depth"]
        b10_val = sd_sheet["B10"].value
        legacy_formula = (
            isinstance(b10_val, str) and b10_val.strip() in ("=Charts!B3",)
        )
        # Skip if user typed a real number — don't clobber their input.
        if b10_val in (None, "") or legacy_formula:
            sd_sheet["B10"].value = (
                '=IF(Charts!B3="","← Type your charge (gr) here",Charts!B3)'
            )
            fixes.append(
                "Seating Depth!B10: smart placeholder formula "
                "(shows 'Type your charge here' when no PL data)"
            )

    # v0.14 fix: Garmin Xero Import + BallisticX Import tabs look "very busy"
    # (Chad 2026-05-14). 20 columns wide with 10 individual shot columns
    # drowns out the aggregate stats most users actually consult. Polish:
    # collapse Shot columns into a [+] outline group, hide gridlines,
    # alternating row stripes, stronger section banners.
    polish_fixes = _polish_import_tabs(wb)
    if polish_fixes:
        fixes.extend(polish_fixes)
        print("Polished import tabs:")
        for f in polish_fixes:
            print(f"    - {f}")

    # v0.14 fix: Load Library Load Name column (C) was cutting off long
    # composite names like "6.5 CM 140 Berger Hybrid / H4350 42.1 / 0.020 jump"
    # (50 chars at col width 25). Widen col C + enable wrap_text on data rows.
    ll_fixes = _polish_load_library(wb)
    if ll_fixes:
        fixes.extend(ll_fixes)
        print("Polished Load Library:")
        for f in ll_fixes:
            print(f"    - {f}")

    # v0.14 fix: After Range Day → Start Here onboarding rebuild
    # (Chad 2026-05-14). Major restructure: simpler organization, friendlier
    # font (Avenir Next), accent section headers, step badges. The tab is
    # the FIRST one users see — its job is orientation, not docs.
    sh_fixes = _polish_start_here_tab(wb)
    if sh_fixes:
        fixes.extend(sh_fixes)
        print("Rebuilt Start Here:")
        for f in sh_fixes:
            print(f"    - {f}")

    # v0.14: Apply shrink_to_fit=True to every numeric/formula cell
    # across visible tabs so users never see ##### when a column is
    # too narrow — Excel auto-shrinks the font instead.
    # Chad 2026-05-14: "I don't ever want hashtags to show up."
    sf_fixes = _enable_shrink_to_fit_on_numeric_cells(wb)
    if sf_fixes:
        fixes.extend(sf_fixes)
        print("Enabled shrink-to-fit on numeric cells:")
        for f in sf_fixes:
            print(f"    - {f}")

    # v0.14: Lock formula + computed cells so users can't accidentally
    # overwrite them. Chad 2026-05-14: "I would like it to be impossible
    # for a user to overwrite any formula." Per-tab whitelist of
    # user-input cells stays unlocked. Import tabs (Garmin Xero +
    # BallisticX) stay fully unprotected so users can fix CSV typos.
    protect_fixes = _protect_workbook(wb)
    if protect_fixes:
        fixes.extend(protect_fixes)
        print("Protected formula + computed cells:")
        for f in protect_fixes:
            print(f"    - {f}")

    # v0.14: Smart placeholder on Seating Depth!B10 (the "Charge:" cell).
    # Normally the cell formula is =Charts!B3 → resolves to the powder
    # ladder winner. But if the user runs ONLY a seating depth test (no
    # powder ladder), Charts!B3 is empty and the user sees a blank cell
    # without knowing they can type the charge themselves. Replace the
    # raw formula with one that displays a placeholder when empty.
    # Chad 2026-05-14: "If someone is only doing a seating test, how
    # does our spreadsheet know how many grains of powder to enter?"
    if "Seating Depth" in wb.sheetnames:
        sd_ws = wb["Seating Depth"]
        existing_b10 = sd_ws["B10"].value
        # Don't clobber a user-typed numeric value (means they already
        # set it manually). Only refresh formula-based or empty cells.
        if existing_b10 is None or (
            isinstance(existing_b10, str) and existing_b10.startswith("=")
        ):
            sd_ws["B10"].value = (
                '=IF(Charts!B3="","← Type your charge (gr) here",Charts!B3)'
            )
            fixes.append(
                "Seating Depth!B10: smart formula shows placeholder when "
                "no powder ladder winner is set"
            )

    # v0.14 fix: Ballistics tab DOPE callout (A4) was a faded gray strip that
    # blended into the page. Chad 2026-05-14: users may not realize the DOPE
    # columns are theirs to fill at the range. Make the callout a prominent
    # accent banner anchored above the DOPE table.
    bal_fixes = _polish_ballistics_callout(wb)
    if bal_fixes:
        fixes.extend(bal_fixes)
        print("Polished Ballistics callout:")
        for f in bal_fixes:
            print(f"    - {f}")

    return fixes


def _write_static_analysis_values(wb, chronograph_records, group_records):
    """Precompute composite scores + Best-in tags + normalized metrics and
    write them as STATIC values to Charts + Load Log + Seating Depth.

    Why: Excel-Mac doesn't reliably recompute the AGGREGATE-based composite
    chain after openpyxl saves. Writing static values forces the right
    display every time, at the cost of the user not being able to change
    weights and see the recompute live (which they could do before only if
    Excel happened to cooperate — many didn't).

    Tags: powder ladder uses "P1"-"P10", seating depth uses "S1"-"S10".
    Matched cross-record by Tag (chrono provides shots/SD/AvgVel, group
    provides MR/SD-Vert/group size).
    """
    fixes = []
    if not chronograph_records:
        return fixes

    # Build tag -> group_record map for fast lookup
    group_by_tag = {g.get("Tag", ""): g for g in group_records}

    # Partition chronograph records into powder-ladder (P*) vs seating-depth (S*)
    pl_chrono = [c for c in chronograph_records if str(c.get("Tag", "")).startswith("P")]
    sd_chrono = [c for c in chronograph_records if str(c.get("Tag", "")).startswith("S")]

    if pl_chrono:
        fixes.extend(_write_pl_static(wb, pl_chrono, group_by_tag))
    if sd_chrono:
        fixes.extend(_write_sd_static(wb, sd_chrono, group_by_tag))

    return fixes


def _pair_chrono_with_groups(chrono_records, group_by_tag):
    """Build a list of unified candidate dicts (one per chrono record that
    has a matching group record). Skip records without a group match —
    they don't have the metrics needed for normalization."""
    candidates = []
    for c in chrono_records:
        tag = c.get("Tag", "")
        grp = group_by_tag.get(tag)
        if not grp:
            continue
        # Skip rows with missing essential metrics
        if c.get("AvgVel") in (None, "") or c.get("SD") in (None, ""):
            continue
        if grp.get("GroupIn") in (None, "") or grp.get("MRIn") in (None, ""):
            continue
        candidates.append({
            "tag": tag,
            "charge_or_jump": c.get("ChargeOrJump"),
            "vel": c.get("AvgVel"),
            "sd": c.get("SD"),
            "group": grp.get("GroupIn"),
            "mr": grp.get("MRIn"),
            "vert": grp.get("SDVertIn") or 0,
            "shots": [s for s in (c.get("Shots") or []) if s is not None],
        })
    return candidates


def _normalize(values):
    """Return 0-1 normalized values where 0 = best (smallest). Identical
    values all return 0.0 (no spread to normalize)."""
    if not values:
        return []
    lo, hi = min(values), max(values)
    rng = hi - lo
    return [0.0 if rng == 0 else (v - lo) / rng for v in values]


def _compute_composites_and_bests(candidates, weights):
    """Given candidates + (w_vel, w_sd, w_mr, w_vert) weights, return:
        composites (floored at 0.001 so winner displays),
        best_in_tags_per_row,
        winner_idx
    """
    norm_group = _normalize([c["group"] for c in candidates])
    norm_sd = _normalize([c["sd"] for c in candidates])
    norm_mr = _normalize([c["mr"] for c in candidates])
    norm_vert = _normalize([c["vert"] for c in candidates])
    w_vel, w_sd, w_mr, w_vert = weights
    composites = [
        w_vel * ng + w_sd * nsd + w_mr * nmr + w_vert * nvert
        for ng, nsd, nmr, nvert in zip(norm_group, norm_sd, norm_mr, norm_vert)
    ]
    composites = [max(0.001, c) for c in composites]

    def best_idx(vals):
        return vals.index(min(vals)) if vals else 0
    bests = {
        "Vel ✓": best_idx([c["group"] for c in candidates]),
        "SD ✓": best_idx([c["sd"] for c in candidates]),
        "MR ✓": best_idx([c["mr"] for c in candidates]),
        "Vert ✓": best_idx([c["vert"] for c in candidates]),
    }
    tags_per_row = []
    for i in range(len(candidates)):
        tags = [t for t, idx in bests.items() if idx == i]
        # When a single row wins all four metrics, "All metrics ✓" is shorter
        # and reads better than concatenating four "X ✓" tags (which wraps
        # awkwardly in the merged "Best in:" cell).
        if len(tags) == 4:
            tags_per_row.append("All metrics ✓")
        else:
            tags_per_row.append(" ".join(tags))
    winner_idx = composites.index(min(composites))
    return composites, tags_per_row, winner_idx, norm_group, norm_sd, norm_mr, norm_vert


def _write_pl_static(wb, pl_chrono, group_by_tag):
    """Powder ladder static-value writes targeting Load Log + Charts."""
    fixes = []
    candidates = _pair_chrono_with_groups(pl_chrono, group_by_tag)
    if len(candidates) < 2:
        return fixes  # need at least 2 to normalize

    weights = (
        DEFAULT_WEIGHTS_CHARTS["B11"],
        DEFAULT_WEIGHTS_CHARTS["D11"],
        DEFAULT_WEIGHTS_CHARTS["F11"],
        DEFAULT_WEIGHTS_CHARTS["H11"],
    )
    composites, tags_per_row, winner_idx, ng, nsd, nmr, nvert = (
        _compute_composites_and_bests(candidates, weights)
    )
    winner = candidates[winner_idx]
    winner_tags = tags_per_row[winner_idx] or "Composite ✓"

    if "Load Log" in wb.sheetnames:
        ll = wb["Load Log"]
        # Winner row 2 (D2/G2/J2/L2 + M2 Best in)
        ll["D2"].value = winner["charge_or_jump"]
        ll["G2"].value = winner["vel"]
        ll["J2"].value = winner["sd"]
        ll["L2"].value = winner["mr"]
        ll["M2"].value = f"Best in:  {winner_tags}"
        # Enable text wrap + bump row 2 height so long Best-in labels fit
        from openpyxl.styles import Alignment
        ll["M2"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        current_h = ll.row_dimensions[2].height if 2 in ll.row_dimensions else None
        if current_h is None or current_h < 38:
            ll.row_dimensions[2].height = 38
        # O16:O25 — composite per row, pad with #N/A
        for i, comp in enumerate(composites):
            ll.cell(row=16 + i, column=15).value = round(comp, 3)
        for r in range(16 + len(composites), 26):
            ll.cell(row=r, column=15).value = "=NA()"
        # P16:P25 — top-3 medal ranking. Template ships with a formula that
        # only labels the single winner ("⭐ WINNER" or blank); Chad asked
        # 2026-05-14 to see 1st/2nd/3rd so the runner-up loads are visible
        # at a glance. Composite is "lower = better" (winner_idx is min);
        # sort ASCENDING so rank 1 = lowest composite.
        ranked = sorted(range(len(composites)), key=lambda i: composites[i])
        medal_by_idx = {}
        if len(ranked) >= 1:
            medal_by_idx[ranked[0]] = "🥇 WINNER"
        if len(ranked) >= 2:
            medal_by_idx[ranked[1]] = "🥈 2nd"
        if len(ranked) >= 3:
            medal_by_idx[ranked[2]] = "🥉 3rd"
        # Enable shrink_to_fit on P column so "🥇 WINNER" fits at width 15
        # without being cut off (Mac Excel renders emoji at ~2 char widths).
        from openpyxl.styles import Alignment as _PAlign
        for i in range(len(composites)):
            cell = ll.cell(row=16 + i, column=16)
            cell.value = medal_by_idx.get(i, "")
            cell.alignment = _PAlign(
                horizontal="center", vertical="center", shrink_to_fit=True
            )
        for r in range(16 + len(composites), 26):
            ll.cell(row=r, column=16).value = ""
        # Widen col P so the winner label fits without shrinking
        ll.column_dimensions["P"].width = 15
        fixes.append(
            f"Load Log!D2/G2/J2/L2/M2 + O16:O{15+len(composites)} + "
            f"P16:P25 top-3 medals static values + col P widened to 15"
        )

    # Auto-fill Ballistics tab header (B5/E5/H5/K5 = rifle/bullet/charge/vel)
    # from the data we already have. The Pocket Range Card reads these
    # cells via data_only=True (cached values) — without auto-fill the
    # template's cross-sheet formulas show blank → card prints with dashes.
    if "Ballistics" in wb.sheetnames and "Load Log" in wb.sheetnames:
        ball = wb["Ballistics"]
        ll = wb["Load Log"]
        # Pull rifle + bullet from Load Log's user-fill cells (header row 5/9)
        rifle = ll["B5"].value
        bullet = ll["B9"].value
        scope = ll["G6"].value
        if rifle and not ball["B5"].value:
            ball["B5"].value = rifle
        if bullet and not ball["E5"].value:
            ball["E5"].value = bullet
        if scope and not ball["B6"].value:
            ball["B6"].value = scope
        # Winning charge + velocity come from the precomputed winner
        ball["H5"].value = winner["charge_or_jump"]
        ball["K5"].value = winner["vel"]
        # Long rifle/scope names get cut off — col B default 13 chars,
        # "Leupold Mark 5HD 5-25x56" is 24 chars. Merge B:C on rows 5/6 so
        # the long names have room without breaking the DOPE table below.
        for r_idx, coord_pair in ((5, "B5:C5"), (6, "B6:C6")):
            merged = any(
                rr.min_row == rr.max_row == r_idx
                and rr.min_col == 2 and rr.max_col == 3
                for rr in ball.merged_cells.ranges
            )
            if not merged:
                ball.merge_cells(coord_pair)
        # Disable wrap so long names fit on one line in the merged width
        from openpyxl.styles import Alignment as _Align
        for coord in ("B5", "B6"):
            existing = ball[coord].alignment
            ball[coord].alignment = _Align(
                horizontal=existing.horizontal or "left",
                vertical=existing.vertical or "center",
                wrap_text=False,
            )
        fixes.append(
            "Ballistics row 5/6: auto-filled rifle/bullet/scope from Load Log, "
            f"charge/vel from winner ({winner['charge_or_jump']}gr @ {winner['vel']} fps); "
            "merged B:C on rows 5+6 for long rifle/scope names"
        )

    if "Charts" in wb.sheetnames:
        ch = wb["Charts"]
        # Pin to test mode for predictable row alignment
        ch["B100"].value = "test"
        # Row 17 headers — overwrite IF($B$100="window",...) formulas + the
        # "Norm Spread" label which only matches D17's window-mode meaning.
        # In test mode the analysis is per-row averages, not 3-row spreads.
        ch["A17"].value = "Charge"
        ch["D17"].value = "Avg Vel (fps)"
        ch["H17"].value = "Norm Vel"
        # B17 / C17 (Low / High) headers — these only describe window-mode
        # rolling-window data. In test mode the columns are blank, so the
        # headers are misleading. Clear them.
        ch["B17"].value = None
        ch["C17"].value = None
        # Row 3-5 winner-summary cells (replace ArrayFormulas + hardcoded G5)
        ch["B3"].value = winner["charge_or_jump"]
        ch["E3"].value = round(composites[winner_idx], 3)
        ch["G3"].value = winner["sd"]
        ch["G4"].value = winner["mr"]
        ch["E5"].value = winner["vel"]
        ch["G5"].value = winner_tags
        # Analysis grid rows 18-25 (floor normalized at 0.001 so 0s display).
        # B/C (Low/High) columns are intentionally LEFT BLANK in test mode —
        # they only carry meaning in window mode (low/middle/high charges of
        # a 3-step rolling window). The template formula returns "" in test
        # mode; we write empty to match.
        for i, c in enumerate(candidates):
            r = 18 + i
            ch.cell(row=r, column=2).value = None  # B Low (window-mode only)
            ch.cell(row=r, column=3).value = None  # C High (window-mode only)
            ch.cell(row=r, column=4).value = c["vel"]
            ch.cell(row=r, column=5).value = c["sd"]
            ch.cell(row=r, column=6).value = c["mr"]
            ch.cell(row=r, column=7).value = c["vert"]
            ch.cell(row=r, column=8).value = round(max(0.001, ng[i]), 3)
            ch.cell(row=r, column=9).value = round(max(0.001, nsd[i]), 3)
            ch.cell(row=r, column=10).value = round(max(0.001, nmr[i]), 3)
            ch.cell(row=r, column=11).value = round(max(0.001, nvert[i]), 3)
            ch.cell(row=r, column=12).value = round(composites[i], 3)
            if tags_per_row[i]:
                ch.cell(row=r, column=18).value = tags_per_row[i]
        # Sorted Vel-vs-Charge helper (T18:U25)
        sorted_pairs = sorted([(c["charge_or_jump"], c["vel"]) for c in candidates])
        for i, (chg, vel) in enumerate(sorted_pairs):
            ch.cell(row=18 + i, column=20).value = chg
            ch.cell(row=18 + i, column=21).value = vel
        for r in range(18 + len(sorted_pairs), 26):
            ch.cell(row=r, column=20).value = "=NA()"
            ch.cell(row=r, column=21).value = "=NA()"
        fixes.append(f"Charts!B-L18:25 + R18:25 + T-U18:25 static values ({len(candidates)} candidates)")

    return fixes


def _write_sd_static(wb, sd_chrono, group_by_tag):
    """Seating depth static-value writes targeting Seating Depth sheet."""
    fixes = []
    candidates = _pair_chrono_with_groups(sd_chrono, group_by_tag)
    if len(candidates) < 2:
        return fixes

    weights = (
        DEFAULT_WEIGHTS_SEATING_DEPTH["C28"],
        DEFAULT_WEIGHTS_SEATING_DEPTH["F28"],
        DEFAULT_WEIGHTS_SEATING_DEPTH["I28"],
        DEFAULT_WEIGHTS_SEATING_DEPTH["L28"],
    )
    composites, tags_per_row, winner_idx, ng, nsd, nmr, nvert = (
        _compute_composites_and_bests(candidates, weights)
    )
    winner = candidates[winner_idx]
    winner_tags = tags_per_row[winner_idx] or "Composite ✓"

    if "Seating Depth" not in wb.sheetnames:
        return fixes

    sd = wb["Seating Depth"]
    # Winner row 2
    sd["D2"].value = winner["charge_or_jump"]
    sd["G2"].value = winner["vel"]
    sd["J2"].value = winner["sd"]
    sd["L2"].value = winner["mr"]
    sd["N2"].value = winner["vert"]
    sd["O2"].value = f"Best in:  {winner_tags}"
    # Enable text wrap + bump row 2 height so long Best-in labels fit
    from openpyxl.styles import Alignment
    sd["O2"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    current_h = sd.row_dimensions[2].height if 2 in sd.row_dimensions else None
    if current_h is None or current_h < 38:
        sd.row_dimensions[2].height = 38
    # O column composite per row (rows 16-25), pad #N/A
    for i, comp in enumerate(composites):
        sd.cell(row=16 + i, column=15).value = round(comp, 3)
    for r in range(16 + len(composites), 26):
        sd.cell(row=r, column=15).value = "=NA()"
    # P16:P25 — top-3 medal ranking (matches Load Log treatment, Chad
    # 2026-05-14). Composite is "lower = better"; sort ASCENDING.
    sd_ranked = sorted(range(len(composites)), key=lambda i: composites[i])
    sd_medals = {}
    if len(sd_ranked) >= 1:
        sd_medals[sd_ranked[0]] = "🥇 WINNER"
    if len(sd_ranked) >= 2:
        sd_medals[sd_ranked[1]] = "🥈 2nd"
    if len(sd_ranked) >= 3:
        sd_medals[sd_ranked[2]] = "🥉 3rd"
    sd_p_align = Alignment(
        horizontal="center", vertical="center", shrink_to_fit=True
    )
    for i in range(len(composites)):
        cell = sd.cell(row=16 + i, column=16)
        cell.value = sd_medals.get(i, "")
        cell.alignment = sd_p_align
    for r in range(16 + len(composites), 26):
        sd.cell(row=r, column=16).value = ""
    sd.column_dimensions["P"].width = 15
    # Row 29 headers — overwrite IF($B$100="window",...) formulas + the
    # "Norm Spread" label which only matches D29's window-mode meaning.
    # B29 / C29 (Low / High) headers cleared in test mode since the
    # columns underneath are blank — meaningless headers look broken.
    sd["A29"].value = "Jump"
    sd["B29"].value = None
    sd["C29"].value = None
    sd["D29"].value = "Avg Vel (fps)"
    sd["H29"].value = "Norm Vel"
    # Analysis grid rows 30-37. B/C (Low/High) are intentionally left blank
    # in test mode — they only carry meaning in window mode (3-step rolling
    # window of jumps).
    # Compute ranks from composites (1 = lowest = best)
    sorted_idx = sorted(range(len(composites)), key=lambda i: composites[i])
    rank_by_idx = {idx: r for r, idx in enumerate(sorted_idx, start=1)}
    for i, c in enumerate(candidates):
        r = 30 + i
        sd.cell(row=r, column=1).value = c["charge_or_jump"]
        sd.cell(row=r, column=2).value = None  # B Low (window-mode only)
        sd.cell(row=r, column=3).value = None  # C High (window-mode only)
        sd.cell(row=r, column=4).value = c["vel"]
        sd.cell(row=r, column=5).value = c["sd"]
        sd.cell(row=r, column=6).value = c["mr"]
        sd.cell(row=r, column=7).value = c["vert"]
        # Floor normalized values at 0.001 so winner's 0 displays (Excel
        # number format `[=NA()]"";0.000` appears to hide exactly-zero)
        sd.cell(row=r, column=8).value = round(max(0.001, ng[i]), 3)
        sd.cell(row=r, column=9).value = round(max(0.001, nsd[i]), 3)
        sd.cell(row=r, column=10).value = round(max(0.001, nmr[i]), 3)
        sd.cell(row=r, column=11).value = round(max(0.001, nvert[i]), 3)
        sd.cell(row=r, column=12).value = round(composites[i], 3)
        sd.cell(row=r, column=13).value = rank_by_idx[i]  # M = Rank
        # N = Best in — always overwrite (empty for non-winners). Otherwise
        # the template's per-row formula displays uncomputed in Excel-Mac.
        sd.cell(row=r, column=14).value = tags_per_row[i] or ""
    for r in range(30 + len(candidates), 38):
        for col in (1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14):
            sd.cell(row=r, column=col).value = "=NA()"
    fixes.append(f"Seating Depth row 2 + O16:O{15+len(composites)} + A-N30:37 static values ({len(candidates)} candidates)")

    return fixes


def _polish_import_tabs(wb):
    """Apply consistent visual polish to Garmin Xero + BallisticX Import tabs.

    Chad called the Garmin tab "very busy" 2026-05-14. Same critique applies
    to BallisticX (same architectural pattern). Shared polish:
      - hide cell gridlines
      - tighten column widths
      - stronger navy section-banner styling on rows 3 / 16 / 28 / 45
      - alternating row stripes on data blocks
      - bold header row
    Garmin-specific: collapse the per-shot columns (E:N = Shot 1-10) into
    an Excel outline group with [+] expander.
    """
    fixes = []

    if "Garmin Xero Import" in wb.sheetnames:
        fixes.extend(_polish_one_import_tab(
            ws=wb["Garmin Xero Import"],
            max_col=20,  # T
            collapse_cols=("E", "F", "G", "H", "I", "J", "K", "L", "M", "N"),
            width_plan={
                "A": 6, "B": 11, "C": 30, "D": 7,
                # E-N hidden in outline group — leave widths alone
                "O": 10, "P": 8, "Q": 8, "R": 10, "S": 11, "T": 28,
            },
            notes_col=20, title_col=3,  # T = Notes, C = Session Title
            sheet_label="Garmin Xero Import",
        ))

    if "BallisticX Import" in wb.sheetnames:
        fixes.extend(_polish_one_import_tab(
            ws=wb["BallisticX Import"],
            max_col=13,  # M
            collapse_cols=(),  # no per-shot cols on BallisticX
            width_plan={
                "A": 6,   # Tag
                "B": 11,  # Date
                "C": 11,  # Distance (yd)
                "D": 9,   # Group (in)
                "E": 9,   # Width (in)
                "F": 9,   # Height (in)
                "G": 9,   # MR (in)
                "H": 9,   # CEP (in)
                "I": 11,  # SD Radial
                "J": 11,  # SD Vert
                "K": 11,  # SD Horiz
                "L": 9,   # Caliber
                "M": 28,  # Notes
            },
            notes_col=13, title_col=None,  # M = Notes, no separate Title col
            sheet_label="BallisticX Import",
        ))

    return fixes


def _polish_one_import_tab(
    ws, max_col, collapse_cols, width_plan, notes_col, title_col, sheet_label,
):
    """Apply the import-tab polish recipe to a single sheet."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fixes = []

    # 1. Hide cell gridlines — they're most of the visual noise.
    if ws.sheet_view is not None:
        ws.sheet_view.showGridLines = False

    # 2. Collapse Shot 1-10 cols (Garmin only) into an outline group.
    if collapse_cols:
        for col_letter in collapse_cols:
            cd = ws.column_dimensions[col_letter]
            cd.outline_level = 1
            cd.hidden = True
        ws.sheet_properties.outlinePr.summaryRight = True
        ws.sheet_properties.outlinePr.summaryBelow = True

    # 3. Tighten column widths.
    for col_letter, w in width_plan.items():
        ws.column_dimensions[col_letter].width = w

    # 4. Stronger section-banner styling on the four big merged headers
    # (rows 3 = POWDER LADDER, 16 = SEATING DEPTH, 28 = FREE-FORM,
    # 45 = HOW THE WIRING WORKS). Same row IDs on both Garmin + BallisticX.
    # Also extend banner / title / wiring-text merges to FULL width
    # (Garmin's template merges A:Q but data goes to col T — Chad 2026-05-14
    # noticed the right edge of banners stopped 3 cols short).
    banner_fill = PatternFill(
        start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid"
    )
    banner_font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
    banner_align = Alignment(horizontal="left", vertical="center", indent=1)

    # Re-merge any row whose merge stops before max_col so the banner
    # spans the full data-table width.
    def _ensure_full_width_merge(target_row, target_max_row=None):
        """Unmerge existing merges that touch target_row and re-merge to span
        cols 1..max_col across rows target_row..target_max_row (default same).
        """
        if target_max_row is None:
            target_max_row = target_row
        # Find existing merges that intersect this row band
        to_remove = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= target_row and rng.max_row <= target_max_row:
                to_remove.append(str(rng))
        for ref in to_remove:
            ws.unmerge_cells(ref)
        ws.merge_cells(start_row=target_row, start_column=1,
                       end_row=target_max_row, end_column=max_col)

    # Banner rows: re-merge to full width + apply navy fill
    for banner_row in (3, 16, 28, 45):
        _ensure_full_width_merge(banner_row)
        # Apply fill/font/align across the FULL row so unmerge+remerge
        # doesn't leave styling holes on the newly-included cells.
        for c in range(1, max_col + 1):
            cell = ws.cell(row=banner_row, column=c)
            cell.fill = banner_fill
        # Top-left of the merge holds the value + font + alignment
        ws.cell(row=banner_row, column=1).font = banner_font
        ws.cell(row=banner_row, column=1).alignment = banner_align
        ws.row_dimensions[banner_row].height = 26

    # Title row 1 + description row 2 also extend to full width
    for title_row in (1, 2):
        _ensure_full_width_merge(title_row)

    # "HOW THE WIRING WORKS" body block at A46 spans multiple rows
    # (A46:Q58 on Garmin, A46:M56 on BallisticX). Re-merge to full width.
    # Find current extent so we preserve the multi-row span.
    wiring_max_row = 46
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == 46 and rng.min_col == 1:
            wiring_max_row = max(wiring_max_row, rng.max_row)
    _ensure_full_width_merge(46, target_max_row=wiring_max_row)

    # 5. Alternating row stripes on the data blocks.
    stripe_fill = PatternFill(
        start_color="FFF5F7FA", end_color="FFF5F7FA", fill_type="solid"
    )
    no_fill = PatternFill(fill_type=None)
    body_font = Font(name="Calibri", size=11)
    body_align = Alignment(horizontal="center", vertical="center")
    notes_align = Alignment(
        horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True
    )
    title_align = Alignment(
        horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True
    )
    for block_rows in (range(6, 16), range(17, 27), range(29, 44)):
        for i, r in enumerate(block_rows):
            fill = stripe_fill if i % 2 == 1 else no_fill
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.font = body_font
                if c == notes_col:
                    cell.alignment = notes_align
                elif title_col is not None and c == title_col:
                    cell.alignment = title_align
                else:
                    cell.alignment = body_align

    # 6. Header row 5 — bold + light-gray fill so it stands out from data.
    header_fill = PatternFill(
        start_color="FFE7ECF1", end_color="FFE7ECF1", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FF1F4E78")
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    for c in range(1, max_col + 1):
        cell = ws.cell(row=5, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[5].height = 32

    # 7. Apply UNIFORM thin borders to every cell in the data blocks +
    # header row. Chad 2026-05-14: "the spreadsheet is not clean. There
    # are a lot of cells that don't have outlines." The template's
    # original borders were applied unevenly (col O bordered, col P not,
    # etc.) — overwrite all data cells with consistent light-gray thin
    # borders so the table reads as one uniform grid.
    thin = Side(style="thin", color="FFBFBFBF")
    uniform = Border(left=thin, right=thin, top=thin, bottom=thin)
    for block_rows in (range(5, 16), range(17, 27), range(29, 44)):
        # range(5,16) intentionally includes header row 5 + data 6-15.
        for r in block_rows:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = uniform

    extras = (
        "collapsed Shot 1-10 cols into [+] outline group, "
        if collapse_cols else ""
    )
    fixes.append(
        f"{sheet_label}: hid gridlines, {extras}tightened column widths, "
        "stronger navy section banners, alternating row stripes, bold header row"
    )
    return fixes


def _polish_ballistics_callout(wb):
    """Make the Ballistics A4 'fill in your DOPE' callout visually prominent.

    Was: faint gray text strip, easy to miss.
    Now: amber/orange accent banner with clear instruction, bigger height,
    bold white text. Anchors users' attention above the DOPE table so they
    know elevation + wind columns are theirs to fill at the range.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    fixes = []
    if "Ballistics" not in wb.sheetnames:
        return fixes
    ws = wb["Ballistics"]

    # Update text — clearer wording per Chad 2026-05-14 ("after range
    # session" framing instead of ambiguous "AT THE RANGE").
    ws["A4"].value = (
        "📍  After your range session — type the elevation and wind values "
        "you dialed at each distance into the columns below. Click counts "
        "fill in automatically."
    )
    # Banner styling: accent (orange) fill, bold white text, taller row
    ws["A4"].fill = PatternFill(
        start_color="FFD97706", end_color="FFD97706", fill_type="solid"
    )
    ws["A4"].font = Font(
        name="Calibri", size=12, bold=True, color="FFFFFFFF"
    )
    ws["A4"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[4].height = 36

    # Row 7 sub-hint right above the DOPE table header (row 8) — second
    # anchor so users who skip the top banner still see it. Subtle italic
    # gray, not loud — A4 is the loud one.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == rng.max_row == 7:
            ws.unmerge_cells(str(rng))
    ws.merge_cells("A7:K7")
    ws["A7"].value = (
        "↓  Type your dialed Mils or MOA into the columns below — "
        "click counts auto-fill in the gray columns next to each."
    )
    ws["A7"].fill = PatternFill(
        start_color="FF42454C", end_color="FF42454C", fill_type="solid"
    )
    ws["A7"].font = Font(
        name="Calibri", size=10, italic=True, color="FFB8BEC9"
    )
    ws["A7"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[7].height = 22

    # Auto-hide the non-matching click columns based on scope type
    # (Chad 2026-05-14): a Mil-turret scope CAN'T dial MOA clicks, so
    # showing the empty MOA columns is just visual noise. Read the user's
    # scope click value from Load Log G7 and hide the irrelevant columns.
    #   0.1 Mil / 0.05 Mil scope  → hide MOA cols (D, E, H, I)
    #   1/4 MOA / 1/8 MOA scope   → hide Mil cols (B, C, F, G)
    #   N/A or unset              → keep both visible (user can decide)
    g7 = ""
    if "Load Log" in wb.sheetnames:
        g7_val = wb["Load Log"]["G7"].value
        g7 = str(g7_val).strip() if g7_val is not None else ""
    if g7 in ("0.1 Mil", "0.05 Mil"):
        for col in ("D", "E", "H", "I"):
            ws.column_dimensions[col].hidden = True
        hidden_msg = "hid MOA columns (D/E/H/I) — scope is Mil-turret"
    elif g7 in ("1/4 MOA", "1/8 MOA"):
        for col in ("B", "C", "F", "G"):
            ws.column_dimensions[col].hidden = True
        hidden_msg = "hid Mil columns (B/C/F/G) — scope is MOA-turret"
    else:
        hidden_msg = (
            f"both Mil + MOA columns visible (scope = {g7!r}, "
            "no auto-hide rule)"
        )

    # Clarify column headers (Chad 2026-05-14): "Mils — Elev" was
    # ambiguous about whether the user enters their TURRET READING or
    # something Loadscope computed. Relabel so the input vs auto-fill
    # distinction is obvious at a glance.
    header_relabels = {
        "B8": "Mils Dialed",
        "C8": "Mil Clicks (auto)",
        "D8": "MOA Dialed",
        "E8": "MOA Clicks (auto)",
        "F8": "Wind Hold Mils\n(per 10 mph wind)",
        "G8": "Wind Mil Clicks (auto)",
        "H8": "Wind Hold MOA\n(per 10 mph wind)",
        "I8": "Wind MOA Clicks (auto)",
    }
    for coord, text in header_relabels.items():
        ws[coord].value = text
        # Wrap text + center so the longer headers display cleanly
        existing = ws[coord].alignment
        ws[coord].alignment = Alignment(
            horizontal=existing.horizontal or "center",
            vertical=existing.vertical or "center",
            wrap_text=True,
        )
    # Bump row 8 height so wrapped 2-line headers fit without clipping
    if (ws.row_dimensions[8].height or 0) < 36:
        ws.row_dimensions[8].height = 36
    # Widen the dial-value columns slightly so "MOA Dialed" / "Mils Dialed"
    # fit on one line (default ~9 chars; need ~12).
    for col in ("B", "D"):
        if (ws.column_dimensions[col].width or 0) < 12:
            ws.column_dimensions[col].width = 12

    fixes.append(
        f"Ballistics!A4: accent banner + A7: sub-hint + {hidden_msg} + "
        "row 8 headers relabeled (Dialed vs auto-clicks distinction)"
    )
    return fixes


def _polish_start_here_tab(wb):
    """Rebuild the After Range Day → Start Here tab as a clean, friendly
    onboarding page. Chad 2026-05-14 asked for major restructure: simpler
    organization, friendlier font (Avenir Next instead of Calibri), and
    polished formatting. The tab is the first one users see — its job is
    orientation, not reference documentation.

    Sections:
      1. Welcome / title
      2. How to import your range data (3 steps)
      3. What's on each tab (visual map)
      4. Starting a new load cycle (quick steps)
      5. Labeling your CSVs (label format reference)
      6. Troubleshooting (common issues)
      7. Footer (support contact)
    """
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    fixes = []

    # Find the sheet (handle both pre-rename and post-rename names)
    sheet_name = None
    for candidate in ("After Range Day", "Start Here"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        return fixes

    ws = wb[sheet_name]
    # Rename if needed
    if sheet_name == "After Range Day":
        ws.title = "Start Here"

    # Clear existing content + merged cells in the working area
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= 50:
            ws.unmerge_cells(str(rng))
    for r in range(1, 60):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font()
            cell.alignment = Alignment()
            cell.border = Border()
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    # Hide gridlines for a cleaner page look
    if ws.sheet_view is not None:
        ws.sheet_view.showGridLines = False

    # Set column widths — single-column layout, B-H wide enough for body text
    ws.column_dimensions["A"].width = 4   # left margin
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["I"].width = 4   # right margin

    # Style tokens — Avenir Next where available, falls back gracefully
    FONT_FAMILY = "Avenir Next"
    NAVY = "FF1F4E78"
    ACCENT = "FFD97706"
    TEXT_BODY = "FF2D3036"
    TEXT_MUTED = "FF6F7682"
    HEADER_FILL_LIGHT = "FFF5F7FA"

    title_font = Font(name=FONT_FAMILY, size=26, bold=True, color=NAVY)
    subtitle_font = Font(name=FONT_FAMILY, size=13, color=TEXT_MUTED)
    h2_font = Font(name=FONT_FAMILY, size=15, bold=True, color=NAVY)
    body_font = Font(name=FONT_FAMILY, size=12, color=TEXT_BODY)
    body_bold = Font(name=FONT_FAMILY, size=12, bold=True, color=TEXT_BODY)
    accent_font = Font(name=FONT_FAMILY, size=12, bold=True, color=ACCENT)
    footer_font = Font(name=FONT_FAMILY, size=11, italic=True, color=TEXT_MUTED)

    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    title_align = Alignment(horizontal="left", vertical="center", indent=1)
    h2_align = Alignment(horizontal="left", vertical="bottom", indent=1)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_underline = Border(
        bottom=Side(style="thin", color="FFD97706")
    )

    def set_cell(coord, value, font=body_font, align=left_align,
                 fill=None, border=None):
        ws[coord].value = value
        ws[coord].font = font
        ws[coord].alignment = align
        if fill is not None:
            ws[coord].fill = fill
        if border is not None:
            ws[coord].border = border

    def merge_row(row, value, font=body_font, align=left_align,
                  height=None, fill=None, border=None):
        cell_ref = f"B{row}"
        ws.merge_cells(f"B{row}:H{row}")
        set_cell(cell_ref, value, font=font, align=align,
                 fill=fill, border=border)
        # Borders on merged cells only render where the cell has the
        # border attribute. Apply to ALL cells in the merge so the
        # accent underline spans full width (Chad 2026-05-14: "the
        # orange underline doesn't cover the entire sentence").
        if border is not None:
            for col in "BCDEFGH":
                ws[f"{col}{row}"].border = border
        if height is not None:
            ws.row_dimensions[row].height = height

    # ---- 1. TITLE + TAGLINE ----
    merge_row(2, "Start Here", font=title_font, align=title_align,
              height=42)
    merge_row(3, "Welcome to Loadscope™. Here's how to find your best load "
                 "from a range trip.",
              font=subtitle_font, align=left_align, height=24)

    # ---- 2. HOW TO LABEL YOUR CSVs ----  (rewritten 2026-05-14 per Chad:
    # column headers + where-to-type-it guidance + per-part breakdown.)
    merge_row(5, "How to label your CSVs",
              font=h2_font, align=h2_align, height=28,
              border=section_underline)
    merge_row(6,
              "Type the label as your session name in Garmin ShotView, "
              "your group name in BallisticX, or by renaming the CSV file "
              "after export. Loadscope reads it from any of those.",
              font=body_font, align=left_align, height=42)
    merge_row(7,
              "Each label has 3 parts, separated by spaces:   "
              "[ test tag ]   [ grains of powder ]   [ powder type ]",
              font=body_bold, align=left_align, height=28)

    # Column headers above the examples — Chad 2026-05-14 wanted it
    # explicit that the left column is the LABEL and the right column
    # is the BREAKDOWN.
    table_header_fill = PatternFill(
        start_color="FFE7ECF1", end_color="FFE7ECF1", fill_type="solid"
    )
    table_header_font = Font(
        name=FONT_FAMILY, size=10, bold=True, color="FF1F4E78",
    )
    ws.row_dimensions[8].height = 22
    ws["B8"].value = "Example label"
    ws["B8"].font = table_header_font
    ws["B8"].fill = table_header_fill
    ws["B8"].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells("B8:C8")
    ws["D8"].value = "What each part means"
    ws["D8"].font = table_header_font
    ws["D8"].fill = table_header_fill
    ws["D8"].alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )
    ws.merge_cells("D8:H8")

    label_examples = [
        ("P1 45.5 H4350",
         "P1     →  Powder Ladder, Load #1\n"
         "45.5   →  45.5 grains of powder\n"
         "H4350  →  Powder type (Hodgdon H4350)"),
        ("S7 0.070 H4350",
         "S7     →  Seating Depth, Test #7\n"
         "0.070  →  0.070-inch jump distance\n"
         "H4350  →  Powder type (Hodgdon H4350)"),
    ]
    code_fill = PatternFill(start_color=HEADER_FILL_LIGHT,
                            end_color=HEADER_FILL_LIGHT, fill_type="solid")
    code_font = Font(name="SF Mono", size=12, bold=True, color=NAVY)
    breakdown_align = Alignment(
        horizontal="left", vertical="center", wrap_text=True, indent=1,
    )
    breakdown_font = Font(name="SF Mono", size=11, color=TEXT_BODY)
    row = 9  # data rows start after header row 8
    for label, breakdown in label_examples:
        ws.row_dimensions[row].height = 64  # 3 lines × ~18 + padding
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font = code_font
        ws[f"B{row}"].fill = code_fill
        ws[f"B{row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:H{row}")
        ws[f"D{row}"].value = breakdown
        ws[f"D{row}"].font = breakdown_font
        ws[f"D{row}"].alignment = breakdown_align
        ws[f"D{row}"].fill = code_fill  # match left side for visual unity
        row += 1

    # ---- 3. HOW TO IMPORT ----
    merge_row(row + 1, "How to import your range data",
              font=h2_font, align=h2_align, height=28,
              border=section_underline)
    row += 2
    steps = [
        ("①",
         "Drop your properly labeled Garmin Xero and BallisticX CSV files "
         "onto Loadscope — either onto the app icon in your Dock, or onto "
         "the drop zone in the Loadscope window."),
        ("②",
         "Click Run Import. Loadscope sorts the files into the right "
         "folders and writes the data into your workbook."),
        ("③",
         "Your workbook opens in Excel with every row filled in. Visit "
         "the Charts tab to see your suggested winner."),
    ]
    for marker, text in steps:
        ws.row_dimensions[row].height = 42
        ws[f"B{row}"].value = marker
        ws[f"B{row}"].font = accent_font
        ws[f"B{row}"].alignment = Alignment(
            horizontal="center", vertical="top"
        )
        ws.merge_cells(f"C{row}:H{row}")
        set_cell(f"C{row}", text, font=body_font, align=Alignment(
            horizontal="left", vertical="top", wrap_text=True, indent=1,
        ))
        row += 1

    # ---- 4. WHAT'S ON EACH TAB ----  (alignment fix per Chad 2026-05-14:
    # tab name on the left was vertical=top but description was vertical=
    # center, so they didn't line up when descriptions wrapped. Both now
    # vertical=top + per-row height computed from description length so
    # nothing clips.)
    merge_row(row + 1, "What's on each tab",
              font=h2_font, align=h2_align, height=28,
              border=section_underline)
    row += 2
    tab_descriptions = [
        ("Load Log",
         "Your powder ladder. Date, powder, bullet, and cartridge auto-fill "
         "from your CSVs. You fill in your rifle setup once (rifle, scope, "
         "click size, primer, brass, CBTO, OAL) — those values flow into "
         "Seating Depth, Ballistics, and your Pocket Range Card automatically. "
         "The orange button jumps you to Charts."),
        ("Seating Depth",
         "Your seating-depth test data. Same idea as Load Log, one row per jump distance."),
        ("Charts",
         "Your suggested best load. Loadscope ranks every powder ladder candidate by composite score. Save the winner to your Library with one click."),
        ("Ballistics",
         "Your DOPE table. Type in the elevation and wind values you dialed at the range — click counts auto-fill. Print a 4×6 Pocket Range Card from the orange button."),
        ("Load Library",
         "Every confirmed load you've saved over time, one row per load."),
        ("Garmin Xero & BallisticX Import",
         "Auto-populated from your CSVs. You usually won't need to edit these — they feed Load Log + Seating Depth."),
    ]
    top_align_indent = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=1,
    )
    top_align_name = Alignment(
        horizontal="left", vertical="top", indent=1,
    )
    # Approximate chars per line for desc col (C:H merged ≈ 84 chars wide
    # at width=14). Each wrapped line ~18px. Min height 28.
    CHARS_PER_LINE = 80
    LINE_HEIGHT = 18
    for tab_name, desc in tab_descriptions:
        n_lines = max(1, (len(desc) + CHARS_PER_LINE - 1) // CHARS_PER_LINE)
        ws.row_dimensions[row].height = max(28, n_lines * LINE_HEIGHT + 4)
        ws[f"B{row}"].value = tab_name
        ws[f"B{row}"].font = body_bold
        ws[f"B{row}"].alignment = top_align_name
        ws.merge_cells(f"C{row}:H{row}")
        set_cell(f"C{row}", desc, font=body_font, align=top_align_indent)
        row += 1

    # ---- 5. NEW LOAD CYCLE ----
    merge_row(row + 1, "Starting a new load cycle (different cartridge, bullet, or powder)",
              font=h2_font, align=h2_align, height=28,
              border=section_underline)
    row += 2
    merge_row(row,
              "In Loadscope, click  Workbook → Start New Cycle…  and give "
              "it a name like \"6.5 CM 140 ELD-M H4350.\" Loadscope archives "
              "your finished workbook and CSVs, then makes a fresh workbook "
              "your active one. From there, drop your new CSVs in and click "
              "Run Import.",
              font=body_font, align=left_align, height=72)
    row += 1

    # ---- 6. TROUBLESHOOTING ----
    merge_row(row + 1, "Something not working?",
              font=h2_font, align=h2_align, height=28,
              border=section_underline)
    row += 2
    issues = [
        ("Excel says 'Protected View'",
         "macOS marks files from email or AirDrop as untrusted. Click the yellow Enable Editing bar at the top of the workbook."),
        ("A row stays blank",
         "The first word of the Garmin label or BallisticX file name must match the row's tag (P1, P2, S1, etc.)."),
        ("Charge or Jump value is blank",
         "The label was missing the number or had a typo. Open the Garmin Xero Import or BallisticX Import tab and fix it directly."),
        ("Run Import won't go",
         "Excel is probably open. Close Excel (Cmd+Q) and click Run Import again."),
        ("Click counts on Ballistics look wrong",
         "Check the Click: dropdown on the Load Log tab and pick the right one for your scope."),
        ("Charts tab is empty",
         "Loadscope needs at least 2 candidate loads with BOTH chronograph AND target data to score them. Drop more CSVs and re-run import."),
        ("Print Pocket Range Card button does nothing",
         "Make sure Loadscope is in /Applications and you've opened it once. The button uses a loadscope:// URL that needs the app registered."),
    ]
    # Same alignment + dynamic-height fix as "What's on each tab" (Chad
    # 2026-05-14): troubleshoot description was vertical=center while
    # title was vertical=top → didn't line up when description wrapped.
    # Now both vertical=top + per-row height computed from fix_text len.
    fix_top_align = Alignment(
        horizontal="left", vertical="top", wrap_text=True, indent=1,
    )
    title_top_align = Alignment(
        horizontal="left", vertical="top", indent=1, wrap_text=True,
    )
    CHARS_PER_LINE_TS = 60  # narrower since D col is C-merged with B is the title
    LINE_HEIGHT_TS = 18
    for problem, fix_text in issues:
        n_lines = max(2, (len(fix_text) + CHARS_PER_LINE_TS - 1) // CHARS_PER_LINE_TS)
        ws.row_dimensions[row].height = n_lines * LINE_HEIGHT_TS + 6
        ws[f"B{row}"].value = problem
        ws[f"B{row}"].font = body_bold
        ws[f"B{row}"].alignment = title_top_align
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:H{row}")
        ws[f"D{row}"].value = fix_text
        ws[f"D{row}"].font = body_font
        ws[f"D{row}"].alignment = fix_top_align
        row += 1

    # ---- 7. FOOTER ----
    row += 1
    merge_row(row,
              "Need more help?  See  Loadscope — Quick Start.docx  in your "
              "project folder   ·   🖨  Cmd+P to print this page.",
              font=footer_font, align=Alignment(
                  horizontal="center", vertical="center", wrap_text=True
              ),
              height=24)
    # Second footer row: clickable mailto: link so users can email support
    # with one click instead of copying the address (Chad 2026-05-14).
    row += 1
    ws.merge_cells(f"B{row}:H{row}")
    ws[f"B{row}"].value = "📧  Email us: support@loadscope.app"
    ws[f"B{row}"].font = Font(
        name=FONT_FAMILY, size=11, bold=True, color=ACCENT, underline="single"
    )
    ws[f"B{row}"].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws[f"B{row}"].hyperlink = (
        "mailto:support@loadscope.app?subject=Loadscope%20support%20request"
    )
    ws.row_dimensions[row].height = 26

    # Print setup — fit-to-1-page-wide so the page prints cleanly on letter
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    fixes.append(
        f"{sheet_name} → Start Here: rebuilt as 6-section onboarding page "
        "(Avenir Next, navy section headers, accent step badges)"
    )
    return fixes


def _enable_shrink_to_fit_on_numeric_cells(wb):
    """Set shrink_to_fit=True on every numeric/formula cell across visible
    tabs. When a column is too narrow to display the value, Excel will
    auto-shrink the font instead of showing #####. Chad 2026-05-14:
    "I don't ever want hashtags to show up and have a user have to go
    in and format themselves."

    Skips text cells (so narrative paragraphs aren't shrunken) and cells
    with wrap_text already enabled (mutually exclusive in Excel).
    """
    from openpyxl.styles import Alignment
    fixes = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        modified = 0
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                is_numeric = isinstance(v, (int, float))
                is_formula = isinstance(v, str) and v.startswith("=")
                if not (is_numeric or is_formula):
                    continue
                existing = cell.alignment
                # wrap_text and shrink_to_fit are mutually exclusive in
                # Excel. Skip cells that have wrap_text on; their content
                # is multi-line and not the narrow-numeric case anyway.
                if existing.wrap_text:
                    continue
                if existing.shrink_to_fit:
                    continue
                cell.alignment = Alignment(
                    horizontal=existing.horizontal,
                    vertical=existing.vertical,
                    text_rotation=existing.text_rotation,
                    wrap_text=False,
                    shrink_to_fit=True,
                    indent=existing.indent,
                )
                modified += 1
        if modified:
            fixes.append(
                f"{ws.title}: shrink-to-fit on {modified} numeric/formula cells"
            )
    return fixes


def _protect_workbook(wb):
    """Lock formula + computed-static cells so users can't accidentally
    overwrite them, while keeping all legitimate input cells editable.
    Chad 2026-05-14: "I would like it to be impossible for a user to
    overwrite any formula."

    Strategy: per-tab whitelist of UNLOCKED ranges. Everything else gets
    locked. Sheet protection enabled with no password. Garmin Xero +
    BallisticX Import tabs stay UNPROTECTED so users can fix CSV typos
    by overwriting the formula cells with their own values.
    """
    from openpyxl.styles import Protection
    fixes = []

    # Per-tab whitelist of user-input cell coordinates. Anything not in
    # the list is locked. Tabs missing from this dict get NO protection.
    UNLOCKED = {
        # Start Here = pure docs, no editable cells. Empty list ⇒ all locked.
        "Start Here": [],
        # Load Log + Seating Depth: rifle/scope/cartridge metadata + click
        # dropdown + bullet/primer/brass + CBTO/OAL + date/temp/notes.
        "Load Log": [
            "B5", "G5", "L5",          # Rifle, Shooter, Cartridge
            "B6", "G6", "L6",          # Barrel, Optic, Chrono
            "G7",                       # Turret dropdown
            "B9", "L9", "O9",          # Bullet, Primer, Brass
            "B10", "G10", "L10",       # CBTO, OAL, Distance
            "B13", "G13", "K13", "L13", # Date, ?, Temp, Notes
        ],
        "Seating Depth": [
            "B5", "G5", "L5",
            "B6", "G6", "L6",
            "G7",
            "B9", "L9", "O9",
            "B10", "G10", "L10",       # B10 = winning charge (smart formula)
            "B13", "G13", "K13", "L13",
        ],
        # Charts: composite weights are user-tunable; B100 is the test/
        # window mode toggle. Everything else (analysis grid, winner
        # cells, sorted helpers) is computed → stays locked.
        "Charts": [
            "C11", "E11", "G11", "I11",  # Composite weights
            "B100",                       # mode toggle
        ],
        # Ballistics: rifle/bullet/charge/vel + zero/sight-ht/twist
        # metadata, plus the user-fill DOPE input columns.
        "Ballistics": (
            ["B5", "E5", "H5", "K5",
             "B6", "E6", "H6", "K6"]
            + [f"B{r}" for r in range(9, 19)]   # Mils Dialed
            + [f"D{r}" for r in range(9, 19)]   # MOA Dialed
            + [f"F{r}" for r in range(9, 19)]   # Wind Hold Mils
            + [f"H{r}" for r in range(9, 19)]   # Wind Hold MOA
            + [f"K{r}" for r in range(9, 19)]   # Notes
        ),
        # Load Library: rows 5-19 (cols B-Q) are user-editable history.
        "Load Library": [
            f"{c}{r}" for c in "BCDEFGHIJKLMNOPQ" for r in range(5, 20)
        ],
        # Garmin Xero + BallisticX Import: deliberately NOT in this dict.
        # Import tabs stay fully unprotected so users can overwrite the
        # formula cells with corrected values when fixing CSV typos.
    }

    locked_default = Protection(locked=True, hidden=False)
    unlocked = Protection(locked=False, hidden=False)

    for sheet_name, unlock_list in UNLOCKED.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Lock every cell first (Excel's default — but explicit so we don't
        # depend on per-cell template defaults).
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = locked_default
        # Unlock the whitelist
        for coord in unlock_list:
            ws[coord].protection = unlocked
        # Enable sheet protection (no password — default; openpyxl
        # rejects setting password=None explicitly). Allow common
        # non-destructive actions.
        ws.protection.sheet = True
        ws.protection.sort = False           # allow sort
        ws.protection.autoFilter = False     # allow filter
        ws.protection.formatCells = False    # allow format changes
        ws.protection.formatColumns = False  # allow column resize
        ws.protection.formatRows = False     # allow row resize
        ws.protection.selectLockedCells = False    # allow click locked cells
        ws.protection.selectUnlockedCells = False  # allow click unlocked cells
        fixes.append(
            f"{sheet_name}: protected (sheet protection on, "
            f"{len(unlock_list)} input cells unlocked)"
        )

    # Explicitly DISABLE protection on the Import tabs (template may have
    # left it on from prior versions; we want users free to fix CSV typos).
    for sheet_name in ("Garmin Xero Import", "BallisticX Import"):
        if sheet_name in wb.sheetnames:
            wb[sheet_name].protection.sheet = False

    return fixes


def _polish_load_library(wb):
    """Widen Load Library Load Name column (C) so long composite names like
    "6.5 CM 140 Berger Hybrid / H4350 42.1 / 0.020 jump" (50 chars) display
    without being cut off. Width 25 → 32 + wrap_text on data rows.
    Chad reported cells C5/C6/C7 cutting off 2026-05-14.
    """
    from openpyxl.styles import Alignment
    fixes = []
    if "Load Library" not in wb.sheetnames:
        return fixes
    ws = wb["Load Library"]

    # Widen the Load Name column.
    ws.column_dimensions["C"].width = 32

    # Enable wrap_text on Load Name cells in data rows 5-19, preserve other
    # alignment attrs (horizontal=center came from the template/generator).
    for r in range(5, 20):
        cell = ws.cell(row=r, column=3)
        existing = cell.alignment
        cell.alignment = Alignment(
            horizontal=existing.horizontal or "center",
            vertical=existing.vertical or "center",
            wrap_text=True,
        )
        # Bump row height so wrapped 2-line names display without clipping.
        # Existing data rows already 32; empty rows are 21.75 — bump to 32.
        current_h = ws.row_dimensions[r].height
        if current_h is None or current_h < 32:
            ws.row_dimensions[r].height = 32

    fixes.append(
        "Load Library!C: widened col C 25→32 + wrap_text on rows 5-19 for long Load Names"
    )
    return fixes


def resize_comment_boxes(xlsx_path, width_px=360, height_px=220):
    """Resize every VML comment shape in an .xlsx file to width_px x height_px.

    Excel defaults comment boxes to 144x79 px which clips most Loadscope
    tooltip text (200-700 chars per comment). 360x220 is Chad's v0.13.3
    proven size — fits the longest comments without scrolling.

    Must run AFTER wb.save() since the VML files only exist on disk inside
    the .xlsx zip. Operates on the file in place via tempfile + replace.
    """
    import re
    import shutil
    import tempfile
    import zipfile
    from pathlib import Path
    xlsx_path = Path(xlsx_path)
    pattern = re.compile(r"width:\d+(?:\.\d+)?px;height:\d+(?:\.\d+)?px")
    replacement = f"width:{width_px}px;height:{height_px}px"
    tmpdir = Path(tempfile.mkdtemp())
    new_zip = tmpdir / "out.xlsx"
    resized = 0
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin:
            with zipfile.ZipFile(new_zip, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    data = zin.read(item)
                    if (item.startswith("xl/drawings/commentsDrawing")
                            and item.endswith(".vml")):
                        text = data.decode("utf-8")
                        n = len(pattern.findall(text))
                        if n:
                            text = pattern.sub(replacement, text)
                            resized += n
                            data = text.encode("utf-8")
                    zout.writestr(item, data)
        shutil.copy(new_zip, xlsx_path)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return resized


# Default composite-score weights — used by the Tools → Reset Composite
# Weights menu item to restore the workbook to Loadscope's recommended
# starting values. Rationale lives in app/help_dialog.py and as a cell
# comment on Charts!A10 + Seating Depth!A28.
DEFAULT_WEIGHTS_CHARTS = {
    "B11": 0.3,   # Velocity
    "D11": 0.2,   # SD
    "F11": 0.2,   # MR
    "H11": 0.3,   # SD-Vert
}
DEFAULT_WEIGHTS_SEATING_DEPTH = {
    "C28": 0.15,  # Velocity
    "F28": 0.25,  # SD
    "I28": 0.25,  # MR
    "L28": 0.35,  # SD-Vert
}


def _inches_to_moa(inches, distance_yd):
    """Convert linear inches to MOA at a given distance in yards.

    1 MOA subtends 1.047 inches at 100 yards, scaling linearly with
    distance. Returns None if either input is missing or distance is
    not positive (can't compute MOA without a known shot distance)."""
    if inches in (None, "") or distance_yd in (None, "", 0):
        return None
    try:
        inches_f = float(inches)
        dist_f = float(distance_yd)
        if dist_f <= 0:
            return None
        return round(inches_f * 100.0 / (dist_f * 1.047), 3)
    except (ValueError, TypeError):
        return None


def gather_suggested_load(workbook_path):
    """Read the current 'suggested load' state from the workbook —
    winning charge from the powder ladder analysis, winning jump from
    the seating-depth analysis, plus the load components and
    performance metrics for the winner row(s).

    Returns a dict ready to be written to one row in the Load Library
    sheet. Raises ValueError if no powder-ladder winner exists.

    Reads cached values via data_only=True; the caller should have run
    a CSV import (or Excel session) before calling this so formulas
    have current cached results.
    """
    import datetime
    import os
    import re as _re
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True, keep_vba=False)
    if "Charts" not in wb.sheetnames or "Load Log" not in wb.sheetnames:
        raise ValueError(
            "This workbook doesn't have the expected sheets (Charts + Load Log)."
        )
    charts = wb["Charts"]
    ll = wb["Load Log"]
    sd = wb["Seating Depth"] if "Seating Depth" in wb.sheetnames else None

    winning_charge = charts["B3"].value
    if winning_charge in (None, "", 0):
        raise ValueError(
            "No suggested charge yet — run a powder ladder first."
        )

    # Find which Load Log candidate row corresponds to the winner so
    # we can pull its metrics. Match by charge weight.
    winner_row = None
    for r in range(16, 26):
        v = ll.cell(r, 2).value  # col B = Charge (gr)
        if v == winning_charge:
            winner_row = r
            break
        # tolerate small float drift
        try:
            if v not in (None, "") and abs(float(v) - float(winning_charge)) < 1e-6:
                winner_row = r
                break
        except (ValueError, TypeError):
            pass

    # Distance for MOA conversion lives at L10 on both Load Log + SD
    distance_yd = ll["L10"].value
    if distance_yd in (None, "") and sd is not None:
        distance_yd = sd["L10"].value

    # Performance metrics from the winning Load Log row (inches → MOA).
    # Columns: H=AvgVel, J=SD, K=Group(in), N=MeanRadius(in).
    if winner_row:
        avg_vel = ll.cell(winner_row, 8).value
        sd_fps = ll.cell(winner_row, 10).value
        group_in = ll.cell(winner_row, 11).value
        mr_in = ll.cell(winner_row, 14).value
    else:
        avg_vel = sd_fps = group_in = mr_in = None

    group_moa = _inches_to_moa(group_in, distance_yd)
    mr_moa = _inches_to_moa(mr_in, distance_yd)

    # Optional: seating-depth winner overrides AvgVel/SD with the more
    # refined measurements from the seating-depth test session.
    winning_jump = None
    if sd is not None:
        winning_jump = sd["D2"].value
        if winning_jump not in (None, "", 0):
            sd_avg = sd["G2"].value  # AvgVel of winning jump
            sd_sd = sd["J2"].value   # SD of winning jump
            sd_mr_moa = sd["L2"].value  # MR (MOA) — SD analysis is already in MOA
            if sd_avg not in (None, ""):
                avg_vel = sd_avg
            if sd_sd not in (None, ""):
                sd_fps = sd_sd
            if sd_mr_moa not in (None, ""):
                mr_moa = sd_mr_moa
        else:
            winning_jump = None

    # Load components from Load Log header (top section).
    rifle = ll["B5"].value
    bullet = ll["B9"].value
    powder = ll["G9"].value
    primer = ll["L9"].value
    brass = ll["O9"].value
    cbto = ll["B10"].value
    notes = ll["L13"].value or ""

    # Parse bullet weight from the Bullet field text (e.g. "140 ELD-M (140 gr)").
    bullet_wt = None
    if bullet:
        m = _re.search(r"(\d+(?:\.\d+)?)\s*gr", str(bullet))
        if m:
            try:
                bullet_wt = float(m.group(1))
            except ValueError:
                pass

    load_name = os.path.splitext(os.path.basename(workbook_path))[0]
    date_added = datetime.date.today().strftime("%Y-%m-%d")

    return {
        "date_added": date_added,
        "load_name": load_name,
        "rifle": rifle,
        "bullet": bullet,
        "bullet_wt": bullet_wt,
        "powder": powder,
        "charge": winning_charge,
        "primer": primer,
        "brass": brass,
        "cbto": cbto,
        "jump": winning_jump,
        "avg_vel": avg_vel,
        "sd_fps": sd_fps,
        "group_moa": group_moa,
        "mr_moa": mr_moa,
        "notes": notes,
    }


def save_suggested_load_to_library(workbook_path, data=None):
    """Append a new Load Library row from the current suggested load.

    If `data` is provided, use it directly (lets a confirmation dialog
    pass an edited version). Otherwise call `gather_suggested_load` to
    build the default row from the workbook's current state.

    Returns (row_number, data_dict). Raises ValueError on any of:
      - no suggested charge yet (gather_suggested_load raised)
      - Load Library sheet missing
      - Load Library full (no empty rows in the 5-19 data range)
    """
    from openpyxl import load_workbook

    if data is None:
        data = gather_suggested_load(workbook_path)

    wb = load_workbook(workbook_path, keep_vba=False)
    wb.template = False
    if "Load Library" not in wb.sheetnames:
        raise ValueError("This workbook doesn't have a Load Library sheet.")
    library = wb["Load Library"]

    # Find next empty row in 5-19 (data rows). Detect "empty" by the
    # # column (A) being blank.
    target_row = None
    for r in range(5, 20):
        if library.cell(r, 1).value in (None, ""):
            target_row = r
            break
    if target_row is None:
        raise ValueError(
            "Load Library is full (rows 5-19 all in use). "
            "Delete or archive a row before saving a new load."
        )

    # Write the row. Column → key mapping mirrors the Load Library
    # header row 4 layout (see workbook for the canonical truth).
    library.cell(target_row, 1).value = target_row - 4  # A: sequential #
    library.cell(target_row, 2).value = data.get("date_added")
    library.cell(target_row, 3).value = data.get("load_name")
    library.cell(target_row, 4).value = data.get("rifle")
    library.cell(target_row, 5).value = data.get("bullet")
    library.cell(target_row, 6).value = data.get("bullet_wt")
    library.cell(target_row, 7).value = data.get("powder")
    library.cell(target_row, 8).value = data.get("charge")
    library.cell(target_row, 9).value = data.get("primer")
    library.cell(target_row, 10).value = data.get("brass")
    library.cell(target_row, 11).value = data.get("cbto")
    library.cell(target_row, 12).value = data.get("jump")
    library.cell(target_row, 13).value = data.get("avg_vel")
    library.cell(target_row, 14).value = data.get("sd_fps")
    library.cell(target_row, 15).value = data.get("group_moa")
    library.cell(target_row, 16).value = data.get("mr_moa")
    library.cell(target_row, 17).value = data.get("notes")

    wb.save(workbook_path)
    return target_row, data


def reset_composite_weights(workbook_path):
    """Restore Charts and Seating Depth composite-score weights to the
    Loadscope defaults. Returns a list of (sheet, cell, old, new) tuples
    describing what changed."""
    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, keep_vba=False)
    wb.template = False
    changes = []
    if "Charts" in wb.sheetnames:
        ws = wb["Charts"]
        for cell_ref, value in DEFAULT_WEIGHTS_CHARTS.items():
            old = ws[cell_ref].value
            ws[cell_ref] = value
            changes.append(("Charts", cell_ref, old, value))
    if "Seating Depth" in wb.sheetnames:
        ws = wb["Seating Depth"]
        for cell_ref, value in DEFAULT_WEIGHTS_SEATING_DEPTH.items():
            old = ws[cell_ref].value
            ws[cell_ref] = value
            changes.append(("Seating Depth", cell_ref, old, value))
    wb.save(workbook_path)
    return changes


def write_group_records(wb, bx_records):
    """Write all target-group records to the BallisticXGroups hidden sheet.

    Sheet name is historical — it now holds group analysis data from any
    target-analysis app, with Source indicating the originating parser.
    """
    if "BallisticXGroups" not in wb.sheetnames:
        print("  ERROR: workbook has no BallisticXGroups sheet")
        return 0
    sht = wb["BallisticXGroups"]
    for row in sht.iter_rows(min_row=2, max_row=sht.max_row):
        for cell in row:
            cell.value = None

    out_row = 2
    for g in bx_records:
        sht.cell(row=out_row, column=1).value = g.get("Tag", "")
        sht.cell(row=out_row, column=2).value = g.get("ChargeOrJump")
        sht.cell(row=out_row, column=3).value = g.get("Powder", "")
        sht.cell(row=out_row, column=4).value = g.get("Date", "")
        sht.cell(row=out_row, column=5).value = g.get("Distance")
        sht.cell(row=out_row, column=6).value = g.get("Caliber", "")
        sht.cell(row=out_row, column=7).value = g.get("GroupIn")
        sht.cell(row=out_row, column=8).value = g.get("WidthIn")
        sht.cell(row=out_row, column=9).value = g.get("HeightIn")
        sht.cell(row=out_row, column=10).value = g.get("MRIn")
        sht.cell(row=out_row, column=11).value = g.get("CEPIn")
        sht.cell(row=out_row, column=12).value = g.get("SDRadIn")
        sht.cell(row=out_row, column=13).value = g.get("SDVertIn")
        sht.cell(row=out_row, column=14).value = g.get("SDHorizIn")
        sht.cell(row=out_row, column=15).value = g.get("ElevOffsetIn")
        sht.cell(row=out_row, column=16).value = g.get("WindOffsetIn")
        sht.cell(row=out_row, column=17).value = g.get("Label", "")
        out_row += 1
    return out_row - 2


# Backwards-compat alias for any callers still using the old name
write_ballisticx = write_group_records


def run_import(workbook_path, project_dir=None, open_excel=True):
    """Read all CSVs in the import folders and write them into the workbook.

    Designed to be called from both the CLI (main()) and from the GUI app.
    Prints progress to stdout — capture with contextlib.redirect_stdout if you
    want to route the output somewhere other than the Terminal.

    Returns a dict:
        {
            "ok": True/False,
            "workbook": <path>,
            "garmin_rows": <int>,
            "ballisticx_rows": <int>,
            "error": <str or None>,
            "safety_stop": <bool>,   # True if no CSVs found and we did nothing
        }
    """
    if project_dir is None:
        project_dir = PROJECT

    print(f"\nWorkbook: {os.path.basename(workbook_path)}")

    # Walk every registered parser and pull records from its import folder.
    # Each file in a parser's folder is checked with that parser's detect() —
    # this means if you accidentally drop a Garmin CSV in BallisticX Imports/,
    # we won't try to parse it as a BallisticX file (it'll just be skipped
    # and you'll see a warning).
    chronograph_records = []
    group_records = []

    for parser in ALL_PARSERS:
        folder = os.path.join(project_dir, parser.IMPORT_FOLDER)
        print(f"\nReading {parser.NAME} CSVs from: {folder}")
        if not os.path.isdir(folder):
            print(f"  (folder doesn't exist — skipping)")
            continue
        for fn in sorted(os.listdir(folder)):
            if not fn.lower().endswith(".csv"):
                continue
            full = os.path.join(folder, fn)
            try:
                if not parser.detect(full):
                    print(f"  skip: {fn}  (doesn't look like a {parser.NAME} CSV)")
                    continue
                result = parser.parse(full)
            except Exception as e:
                print(f"  WARNING: {parser.NAME} parser failed on {fn}: {e}")
                continue

            if result is None:
                continue
            if isinstance(result, list):
                # Group parser — list of records
                group_records.extend(result)
                print(f"  parsed: {fn}  →  {len(result)} group(s)")
            else:
                # Chronograph parser — single record. Skip records that have
                # neither shots nor an average velocity — these would otherwise
                # write a Tag-only row to GarminSessions and shove a real row's
                # data into the wrong cell when re-imported.
                shots_count = len(result.get("Shots") or [])
                if shots_count == 0 and result.get("AvgVel") in (None, ""):
                    print(f"  skip: {fn}  (no shots — empty session)")
                    continue
                chronograph_records.append(result)
                print(
                    f"  parsed: {fn}  →  Tag={result.get('Tag')!r}, "
                    f"Charge={result.get('ChargeOrJump')}, "
                    f"Powder={result.get('Powder')!r}, {shots_count} shots"
                )

    print(f"\nTotals: {len(chronograph_records)} chronograph session(s), "
          f"{len(group_records)} target group(s)")

    # ---- Data validation (warnings only — non-blocking) ----
    warnings = []
    for d in chronograph_records:
        # We don't have the original filename per-record any more; use Tag/SessionTitle as identifier
        ident = d.get("SessionTitle") or d.get("Tag") or "?"
        _validate_chronograph_record(d, ident, warnings)
    for g in group_records:
        ident = g.get("Label") or g.get("Tag") or "?"
        _validate_group_record(g, ident, warnings)
    warnings += _check_duplicate_tags(chronograph_records, "Garmin/chronograph data")
    warnings += _check_duplicate_tags(group_records, "BallisticX/group data")
    if warnings:
        print(f"\nValidation warnings ({len(warnings)}):")
        for w in warnings:
            print(w)
        print("(Data was still imported — review the warnings above and your "
              "source CSVs to confirm the values are correct.)")

    # SAFETY CHECK — if no records anywhere, refuse to wipe the workbook.
    if not chronograph_records and not group_records:
        print("\n" + "=" * 60)
        print("SAFETY STOP — no CSVs found in either import folder.")
        print("=" * 60)
        print("\nNothing was written to the workbook. The existing data is untouched.")
        if open_excel:
            print("\nOpening the workbook anyway so you can review it.")
            subprocess.run(["open", workbook_path])
        return {
            "ok": True,
            "workbook": workbook_path,
            "garmin_rows": 0,
            "ballisticx_rows": 0,
            "error": None,
            "safety_stop": True,
        }

    # Backup the workbook before we modify it. If anything goes wrong (CSV
    # parse error mid-import, openpyxl bug, power outage), the user can
    # recover from the most recent backup. Retention count comes from settings.
    backup_keep = _read_backup_keep_setting(default=5)
    if backup_keep > 0:
        _rotate_workbook_backups(workbook_path, keep=backup_keep)
    else:
        print("  (Backups disabled in settings — skipping pre-import backup.)")

    # Open workbook and write data
    print(f"\nWriting to workbook…")
    try:
        wb = load_workbook(workbook_path, keep_vba=False)
    except Exception as e:
        msg = f"Couldn't open workbook: {e}. Is it open in Excel? Close it and try again."
        print(f"  ERROR: {msg}")
        return {
            "ok": False, "workbook": workbook_path,
            "garmin_rows": 0, "ballisticx_rows": 0,
            "error": msg, "safety_stop": False,
        }

    # Force save as regular workbook (not template)
    wb.template = False

    # Migrate any legacy 7-shot schema to v0.13.3 10-shot BEFORE the
    # writer runs — otherwise the writer's new column targets collide
    # with legacy data positions.
    migrate_schema_to_10_shot(wb)

    n_g = write_chronograph_records(wb, chronograph_records)
    n_b = write_group_records(wb, group_records)
    print(f"  Wrote {n_g} chronograph row(s) and {n_b} target-group row(s)")

    # Auto-repair template-level bugs in older workbooks and auto-fill
    # session metadata from the imported data. Safe to call every time --
    # idempotent and never overwrites user input.
    apply_workbook_repairs(wb, group_records, chronograph_records=chronograph_records)

    try:
        wb.save(workbook_path)
        print(f"  Saved.")
        # Post-save: resize VML comment boxes (Excel defaults them to 144x79
        # which clips most Loadscope tooltip text). Must run after save since
        # the VML files only exist inside the saved .xlsx zip.
        try:
            n_resized = resize_comment_boxes(workbook_path)
            if n_resized:
                print(f"  Resized {n_resized} comment box(es) to 360x220 px")
        except Exception as e:
            print(f"  (Couldn't resize comment boxes: {e} — not fatal)")
    except PermissionError:
        msg = "Workbook is open in Excel. Close it and try again."
        print(f"  ERROR: {msg}")
        return {
            "ok": False, "workbook": workbook_path,
            "garmin_rows": n_g, "ballisticx_rows": n_b,
            "error": msg, "safety_stop": False,
        }

    # v0.14: detect SD-only case for the auto-prompt UX (Chad 2026-05-14).
    # If user imported seating-depth records but NO powder-ladder records
    # AND Charts!B3 (winning charge) is still empty, Loadscope can't infer
    # what charge they used. main.py reads `needs_sd_charge` and prompts.
    # When the prompt IS needed, defer opening Excel — main.py will write
    # the user's answer to the workbook first, then open it.
    sd_imported = any(
        str(c.get("Tag", "")).startswith("S") for c in chronograph_records
    )
    pl_imported = any(
        str(c.get("Tag", "")).startswith("P") for c in chronograph_records
    )
    needs_sd_charge = False
    try:
        from openpyxl import load_workbook as _lw
        wb_check = _lw(workbook_path, data_only=False)
        if "Charts" in wb_check.sheetnames:
            charts_b3 = wb_check["Charts"]["B3"].value
            charts_b3_empty = charts_b3 in (None, "", 0)
        else:
            charts_b3_empty = True
        needs_sd_charge = (
            sd_imported and not pl_imported and charts_b3_empty
        )
    except Exception:
        pass  # detection is best-effort; failure shouldn't block the import

    if open_excel and not needs_sd_charge:
        print(f"\nOpening workbook in Excel…")
        subprocess.run(["open", workbook_path])

    print("\nDONE.")
    return {
        "ok": True, "workbook": workbook_path,
        "garmin_rows": n_g, "ballisticx_rows": n_b,
        "error": None, "safety_stop": False,
        "needs_sd_charge": needs_sd_charge,
    }


def main():
    """CLI entry point — used by 'Import Rifle Data.command'."""
    print("=" * 60)
    print("Rifle Loads — Import Data")
    print("=" * 60)

    workbook_path = find_workbook()
    if not workbook_path:
        print("\nERROR: No working .xlsx workbook found in:")
        print(f"  {PROJECT}")
        print("\nMake sure you have a saved working file (e.g., 'My Load 6.5CM.xlsx').")
        print("(The .xltx template is excluded — make a working copy via File > Save As first.)")
        sys.exit(1)

    result = run_import(workbook_path)
    print("=" * 60)
    if not result["ok"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
