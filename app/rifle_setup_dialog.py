"""Native "Rifle & Setup" form — UI-change Phase 1 ([[loadscope-appify-scoping]]).

The first slice of bringing the clean Loadscope format to the REAL
workbook: a themed form for the hand-typed identity/setup data, so the
user edits it in Loadscope instead of hunting cells in the Excel grid.
It reads from and writes back into the workbook's Load Log header (the
single source of truth — its formulas propagate to the Ballistics
header), exactly the openpyxl pattern apply_workbook_repairs already
uses. Excel stays the calc engine.

The field -> Load Log cell map + the read/write helpers are pure
(no Qt), so they're unit-testable without a display.

Deliberately EXCLUDED from Phase 1 (handled elsewhere / Phase 2):
  - Powder (Load Log!G9) is a FORMULA auto-derived from the Garmin
    import — writing a literal would clobber it.
  - Date (Load Log!B13) is a datetime-typed cell feeding formulas;
    typed-widget editing comes in Phase 2 to avoid format corruption.
"""

import os

try:
    from PyQt5.QtGui import QFont
    from PyQt5.QtWidgets import (
        QDialog,
        QDialogButtonBox,
        QFormLayout,
        QLabel,
        QLineEdit,
        QMessageBox,
        QPushButton,
        QVBoxLayout,
        QWidget,
    )
except ImportError:  # allow import for unit tests without PyQt5
    QDialog = QDialogButtonBox = QFormLayout = QLabel = QLineEdit = None
    QMessageBox = QVBoxLayout = QFont = QWidget = QPushButton = None

from openpyxl import load_workbook

import component_data as _cd
import smart_fields as _sf

# (section, label, Load Log value-cell, kind). kind drives the widget:
#   text|number  -> QLineEdit (number coerces to float on save)
#   turret       -> LockedCombo (exact accepted strings — safeguard)
#   bullet|primer-> CascadeField (manufacturer -> item, + Other)
#   brass|chrono|shooter -> HistoryCombo (curated + remembered history)
#   distance     -> HistoryCombo (presets + free) ; saved numeric
# Verified against the demo workbook's Load Log header layout.
RIFLE_SETUP_FIELDS = [
    ("Rifle & Shooter", "Rifle",     "B5",  "text"),
    ("Rifle & Shooter", "Shooter",   "G5",  "shooter"),
    ("Rifle & Shooter", "Cartridge", "L5",  "text"),
    ("Rifle & Shooter", "Barrel",    "B6",  "text"),
    ("Rifle & Shooter", "Optic",     "G6",  "text"),
    ("Rifle & Shooter", "Chrono",    "L6",  "chrono"),
    ("Rifle & Shooter", "Scope click (turret)", "G7", "turret"),
    ("Load Components", "Bullet",    "B9",  "bullet"),
    ("Load Components", "Primer",    "L9",  "primer"),
    ("Load Components", "Brass",     "O9",  "brass"),
    ("Load Components", "CBTO (in)", "B10", "number"),
    ("Load Components", "OAL (in)",  "G10", "number"),
    ("Load Components", "Distance (yd)", "L10", "distance"),
    ("Test Session",    "Temp (°F)", "G13", "number"),
    ("Test Session",    "Conditions / notes", "L13", "text"),
]

# kinds whose value is written to the cell as a float (keeps the
# workbook's numeric formulas working).
_NUMERIC_KINDS = ("number", "distance")

_SHEET = "Load Log"


def read_rifle_setup(workbook_path):
    """Return {cell: current value as str} for every RIFLE_SETUP_FIELDS
    cell. Pure — no Qt. Empty string for blank cells."""
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    ws = wb[_SHEET]
    out = {}
    for _section, _label, cell, _kind in RIFLE_SETUP_FIELDS:
        v = ws[cell].value
        out[cell] = "" if v is None else str(v)
    return out


def write_rifle_setup(workbook_path, values):
    """Write edited values back to the Load Log header cells.

    `values` is {cell: str}. "number" fields are coerced to float when
    parseable (so they keep feeding the workbook's numeric formulas);
    an unparseable number is skipped rather than corrupting the cell.
    Only cells whose value actually changed are touched. Pure — no Qt.
    Returns the list of changed cells. Raises PermissionError if the
    workbook is open in Excel (caller surfaces a friendly message).
    """
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    ws = wb[_SHEET]
    kinds = {cell: kind for _s, _l, cell, kind in RIFLE_SETUP_FIELDS}
    changed = []
    for cell, raw in values.items():
        if cell not in kinds:
            continue
        new = raw.strip()
        if kinds[cell] in _NUMERIC_KINDS:
            if new == "":
                continue
            try:
                new = float(new)
            except ValueError:
                continue  # don't corrupt a numeric cell with junk
        cur = ws[cell].value
        cur_norm = "" if cur is None else cur
        if new == cur_norm or (isinstance(new, float)
                                and _as_float(cur) == new):
            continue
        ws[cell] = new
        changed.append(cell)
    if changed:
        wb.save(workbook_path)
    return changed


def _as_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


if QWidget is not None:

    class _LineField(QLineEdit):
        """Plain text/number field with the smart-widget interface so
        the panel + save() treat every field uniformly."""

        def value(self):
            return self.text().strip()

        def set_value(self, s):
            self.setText(s or "")

        def commit(self):
            pass

    def _build_field(kind):
        if kind == "turret":
            return _sf.LockedCombo(_cd.turret_clicks())
        if kind == "bullet":
            return _sf.CascadeField("bullet")
        if kind == "primer":
            return _sf.CascadeField("primer")
        if kind == "brass":
            return _sf.HistoryCombo("brass", _cd.brass_options())
        if kind == "chrono":
            return _sf.HistoryCombo("chrono", _cd.chronographs())
        if kind == "shooter":
            return _sf.HistoryCombo("shooter", [])
        if kind == "distance":
            return _sf.HistoryCombo("distance",
                                    ["50", "100", "200", "300"])
        return _LineField()

    class RifleSetupPanel(QWidget):
        """Embeddable Rifle & Setup form (title + intro + fields). No
        dialog chrome — the shell embeds this directly; RifleSetupDialog
        wraps it with Save/Cancel for the menu flow. save() does the
        write via the pure helpers and returns a result tuple so the
        caller decides how to surface success/errors (no QMessageBox
        here -> testable + embeddable)."""

        def __init__(self, workbook_path, parent=None, show_intro=True,
                     with_save=False):
            super().__init__(parent)
            self._wb_path = workbook_path
            try:
                import theme
                self._t = theme
            except ImportError:
                self._t = None

            layout = QVBoxLayout(self)
            layout.setContentsMargins(24, 20, 24, 16)
            layout.setSpacing(12)

            title = QLabel("Rifle & Setup")
            tf = QFont()
            tf.setPointSize(17)
            tf.setWeight(QFont.DemiBold)
            title.setFont(tf)
            if self._t:
                title.setStyleSheet(f"color: {self._t.TEXT_PRIMARY};")
            layout.addWidget(title)

            if show_intro:
                intro = QLabel(
                    "Edit your rifle, shooter, components, and session "
                    "info here — Loadscope writes it straight into your "
                    "workbook.")
                intro.setWordWrap(True)
                if self._t:
                    intro.setStyleSheet(
                        f"color: {self._t.TEXT_SECONDARY};")
                layout.addWidget(intro)

            try:
                current = read_rifle_setup(workbook_path)
                self.read_error = None
            except Exception as e:
                current = {c: "" for _s, _l, c, _k in RIFLE_SETUP_FIELDS}
                self.read_error = str(e)

            # ONE shared form for every section so the input column
            # lines up across Rifle & Shooter / Load Components / Test
            # Session. (Each section having its own QFormLayout sized its
            # label column independently -> fields offset per section.)
            # Section headers are full-width spanning rows.
            self._edits = {}
            form = QFormLayout()
            form.setVerticalSpacing(10)
            form.setHorizontalSpacing(16)
            form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
            layout.addLayout(form)
            last_section = None
            for section, label, cell, kind in RIFLE_SETUP_FIELDS:
                if section != last_section:
                    hdr = QLabel(section.upper())
                    if self._t:
                        hdr.setStyleSheet(
                            f"color: {self._t.TEXT_TERTIARY}; "
                            f"font-size: {self._t.FONT_SIZE_TINY}px; "
                            f"text-transform: uppercase; letter-spacing: 1px; "
                            f"font-weight: bold; padding-top: 8px;")
                    form.addRow(hdr)  # spans both columns
                    last_section = section
                w = _build_field(kind)
                w.set_value(current.get(cell, ""))
                w.setMinimumHeight(34)
                if isinstance(w, _LineField):
                    w.setStyleSheet("QLineEdit { padding: 6px 10px; "
                                    "font-size: 14px; }")
                self._edits[cell] = w
                form.addRow(label + ":", w)
            layout.addStretch(1)

            # When embedded in the shell there's no dialog OK button, so
            # the panel carries its own Save + inline status.
            self._status = None
            if with_save:
                row = QVBoxLayout()
                self._status = QLabel("")
                self._status.setWordWrap(True)
                row.addWidget(self._status)
                btn = QPushButton("Save")
                btn.setObjectName("primary")
                btn.clicked.connect(self._on_save_clicked)
                row.addWidget(btn)
                layout.addLayout(row)

        def _on_save_clicked(self):
            status, payload = self.save()
            if not self._status:
                return
            if status == "ok":
                msg = ("Saved." if payload else "No changes to save.")
                col = self._t.LOG_SUCCESS if self._t else "green"
            elif status == "locked":
                msg = ("The workbook is open in Excel — close it and "
                       "click Save again.")
                col = self._t.LOG_ERROR if self._t else "red"
            else:
                msg = f"Couldn't save: {payload}"
                col = self._t.LOG_ERROR if self._t else "red"
            self._status.setText(msg)
            self._status.setStyleSheet(f"color: {col};")

        def save(self):
            """Returns (status, payload): ("ok", changed_list) |
            ("locked", None) | ("error", message)."""
            values = {c: w.value() for c, w in self._edits.items()}
            try:
                changed = write_rifle_setup(self._wb_path, values)
            except PermissionError:
                return ("locked", None)
            except Exception as e:
                return ("error", str(e))
            # Remember history-backed entries only after a clean save.
            for w in self._edits.values():
                try:
                    w.commit()
                except Exception:
                    pass
            return ("ok", changed)


if QDialog is not None:

    class RifleSetupDialog(QDialog):
        """Thin modal wrapper around RifleSetupPanel — preserves the
        exact menu-launched flow (Workbook → Rifle & Setup)."""

        def __init__(self, workbook_path, parent=None):
            super().__init__(parent)
            self.saved = False
            self.setWindowTitle("Loadscope — Rifle & Setup")
            self.setModal(True)
            self.setMinimumWidth(560)

            layout = QVBoxLayout(self)
            layout.setContentsMargins(0, 0, 0, 0)
            self._panel = RifleSetupPanel(workbook_path, parent=self)
            layout.addWidget(self._panel)
            if self._panel.read_error:
                QMessageBox.critical(
                    self, "Couldn't read workbook",
                    "Loadscope couldn't read the workbook:\n\n"
                    f"{self._panel.read_error}")

            btns = QDialogButtonBox(
                QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            ok = btns.button(QDialogButtonBox.Ok)
            if ok is not None:
                ok.setObjectName("primary")
                ok.setText("Save")
            btns.accepted.connect(self._save)
            btns.rejected.connect(self.reject)
            cont = QVBoxLayout()
            cont.setContentsMargins(24, 0, 24, 16)
            cont.addWidget(btns)
            layout.addLayout(cont)

        def _save(self):
            status, payload = self._panel.save()
            if status == "locked":
                QMessageBox.warning(
                    self, "Workbook is open",
                    "Close the workbook in Excel and try again so "
                    "Loadscope can save your changes.")
                return
            if status == "error":
                QMessageBox.critical(
                    self, "Couldn't save",
                    f"Loadscope couldn't write to the workbook:\n\n{payload}")
                return
            self.saved = bool(payload)
            self.accept()


def show_rifle_setup(workbook_path, parent=None):
    """Open the Rifle & Setup editor modally. Returns True if saved."""
    dlg = RifleSetupDialog(workbook_path, parent=parent)
    dlg.exec_()
    return dlg.saved
