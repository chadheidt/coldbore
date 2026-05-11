"""
Load card generator.

Reads the user's working workbook, pulls the suggested winning charge and
seating depth (plus the rifle/cartridge/components info from the Load Log
header), and writes a single-page HTML "load card" formatted for printing.
The user opens it in their browser and chooses File → Print → Save as PDF
(or just prints it for the range bag).

Tools → Generate Load Card… in the app launches this.
"""

import os
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# Where to read each piece of data from. Cell coordinates are the merged-cell
# ANCHOR cells (top-left of each merged range). Adjacent label cells (e.g.,
# "Rifle:" in A5) are intentionally NOT in this map — they're labels, not values.
LOAD_LOG_FIELDS = {
    "rifle":        ("Load Log", "B5"),   # B5:E5 merged, holds rifle name
    "shooter":      ("Load Log", "G5"),   # G5:J5 merged
    "cartridge":    ("Load Log", "L5"),   # L5:N5 merged
    "barrel":       ("Load Log", "B6"),   # B6:E6 merged
    "optic":        ("Load Log", "G6"),   # G6:J6 merged
    "chrono":       ("Load Log", "L6"),   # L6:N6 merged
    "bullet":       ("Load Log", "B9"),   # B9:D9 merged
    "powder":       ("Load Log", "F9"),   # F9:G9 merged, has a LOOKUP formula
    "primer":       ("Load Log", "I9"),   # I9:J9 merged
    "brass":        ("Load Log", "L9"),   # L9:N9 merged
    "cbto":         ("Load Log", "B10"),  # B10:E10 merged
    "off_lands":    ("Load Log", "G10"),  # G10:J10 merged
    "distance":     ("Load Log", "L10"),  # L10:N10 merged — yardage, NOT bullet weight
}

# Suggested winners
SUGGESTED_CHARGE_CELL = ("Charts", "B3")
SUGGESTED_JUMP_CELL = ("Seating Depth", "D2")
SUGGESTED_AVG_VEL_CELL = ("Charts", "E5")  # Avg velocity at suggested charge
SUGGESTED_SD_CELL = ("Charts", "G3")
SUGGESTED_GROUP_CELL = ("Charts", "G4")
SUGGESTED_MR_CELL = ("Charts", "L2")  # mean radius at suggested jump


def _cell(wb, sheet_name, coord):
    """Read a cell value. Returns None if sheet missing or value is None."""
    if sheet_name not in wb.sheetnames:
        return None
    return wb[sheet_name][coord].value


def _fmt(value, blank="—"):
    """Format a cell value for display. None / empty / 0 numerics get a dash."""
    if value is None:
        return blank
    if isinstance(value, str):
        return value.strip() or blank
    if isinstance(value, float):
        if value != value:  # NaN
            return blank
        return f"{value:g}"
    return str(value)


def _extract_bullet_weight(bullet_str):
    """Try to pull a bullet weight in grains from a description like
    '140 ELD-M' → 140 or 'Berger 215 Hybrid' → 215.
    Returns None if no plausible weight is found."""
    if not bullet_str:
        return None
    import re
    # Look for an integer 30-800 (typical rifle bullet weight range)
    for match in re.finditer(r"\b(\d{2,3})\b", str(bullet_str)):
        n = int(match.group(1))
        if 30 <= n <= 800:
            return n
    return None


def _gather(workbook_path):
    """Open the workbook with computed values and pull every field for the card."""
    # data_only=True returns computed cell values (not formulas) — key for
    # showing the *result* of the suggested-charge formula, not "=Charts!B3".
    wb = load_workbook(workbook_path, data_only=True, keep_vba=False)

    data = {}
    for key, (sheet, coord) in LOAD_LOG_FIELDS.items():
        data[key] = _cell(wb, sheet, coord)

    # Bullet weight isn't a separate cell — extract it from the bullet description
    data["bullet_wt"] = _extract_bullet_weight(data.get("bullet"))

    data["suggested_charge"] = _cell(wb, *SUGGESTED_CHARGE_CELL)
    data["suggested_jump"] = _cell(wb, *SUGGESTED_JUMP_CELL)
    data["suggested_avg_vel"] = _cell(wb, *SUGGESTED_AVG_VEL_CELL)
    data["suggested_sd"] = _cell(wb, *SUGGESTED_SD_CELL)
    data["suggested_group"] = _cell(wb, *SUGGESTED_GROUP_CELL)
    data["suggested_mr"] = _cell(wb, *SUGGESTED_MR_CELL)

    data["workbook_name"] = os.path.splitext(os.path.basename(workbook_path))[0]
    data["generated_at"] = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    return data


def _build_html(d):
    """Render the load card as a clean printable HTML page."""
    # Inline CSS only — no external dependencies. Tuned for one 8.5x11 page
    # with a 4x6-card-shaped main panel that works well as a single load card
    # AND fills a letter-sized page with whitespace if printed at full scale.
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Load Card — {_fmt(d['cartridge'])}</title>
<style>
  @page {{ margin: 0.5in; }}
  body {{
    font-family: -apple-system, "SF Pro Text", "Helvetica Neue", sans-serif;
    color: #1a1a1a;
    margin: 0;
    padding: 24px;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .card {{
    max-width: 720px;
    margin: 0 auto;
    border: 2px solid #1a1a1a;
    border-radius: 12px;
    overflow: hidden;
  }}
  .header {{
    background: #1f2229;
    color: #f5f5f5;
    padding: 18px 24px;
    text-align: center;
  }}
  .header h1 {{
    margin: 0;
    font-size: 24px;
    letter-spacing: 1px;
  }}
  .header .subtitle {{
    color: #d97706;
    font-size: 13px;
    letter-spacing: 2px;
    margin-top: 4px;
    text-transform: uppercase;
  }}
  .winner {{
    background: #fff8e1;
    padding: 18px 24px;
    border-bottom: 1px solid #1a1a1a;
  }}
  .winner-title {{
    font-size: 11px;
    letter-spacing: 2px;
    color: #5d4037;
    text-transform: uppercase;
    font-weight: bold;
    margin-bottom: 6px;
  }}
  .winner-row {{
    display: flex;
    gap: 24px;
    flex-wrap: wrap;
  }}
  .winner-cell {{
    flex: 1;
    min-width: 140px;
  }}
  .winner-label {{ font-size: 11px; color: #555; text-transform: uppercase; letter-spacing: 1px; }}
  .winner-value {{ font-size: 22px; font-weight: 600; color: #1a1a1a; margin-top: 2px; }}
  .winner-unit  {{ font-size: 13px; color: #555; font-weight: normal; }}

  .body {{ padding: 18px 24px; }}
  .section-title {{
    font-size: 11px;
    letter-spacing: 2px;
    color: #888;
    text-transform: uppercase;
    font-weight: bold;
    border-bottom: 1px solid #ccc;
    padding-bottom: 4px;
    margin: 14px 0 8px;
  }}
  .grid {{
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 8px 16px;
  }}
  .grid .field {{ font-size: 13px; line-height: 1.4; }}
  .grid .field .label {{ color: #888; font-size: 11px; }}
  .grid .field .value {{ font-weight: 500; color: #1a1a1a; }}
  .stats {{ display: grid; grid-template-columns: 1fr 1fr 1fr 1fr; gap: 8px 16px; }}
  .stats .field {{ font-size: 13px; }}
  .stats .field .label {{ color: #888; font-size: 11px; }}
  .stats .field .value {{ font-weight: 600; font-size: 16px; color: #1a1a1a; }}

  .footer {{
    background: #f5f5f5;
    padding: 10px 24px;
    font-size: 10px;
    color: #888;
    text-align: center;
    border-top: 1px solid #ccc;
  }}
</style>
</head>
<body>
<div class="card">

  <div class="header">
    <h1>{_fmt(d['cartridge']) or 'Load Card'}</h1>
    <div class="subtitle">True Zero Load Card</div>
  </div>

  <div class="winner">
    <div class="winner-title">Suggested winning load</div>
    <div class="winner-row">
      <div class="winner-cell">
        <div class="winner-label">Powder charge</div>
        <div class="winner-value">{_fmt(d['suggested_charge'])} <span class="winner-unit">gr</span></div>
      </div>
      <div class="winner-cell">
        <div class="winner-label">Bullet jump</div>
        <div class="winner-value">{_fmt(d['suggested_jump'])} <span class="winner-unit">in</span></div>
      </div>
      <div class="winner-cell">
        <div class="winner-label">Powder type</div>
        <div class="winner-value" style="font-size: 16px;">{_fmt(d['powder'])}</div>
      </div>
    </div>
  </div>

  <div class="body">
    <div class="section-title">Components</div>
    <div class="grid">
      <div class="field"><div class="label">Bullet</div><div class="value">{_fmt(d['bullet'])}</div></div>
      <div class="field"><div class="label">Bullet weight</div><div class="value">{_fmt(d['bullet_wt'])} gr</div></div>
      <div class="field"><div class="label">Powder</div><div class="value">{_fmt(d['powder'])}</div></div>
      <div class="field"><div class="label">Primer</div><div class="value">{_fmt(d['primer'])}</div></div>
      <div class="field"><div class="label">Brass</div><div class="value">{_fmt(d['brass'])}</div></div>
      <div class="field"><div class="label">CBTO</div><div class="value">{_fmt(d['cbto'])}</div></div>
    </div>

    <div class="section-title">Rifle &amp; setup</div>
    <div class="grid">
      <div class="field"><div class="label">Rifle</div><div class="value">{_fmt(d['rifle'])}</div></div>
      <div class="field"><div class="label">Barrel</div><div class="value">{_fmt(d['barrel'])}</div></div>
      <div class="field"><div class="label">Cartridge</div><div class="value">{_fmt(d['cartridge'])}</div></div>
      <div class="field"><div class="label">Optic</div><div class="value">{_fmt(d['optic'])}</div></div>
      <div class="field"><div class="label">Chrono</div><div class="value">{_fmt(d['chrono'])}</div></div>
      <div class="field"><div class="label">Distance tested</div><div class="value">{_fmt(d['distance'])} yd</div></div>
    </div>

    <div class="section-title">Performance at suggested load</div>
    <div class="stats">
      <div class="field"><div class="label">Avg velocity</div><div class="value">{_fmt(d['suggested_avg_vel'])} fps</div></div>
      <div class="field"><div class="label">SD</div><div class="value">{_fmt(d['suggested_sd'])} fps</div></div>
      <div class="field"><div class="label">Group</div><div class="value">{_fmt(d['suggested_group'])}"</div></div>
      <div class="field"><div class="label">Mean radius</div><div class="value">{_fmt(d['suggested_mr'])}"</div></div>
    </div>
  </div>

  <div class="footer">
    Generated by True Zero from <b>{_fmt(d['workbook_name'])}</b> on {d['generated_at']} ·
    Always cross-check loads against published reloading manuals.
  </div>
</div>
</body>
</html>
"""


def generate_load_card(workbook_path, open_after=True):
    """Generate an HTML load card from the workbook. Returns the path to the
    generated file. If `open_after`, opens it in the user's default browser
    so they can View → Print → Save as PDF."""
    data = _gather(workbook_path)
    html = _build_html(data)

    # Save next to the workbook in a Load Cards/ subfolder for organization
    project_dir = os.path.dirname(os.path.abspath(workbook_path))
    out_dir = os.path.join(project_dir, "Load Cards")
    os.makedirs(out_dir, exist_ok=True)

    workbook_base = os.path.splitext(os.path.basename(workbook_path))[0]
    stamp = datetime.now().strftime("%Y-%m-%d %H-%M")
    out_path = os.path.join(out_dir, f"{workbook_base} — Load Card {stamp}.html")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    if open_after:
        subprocess.run(["open", out_path])

    return out_path
