"""Native "Range Session" form — UI-change Phase 2 ([[loadscope-appify-scoping]]).

Session date + the DOPE table (elevation/wind you dialed at each
distance) edited in a clean Loadscope grid instead of the Excel cells.
This is THE screen the ballistic solver will later pre-fill with
predicted DOPE (the user then confirms/overwrites at the range).

Writes into the Ballistics DOPE table (rows 9-18 = 100..1000 yd) and
the Load Log session Date (B13). The click columns (C/E/G/I) are
workbook FORMULAS that auto-fill from the click value — never touched.
Only the shooter's unit is shown (Mil OR MOA, from Load Log!G7) — same
principle as the printed Pocket Card.

Field map + read/write helpers are pure (no Qt) and unit-tested.
"""

import os
from datetime import datetime

try:
    from PyQt5.QtGui import QFont
    from PyQt5.QtWidgets import (
        QComboBox,
        QDialog,
        QDialogButtonBox,
        QGridLayout,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMessageBox,
        QPushButton,
        QScrollArea,
        QVBoxLayout,
        QWidget,
    )
except ImportError:  # allow import for unit tests without PyQt5
    QDialog = QDialogButtonBox = QGridLayout = QLabel = QLineEdit = None
    QMessageBox = QScrollArea = QVBoxLayout = QWidget = QFont = None
    QPushButton = QComboBox = QHBoxLayout = None

from openpyxl import load_workbook

import dope_solver

# "How Loadscope predicts" — the confidence-building transparency note
# (Chad-approved gate-4 copy; honest about measured-drag solvers).
PREDICT_NOTE = (
    "Loadscope predicts your DOPE with a G7/G1 ballistic-coefficient "
    "solver — the same model class as Strelok, GeoBallistics, and "
    "the free Hornady and Berger calculators, and validated against "
    "JBM. With good inputs it's accurate to about 1000 yards; measured-"
    "drag solvers (Applied Ballistics, Hornady 4DOF) pull ahead deep "
    "in the transonic range. Grey values are PREDICTED — always "
    "confirm at the range; type your dialed value and it turns solid.")


def _cell_save_text(confirmed, widget_text, orig_text):
    """What to persist for a DOPE cell. CONFIRMED -> the widget text;
    a still-PREDICTED (untouched) cell -> its original workbook value
    (so a live prediction is NEVER written to the workbook). Pure."""
    return widget_text if confirmed else orig_text

DOPE_ROWS = range(9, 19)          # Ballistics rows 9..18 = 100..1000 yd
_RANGE_COL = "A"
_NOTES_COL = "K"
_CLICK_CELL = ("Load Log", "G7")  # decides Mil vs MOA
_DATE_CELL = ("Load Log", "B13")
_BAL = "Ballistics"
# MM/DD/YY is the displayed + preferred input format (Chad 2026-05-15);
# the others are still accepted on input so a typed Y-M-D won't error.
_DATE_FORMATS = ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d")
_DATE_DISPLAY = "%m/%d/%y"


def dope_unit(workbook_path):
    """'moa' or 'mil' from the scope click value. Defaults to mil."""
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    sheet, cell = _CLICK_CELL
    v = str(wb[sheet][cell].value or "").lower()
    return "moa" if "moa" in v else "mil"


def _unit_cols(unit):
    """(elevation col, wind col) the USER types into for this unit.
    Mil: B (Mils Dialed) / F (Wind Hold Mils).
    MOA: D (MOA Dialed) / H (Wind Hold MOA)."""
    return ("D", "H") if unit == "moa" else ("B", "F")


def _date_str(v):
    if isinstance(v, datetime):
        return v.strftime(_DATE_DISPLAY)
    return "" if v is None else str(v)


def read_dope(workbook_path):
    """Return {'unit', 'date' (YYYY-MM-DD str), 'rows': [{row, range,
    elev, wind, notes}]}. Pure — no Qt."""
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    bal = wb[_BAL]
    unit = "moa" if "moa" in str(
        wb[_CLICK_CELL[0]][_CLICK_CELL[1]].value or "").lower() else "mil"
    ecol, wcol = _unit_cols(unit)
    rows = []
    for r in DOPE_ROWS:
        def s(col):
            v = bal[f"{col}{r}"].value
            return "" if v is None else str(v)
        rows.append({
            "row": r,
            "range": s(_RANGE_COL),
            "elev": s(ecol),
            "wind": s(wcol),
            "notes": s(_NOTES_COL),
        })
    return {
        "unit": unit,
        "date": _date_str(wb[_DATE_CELL[0]][_DATE_CELL[1]].value),
        "rows": rows,
    }


def _coerce(raw):
    """'' -> '' (clears the cell; the click formula yields blank).
    A number -> float. Junk -> None (skip; don't corrupt)."""
    t = raw.strip()
    if t == "":
        return ""
    try:
        return float(t)
    except ValueError:
        return None


def write_dope(workbook_path, unit, rows, date_str=None):
    """Write elevation/wind/notes (unit-appropriate cells) + optional
    session date. Never touches the click formula columns. Returns the
    list of changed cells. Raises PermissionError if open in Excel.
    Pure — no Qt."""
    wb = load_workbook(workbook_path, data_only=False, keep_vba=False)
    bal = wb[_BAL]
    ecol, wcol = _unit_cols(unit)
    changed = []

    def maybe(cell, newval):
        cur = bal[cell].value
        cur_norm = "" if cur is None else cur
        if newval is None:                       # unparseable -> skip
            return
        if newval == cur_norm:
            return
        if isinstance(newval, float):
            try:
                if cur is not None and float(cur) == newval:
                    return
            except (TypeError, ValueError):
                pass
        bal[cell] = newval
        changed.append(cell)

    by_row = {r["row"]: r for r in rows}
    for r in DOPE_ROWS:
        row = by_row.get(r)
        if not row:
            continue
        maybe(f"{ecol}{r}", _coerce(row.get("elev", "")))
        maybe(f"{wcol}{r}", _coerce(row.get("wind", "")))
        n = row.get("notes", "").strip()
        ncur = bal[f"{_NOTES_COL}{r}"].value
        if n != ("" if ncur is None else str(ncur)):
            bal[f"{_NOTES_COL}{r}"] = n
            changed.append(f"{_NOTES_COL}{r}")

    if date_str is not None:
        d = date_str.strip()
        ll = wb[_DATE_CELL[0]]
        cur = ll[_DATE_CELL[1]].value
        if d == "":
            pass  # don't clear a date by blanking the field
        else:
            parsed = None
            for fmt in _DATE_FORMATS:
                try:
                    parsed = datetime.strptime(d, fmt)
                    break
                except ValueError:
                    continue
            if parsed is not None and parsed != cur:
                ll[_DATE_CELL[1]] = parsed
                changed.append(f"{_DATE_CELL[0]}!{_DATE_CELL[1]}")

    if changed:
        wb.save(workbook_path)
    return changed


if QWidget is not None:

    class DopeEntryPanel(QWidget):
        """Embeddable Range Session & DOPE form with the gate-4
        predicted-vs-confirmed overlay. Grey/italic = PREDICTED (a live
        solver estimate, never written to the workbook); type your
        dialed value and the cell turns solid = CONFIRMED (saved).
        save() returns a result tuple (testable + embeddable)."""

        def __init__(self, workbook_path, parent=None, show_intro=True,
                     with_save=False):
            super().__init__(parent)
            self._wb_path = workbook_path
            try:
                import theme
                self._t = theme
            except ImportError:
                self._t = None
            self._c_primary = self._t.TEXT_PRIMARY if self._t else "#000"
            self._c_pred = self._t.TEXT_TERTIARY if self._t else "#888"

            try:
                data = read_dope(workbook_path)
                self.read_error = None
            except Exception as e:
                data = {"unit": "mil", "date": "", "rows": []}
                self.read_error = str(e)
            self._unit = data["unit"]
            ulabel = "MOA" if self._unit == "moa" else "Mils"

            # live predicted overlay (never written to the workbook)
            try:
                self._pred = dope_solver.predicted_dope(workbook_path)
            except Exception as e:                    # never crash UI
                self._pred = {"status": dope_solver.ERROR,
                              "message": str(e), "rows": {}}

            layout = QVBoxLayout(self)
            layout.setContentsMargins(24, 20, 24, 16)
            layout.setSpacing(10)

            title = QLabel("Range Session & DOPE")
            tf = QFont()
            tf.setPointSize(17)
            tf.setWeight(QFont.DemiBold)
            title.setFont(tf)
            if self._t:
                title.setStyleSheet(f"color: {self._c_primary};")
            layout.addWidget(title)

            if show_intro:
                intro = QLabel(
                    f"Grey values are PREDICTED for your load. Type what "
                    f"you actually dialed at the range, in {ulabel} (your "
                    "scope's unit), and the cell turns solid = confirmed. "
                    "Click counts auto-fill in the workbook.")
                intro.setWordWrap(True)
                if self._t:
                    intro.setStyleSheet(f"color: {self._t.TEXT_SECONDARY};")
                layout.addWidget(intro)

            layout.addLayout(self._build_predict_controls())

            date_row = QGridLayout()
            date_row.addWidget(QLabel("Session date (MM/DD/YY):"), 0, 0)
            self._date_edit = QLineEdit(data.get("date", ""))
            self._date_edit.setMaximumWidth(160)
            date_row.addWidget(self._date_edit, 0, 1)
            date_row.setColumnStretch(2, 1)
            layout.addLayout(date_row)

            grid = QGridLayout()
            grid.setHorizontalSpacing(10)
            grid.setVerticalSpacing(6)
            for col, head in enumerate(
                    ["Range (yd)", f"Elev ({ulabel})",
                     f"Wind/10mph ({ulabel})", "Notes", ""]):
                h = QLabel(head)
                if self._t:
                    h.setStyleSheet(
                        f"color: {self._t.TEXT_TERTIARY}; "
                        f"font-weight: bold;")
                grid.addWidget(h, 0, col)

            pred_rows = self._pred.get("rows", {})
            self._row_edits = {}        # rnum -> (elev, wind, notes) compat
            self._cells = {}            # (rnum, 'elev'|'wind') -> state
            self._tags = {}             # rnum -> QLabel "predicted"
            for i, row in enumerate(data["rows"], start=1):
                rnum = row["row"]
                rng = QLabel(str(row["range"]))
                if self._t:
                    rng.setStyleSheet(f"color: {self._c_primary};")
                e_el = QLineEdit()
                e_wd = QLineEdit()
                e_nt = QLineEdit(row["notes"])
                for fld, w, orig in (("elev", e_el, row["elev"]),
                                     ("wind", e_wd, row["wind"])):
                    w.setMaximumWidth(120)
                    st = {"w": w, "orig": orig, "confirmed": False}
                    self._cells[(rnum, fld)] = st
                    pv = pred_rows.get(rnum, {}).get(fld, "")
                    if str(orig).strip() != "":
                        w.setText(str(orig))
                        st["confirmed"] = True
                    elif pv != "":
                        w.setText(pv)            # predicted ghost
                    else:
                        st["confirmed"] = True   # plain empty cell
                    self._paint_cell(rnum, fld)
                    w.textEdited.connect(
                        lambda _t, rn=rnum, f=fld: self._on_edit(rn, f))
                tag = QLabel("predicted")
                if self._t:
                    tag.setStyleSheet(
                        f"color: {self._c_pred}; font-style: italic;")
                self._tags[rnum] = tag
                grid.addWidget(rng, i, 0)
                grid.addWidget(e_el, i, 1)
                grid.addWidget(e_wd, i, 2)
                grid.addWidget(e_nt, i, 3)
                grid.addWidget(tag, i, 4)
                self._row_edits[rnum] = (e_el, e_wd, e_nt)
                self._refresh_tag(rnum)

            grid.setRowStretch(len(data["rows"]) + 2, 1)
            holder = QWidget()
            holder.setLayout(grid)
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(holder)
            layout.addWidget(scroll, stretch=1)

            note = QLabel(PREDICT_NOTE)
            note.setWordWrap(True)
            if self._t:
                f = QFont()
                f.setPointSize(10)
                note.setFont(f)
                note.setStyleSheet(f"color: {self._t.TEXT_TERTIARY};")
            layout.addWidget(note)

            self._status = None
            if with_save:
                self._status = QLabel("")
                self._status.setWordWrap(True)
                layout.addWidget(self._status)
                btn = QPushButton("Save DOPE")
                btn.setObjectName("primary")
                btn.clicked.connect(self._on_save_clicked)
                layout.addWidget(btn)
            self._refresh_pred_status()

        # ---- predicted overlay --------------------------------------
        def _build_predict_controls(self):
            """BC + atmosphere + wind inputs that drive the prediction.
            Defaults come from the live solver result / workbook."""
            atm = self._pred.get("atmosphere", {})
            grid = QGridLayout()
            grid.setHorizontalSpacing(8)
            grid.setVerticalSpacing(6)

            def field(text, col, val, width=70):
                lab = QLabel(text)
                if self._t:
                    lab.setStyleSheet(
                        f"color: {self._t.TEXT_TERTIARY};")
                e = QLineEdit("" if val is None else str(val))
                e.setMaximumWidth(width)
                grid.addWidget(lab, 0, col)
                grid.addWidget(e, 0, col + 1)
                return e

            bc0 = self._pred.get("bc")
            self._bc_edit = field("BC", 0,
                                   "" if bc0 is None else f"{bc0:.3f}", 80)
            self._model_combo = QComboBox()
            self._model_combo.addItems(["G7", "G1"])
            self._model_combo.setCurrentText(
                self._pred.get("bc_model") or "G7")
            grid.addWidget(self._model_combo, 0, 2)
            self._temp_edit = field("Temp F", 3,
                                     atm.get("temp_f", 59.0))
            self._press_edit = field("Press inHg", 5,
                                      atm.get("pressure_inhg", 29.92))
            self._wind_edit = field("Wind mph", 7,
                                     atm.get("wind_mph", 10.0))
            self._clock_edit = field("@ o'clock", 9,
                                      atm.get("wind_clock", 3.0))
            upd = QPushButton("Update predictions")
            upd.clicked.connect(self._on_update)
            grid.addWidget(upd, 0, 11)
            self._bc_src = QLabel("")
            if self._t:
                self._bc_src.setStyleSheet(
                    f"color: {self._t.TEXT_TERTIARY};")
            grid.addWidget(self._bc_src, 1, 0, 1, 12)
            grid.setColumnStretch(12, 1)
            return grid

        def _paint_cell(self, rnum, fld):
            st = self._cells[(rnum, fld)]
            if st["confirmed"]:
                st["w"].setStyleSheet(
                    f"color: {self._c_primary}; font-style: normal;")
            else:
                st["w"].setStyleSheet(
                    f"color: {self._c_pred}; font-style: italic;")

        def _refresh_tag(self, rnum):
            pred = any(not self._cells[(rnum, f)]["confirmed"]
                       and self._cells[(rnum, f)]["w"].text().strip() != ""
                       for f in ("elev", "wind"))
            self._tags[rnum].setVisible(pred)

        def _on_edit(self, rnum, fld):
            st = self._cells[(rnum, fld)]
            if not st["confirmed"]:
                st["confirmed"] = True
                self._paint_cell(rnum, fld)
            self._refresh_tag(rnum)

        def _refresh_pred_status(self):
            s = self._pred.get("status")
            if s == dope_solver.OK:
                msg = (f"Predicted from BC {self._pred['bc']:.3f} "
                       f"{self._pred['bc_model']} "
                       f"({self._pred['bc_source']}).")
            else:
                msg = self._pred.get("message", "")
            self._bc_src.setText(msg)

        def _apply_predictions(self):
            rows = self._pred.get("rows", {})
            for (rnum, fld), st in self._cells.items():
                if st["confirmed"]:
                    continue
                st["w"].setText(rows.get(rnum, {}).get(fld, ""))
                self._paint_cell(rnum, fld)
            for rnum in self._tags:
                self._refresh_tag(rnum)
            self._refresh_pred_status()

        def _on_update(self):
            def f(edit, dflt):
                try:
                    return float(edit.text().strip())
                except (ValueError, AttributeError):
                    return dflt
            bc_txt = self._bc_edit.text().strip()
            kw = dict(
                temp_f=f(self._temp_edit, None),
                pressure_inhg=f(self._press_edit, 29.92),
                wind_mph=f(self._wind_edit, 10.0),
                wind_clock=f(self._clock_edit, 3.0))
            if bc_txt:
                kw["manual_bc"] = f(self._bc_edit, None)
                kw["manual_model"] = self._model_combo.currentText()
            try:
                self._pred = dope_solver.predicted_dope(self._wb_path, **kw)
            except Exception as e:
                self._pred = {"status": dope_solver.ERROR,
                              "message": str(e), "rows": {}}
            self._apply_predictions()

        def _on_save_clicked(self):
            status, payload = self.save()
            if not self._status:
                return
            if status == "ok":
                msg = "Saved." if payload else "No changes to save."
                col = self._t.LOG_SUCCESS if self._t else "green"
            elif status == "locked":
                msg = ("The workbook is open in Excel — close it and "
                       "click Save again.")
                col = self._t.LOG_ERROR if self._t else "red"
            else:
                msg = f"Couldn't save: {payload}"
                col = self._t.LOG_ERROR if self._t else "red"
            self._status.setText(msg)
            self._status.setStyleSheet(f"color: {col};")

        def save(self):
            """Write CONFIRMED DOPE only. A still-predicted (untouched)
            cell is NEVER written — the workbook stays confirmed-only.
            Returns (status, payload)."""
            rows = []
            for rn, (el, wd, nt) in self._row_edits.items():
                se = self._cells[(rn, "elev")]
                sw = self._cells[(rn, "wind")]
                rows.append({
                    "row": rn,
                    "elev": _cell_save_text(se["confirmed"], el.text(),
                                            se["orig"]),
                    "wind": _cell_save_text(sw["confirmed"], wd.text(),
                                            sw["orig"]),
                    "notes": nt.text()})
            try:
                return ("ok", write_dope(self._wb_path, self._unit, rows,
                                         self._date_edit.text()))
            except PermissionError:
                return ("locked", None)
            except Exception as e:
                return ("error", str(e))


if QDialog is not None:

    class DopeEntryDialog(QDialog):
        """Thin modal wrapper around DopeEntryPanel — preserves the
        menu-launched flow (Workbook → Range Session & DOPE)."""

        def __init__(self, workbook_path, parent=None):
            super().__init__(parent)
            self.saved = False
            self.setWindowTitle("Loadscope — Range Session & DOPE")
            self.setModal(True)
            self.setMinimumWidth(620)

            layout = QVBoxLayout(self)
            layout.setContentsMargins(0, 0, 0, 0)
            self._panel = DopeEntryPanel(workbook_path, parent=self)
            layout.addWidget(self._panel)
            if self._panel.read_error:
                QMessageBox.critical(
                    self, "Couldn't read workbook",
                    "Loadscope couldn't read the workbook:\n\n"
                    f"{self._panel.read_error}")

            btns = QDialogButtonBox(
                QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            ok = btns.button(QDialogButtonBox.Ok)
            if ok is not None:
                ok.setObjectName("primary")
                ok.setText("Save")
            btns.accepted.connect(self._save)
            btns.rejected.connect(self.reject)
            cont = QVBoxLayout()
            cont.setContentsMargins(24, 0, 24, 16)
            cont.addWidget(btns)
            layout.addLayout(cont)

        def _save(self):
            status, payload = self._panel.save()
            if status == "locked":
                QMessageBox.warning(
                    self, "Workbook is open",
                    "Close the workbook in Excel and try again so "
                    "Loadscope can save your DOPE.")
                return
            if status == "error":
                QMessageBox.critical(
                    self, "Couldn't save",
                    f"Loadscope couldn't write to the workbook:\n\n{payload}")
                return
            self.saved = bool(payload)
            self.accept()


def show_dope_entry(workbook_path, parent=None):
    """Open the Range Session & DOPE editor modally. True if saved."""
    dlg = DopeEntryDialog(workbook_path, parent=parent)
    dlg.exec_()
    return dlg.saved
