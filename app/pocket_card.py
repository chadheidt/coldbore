"""Pocket Range Card (DOPE card) generator.

Reads the user's Ballistics tab (rifle/cartridge/scope info + the 100-1000yd
elev/wind table they've filled in) and writes a single 4x6 landscape HTML
page sized for printing as a pocket-card range reference.

User opens the HTML in their browser and chooses File → Print → Save as PDF,
or prints directly to 4x6 index card stock if they have it. Standard
letter-paper printing also works (the card just appears centered with
whitespace around it).

Workbook → Print Pocket Range Card… in the app launches this.
"""

import base64
import os
import subprocess
import sys
from datetime import datetime

from openpyxl import load_workbook

import dope_solver


def _logo_data_uri():
    """Read the Loadscope icon and return a base64 data URI so the
    generated HTML stays a single self-contained file (works offline,
    portable to email, doesn't break if the user moves it)."""
    # Look up the icon in the package's resources. Two location candidates
    # depending on whether we're in dev or in the .app bundle:
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    # py2app: __file__ is inside Contents/Resources/lib/pythonXX.zip, so
    # resolve the real Contents/Resources/ via sys.executable first
    # (proven-good pattern from setup_wizard.find_bundled_template()).
    if getattr(sys, "frozen", False):
        try:
            exe = os.path.abspath(sys.executable)
            candidates.append(os.path.join(
                os.path.dirname(os.path.dirname(exe)), "Resources", "icon.png"))
        except (OSError, ValueError):
            pass
    candidates += [
        os.path.normpath(os.path.join(here, "..", "docs", "assets", "icon.png")),
        os.path.normpath(os.path.join(here, "resources", "icon.png")),
        os.path.normpath(os.path.join(here, "..", "Resources", "icon.png")),
    ]
    for path in candidates:
        if os.path.isfile(path):
            with open(path, "rb") as f:
                return "data:image/png;base64," + base64.b64encode(f.read()).decode("ascii")
    # Icon missing — return empty so the card still renders, just no logo
    return ""


# Header info on the Ballistics tab. Cell coordinates are the value cells
# (not the labels). Some are merged ranges — using the top-left anchor.
HEADER_FIELDS = {
    "rifle":    ("Ballistics", "B5"),
    "bullet":   ("Ballistics", "E5"),
    "charge":   ("Ballistics", "H5"),
    "vel":      ("Ballistics", "K5"),
    "scope":    ("Ballistics", "B6"),
    "zero":     ("Ballistics", "E6"),
    "sight_ht": ("Ballistics", "H6"),
    "twist":    ("Ballistics", "K6"),
    # Scope click value lives on Powder Charge Log (single source of truth — the
    # Ballistics click-column formulas all reference it). Showing it on
    # the printed card tells the shooter which scope click value the
    # click counts assume.
    "click":    ("Powder Charge Log", "G7"),
}

# DOPE table rows on the Ballistics tab. Data lives in rows 9..18
# (100yd through 1000yd in 100yd increments). Columns:
#   A=Range, B=Mils Elev, C=Mil Clicks, D=MOA Elev, E=MOA Clicks,
#   F=Wind Mils/10mph, G=Wind Mil Clk, H=Wind MOA/10mph, I=Wind MOA Clk,
#   J=TOF (sec)
DOPE_ROWS = range(9, 19)
DOPE_COL_KEYS = ["range", "mils_elev", "mils_clk", "moa_elev", "moa_clk",
                 "wind_mils", "wind_mils_clk", "wind_moa", "wind_moa_clk",
                 "tof"]
DOPE_COL_LETTERS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]


def _cell(wb, sheet, coord):
    if sheet not in wb.sheetnames:
        return None
    return wb[sheet][coord].value


def _fmt(v, blank="—"):
    """Format a cell value for display in the card. Returns blank glyph
    for None / empty / NaN."""
    if v is None:
        return blank
    if isinstance(v, str):
        return v.strip() or blank
    if isinstance(v, float):
        if v != v:  # NaN
            return blank
        if v == int(v):
            return str(int(v))
        # 2 decimal places, trimming trailing zeros
        return f"{v:.2f}".rstrip("0").rstrip(".")
    return str(v)


def _gather(workbook_path):
    """Read all card data from the workbook. Uses data_only=True so the
    LOOKUP formulas on the Ballistics tab evaluate to their cached values."""
    wb = load_workbook(workbook_path, data_only=True, keep_vba=False)
    if "Ballistics" not in wb.sheetnames:
        raise ValueError("This workbook doesn't have a Ballistics tab.")

    data = {"header": {}, "dope": []}
    for key, (sheet, coord) in HEADER_FIELDS.items():
        data["header"][key] = _cell(wb, sheet, coord)

    bal = wb["Ballistics"]
    rows_filled = 0
    for r in DOPE_ROWS:
        row = {"row_num": r}
        for letter, key in zip(DOPE_COL_LETTERS, DOPE_COL_KEYS):
            row[key] = bal[f"{letter}{r}"].value
        # A row is considered "filled" if there's any elev or wind data
        if any(row.get(k) not in (None, "") for k in
               ("mils_elev", "moa_elev", "wind_mils", "wind_moa")):
            rows_filled += 1
        data["dope"].append(row)

    # v0.14.5 (Chad): the printed card should show ONLY the unit the
    # shooter actually dials (Mil OR MOA), not both with the other half
    # blank. Derive it from the scope click value (e.g. "0.1 Mil",
    # "0.25 MOA"). Default to Mil for anything ambiguous/unknown.
    click_str = str(data["header"].get("click") or "").lower()
    data["unit"] = "moa" if "moa" in click_str else "mil"

    # Gate-4: fill rows the shooter hasn't confirmed yet with the LIVE
    # solver prediction so the card is useful before the first range
    # trip ("range card ready before you leave the truck"). Predicted
    # rows are flagged so the card carries the verify-at-range banner.
    # The workbook is NOT modified — predictions are display-only.
    elev_k = "moa_elev" if data["unit"] == "moa" else "mils_elev"
    wind_k = "wind_moa" if data["unit"] == "moa" else "wind_mils"
    data["has_predicted"] = False
    try:
        pred = dope_solver.predicted_dope(workbook_path)
    except Exception:
        pred = {"status": "error", "rows": {}}
    if pred.get("status") == dope_solver.OK:
        prows = pred.get("rows", {})
        for row in data["dope"]:
            r = row["row_num"]
            pr = prows.get(r)
            if not pr:
                continue
            # Drop / velocity / energy are solver REFERENCE values — not
            # something the shooter dials or confirms — so show them on
            # every row (confirmed or predicted alike).
            row["drop"] = pr.get("drop", "")
            row["velocity"] = pr.get("velocity", "")
            row["energy"] = pr.get("energy", "")
            confirmed = (row.get(elev_k) not in (None, "")
                         or row.get(wind_k) not in (None, ""))
            if confirmed:
                continue
            try:
                row[elev_k] = float(pr["elev"])
                row[wind_k] = float(pr["wind"])
            except (TypeError, ValueError, KeyError):
                continue
            row["predicted"] = True
            data["has_predicted"] = True

    if rows_filled == 0 and not data["has_predicted"]:
        hint = ("Fill in the elevation and wind values for at least one "
                "range (rows 9-18) before generating a Pocket Range "
                "Card.")
        if pred.get("status") == dope_solver.NO_BC:
            hint = ("Enter this bullet's G7 (or G1) ballistic "
                    "coefficient on the Range & DOPE screen and Loadscope "
                    "can predict a starting card for you.")
        raise ValueError(
            "Ballistics tab has no DOPE data yet. " + hint)
    elev_clk_k = "moa_clk" if data["unit"] == "moa" else "mils_clk"
    wind_clk_k = "wind_moa_clk" if data["unit"] == "moa" else "wind_mils_clk"
    # Only show a click-count column if it actually has values — an
    # all-"—" column is just noise on a pocket card.
    data["show_elev_clk"] = any(
        r.get(elev_clk_k) not in (None, "") for r in data["dope"])
    data["show_wind_clk"] = any(
        r.get(wind_clk_k) not in (None, "") for r in data["dope"])

    data["workbook_name"] = os.path.splitext(os.path.basename(workbook_path))[0]
    data["generated_at"] = datetime.now().strftime("%b %d, %Y")
    return data


def _build_html(d, layout="field"):
    """Render the pocket card HTML.

    layout:
      "card"  — bare 6x4 page, one card (used to render the website /
                demo-tour screenshot; never printed by users).
      "field" — US Letter page, TWO true-size cards pinned top-left
                with dashed cut guides (DEFAULT — what the app prints;
                a Letter page prints 1:1 on any home printer, unlike a
                6x4 page which printers shrink/float).
      "large" — US Letter page, ONE card scaled up to fill the sheet
                for desk/eyesight use.
    """
    h = d["header"]
    title_line = f"{_fmt(h.get('rifle'))} • {_fmt(h.get('bullet'))} • {_fmt(h.get('charge'), '—')}gr • {_fmt(h.get('vel'), '—')} fps"
    logo_uri = _logo_data_uri()
    logo_img = (
        f'<img src="{logo_uri}" alt="Loadscope" class="logo"/>'
        if logo_uri else ''
    )

    # v0.14.5: render ONLY the shooter's unit (Mil or MOA), and only
    # include click-count columns that actually have data.
    unit = d.get("unit", "mil")
    show_elev_clk = d.get("show_elev_clk", False)
    show_wind_clk = d.get("show_wind_clk", False)
    if unit == "moa":
        elev_k, elev_clk_k = "moa_elev", "moa_clk"
        wind_k, wind_clk_k = "wind_moa", "wind_moa_clk"
        unit_lbl = "MOA"
    else:
        elev_k, elev_clk_k = "mils_elev", "mils_clk"
        wind_k, wind_clk_k = "wind_mils", "wind_mils_clk"
        unit_lbl = "Mils"

    # DOPE table rows
    # Skip rows where range is empty (shouldn't happen — A9..A18 are
    # 100..1000); the LAST rendered row caps the dial box's bottom edge.
    renderable = [r for r in d["dope"] if r.get("range") not in (None, "")]
    last_i = len(renderable) - 1
    dope_rows_html = []
    for i, row in enumerate(renderable):
        b = " db" if i == last_i else ""
        cells = [f"<td class='r'>{_fmt(row.get('range'))}</td>",
                 f"<td class='dial dl{b}'>{_fmt(row.get(elev_k))}</td>"]
        if show_elev_clk:
            cells.append(f"<td class='dial clk{b}'>{_fmt(row.get(elev_clk_k))}</td>")
        wind_edge = "" if show_wind_clk else " dr"
        cells.append(f"<td class='dial{wind_edge}{b}'>{_fmt(row.get(wind_k))}</td>")
        if show_wind_clk:
            cells.append(f"<td class='dial clk dr{b}'>{_fmt(row.get(wind_clk_k))}</td>")
        cells.append(f"<td class='tof'>{_fmt(row.get('drop'))}</td>")
        cells.append(f"<td class='tof'>{_fmt(row.get('velocity'))}</td>")
        cells.append(f"<td class='tof'>{_fmt(row.get('energy'))}</td>")
        cells.append(f"<td class='tof'>{_fmt(row.get('tof'))}</td>")
        tr_cls = " class='pred'" if row.get("predicted") else ""
        dope_rows_html.append(f"<tr{tr_cls}>" + "".join(cells) + "</tr>")
    dope_rows_str = "\n".join(dope_rows_html)
    pred_banner = (
        '<div class="pred-banner">PREDICTED DOPE (italic rows) - '
        'estimated from your bullet, velocity &amp; atmosphere. '
        'VERIFY AT THE RANGE before relying on these values.</div>'
        if d.get("has_predicted") else "")

    # Dynamic header matching the columns we actually render.
    elev_span = 2 if show_elev_clk else 1
    wind_span = 2 if show_wind_clk else 1
    elev_unit_th = f'<th class="dial dl">{unit_lbl}</th>'
    elev_clk_th = '<th class="dial">clk</th>' if show_elev_clk else ''
    if show_wind_clk:
        wind_unit_th = f'<th class="dial">{unit_lbl}</th>'
        wind_clk_th = '<th class="dial dr">clk</th>'
    else:
        wind_unit_th = f'<th class="dial dr">{unit_lbl}</th>'
        wind_clk_th = ''
    thead_html = (
        '<tr>'
        '<th rowspan="2">Range<br>(yd)</th>'
        f'<th class="spanner dial dl dt" colspan="{elev_span}">Elev — {unit_lbl}</th>'
        f'<th class="spanner dial dr dt" colspan="{wind_span}">Wind 10 mph ({unit_lbl})</th>'
        '<th rowspan="2">Drop<br>(in)</th>'
        '<th rowspan="2">Vel<br>(fps)</th>'
        '<th rowspan="2">Energy<br>(ft&middot;lb)</th>'
        '<th rowspan="2">TOF<br>(s)</th>'
        '</tr><tr>'
        f'{elev_unit_th}{elev_clk_th}{wind_unit_th}{wind_clk_th}'
        '</tr>'
    )

    card_div = f'''<div class="card">
  <div class="title-bar">
    <div class="brand">
      {logo_img}
      <span>Loadscope<span class="tm">™</span></span>
    </div>
    <div class="subtitle">Pocket Range Card</div>
  </div>
  <div class="title-line">{title_line}</div>
  <div class="setup">
    <div class="field"><span class="label">Scope:</span><span>{_fmt(h.get('scope'))}</span></div>
    <div class="field"><span class="label">Click:</span><span>{_fmt(h.get('click'))}</span></div>
    <div class="field"><span class="label">Zero:</span><span>{_fmt(h.get('zero'))} yd</span></div>
    <div class="field"><span class="label">Sight Ht:</span><span>{_fmt(h.get('sight_ht'))} in</span></div>
    <div class="field"><span class="label">Twist:</span><span>{_fmt(h.get('twist'))}</span></div>
  </div>
  {pred_banner}
  <table>
    <thead>
      {thead_html}
    </thead>
    <tbody>
      {dope_rows_str}
    </tbody>
  </table>
  <div class="footer">Generated {d['generated_at']}</div>
</div>'''

    if layout == "card":
        # Bare 6x4 — for the website / demo-tour screenshot only.
        page_rule = "@page { size: 6in 4in; margin: 0.15in; }"
        layout_css = ""
        body_inner = card_div
    elif layout == "large":
        # One card scaled up to fill a Letter sheet (desk / eyesight).
        page_rule = "@page { size: letter; margin: 0.45in; }"
        layout_css = ".big { zoom: 1.30; }"
        body_inner = f'<div class="big">{card_div}</div>'
    else:  # "field" (default): Letter, TWO true-size cards, top-left.
        page_rule = "@page { size: letter; margin: 0.3in; }"
        layout_css = (".cut { outline: 0.5pt dashed #999; "
                      "outline-offset: 4pt; margin: 0 0 0.5in 0; }")
        body_inner = (f'<div class="cut">{card_div}</div>'
                      f'<div class="cut">{card_div}</div>')

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Pocket Range Card — {_fmt(h.get('rifle'))}</title>
<style>
  {page_rule}
  * {{ box-sizing: border-box; }}
  html, body {{
    margin: 0;
    padding: 0;
    font-family: -apple-system, "Helvetica Neue", Helvetica, Arial, sans-serif;
    color: #000;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
  }}
  .card {{
    width: 5.8in;
    min-height: 3.6in;
    padding: 0.1in;
    border: 1.5pt solid #000;
  }}
  .title-bar {{
    background: #d97706;
    color: #fff;
    padding: 7pt 10pt;
    margin: -0.1in -0.1in 0.07in -0.1in;
    display: flex;
    justify-content: space-between;
    align-items: center;
    border-bottom: 1pt solid #b45309;
  }}
  .title-bar .brand {{
    display: flex;
    align-items: center;
    gap: 6pt;
    font-weight: 700;
    font-size: 12pt;
    letter-spacing: 0.3pt;
  }}
  .title-bar .brand .logo {{
    width: 22pt;
    height: 22pt;
    display: block;
  }}
  .title-bar .brand .tm {{
    font-size: 7pt;
    font-weight: 600;
    vertical-align: super;
    margin-left: 1pt;
  }}
  .title-bar .subtitle {{
    font-size: 9pt;
    font-weight: 700;
    letter-spacing: 0.5pt;
    text-transform: uppercase;
  }}
  .title-line {{
    font-size: 9pt;
    font-weight: 600;
    margin: 3pt 0 2pt;
    text-align: center;
  }}
  .setup {{
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 2pt 8pt;
    font-size: 7pt;
    margin-bottom: 4pt;
    padding-bottom: 3pt;
    border-bottom: 0.5pt solid #aaa;
  }}
  .setup .field {{ display: flex; gap: 3pt; }}
  .setup .label {{ color: #555; font-weight: 600; }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 9pt;
    font-variant-numeric: tabular-nums;
  }}
  th {{
    background: #f0eee9;
    font-weight: 700;
    font-size: 7.5pt;
    padding: 2pt 1pt;
    text-align: center;
    border-bottom: 0.8pt solid #555;
    line-height: 1.05;
  }}
  th.spanner {{
    background: #1f2229;
    color: #fff;
    border-bottom: 0;
    border-right: 0.5pt solid #555;
    padding: 1pt;
  }}
  th.spanner:last-child {{ border-right: 0; }}
  td {{
    padding: 1.4pt 1pt;
    text-align: center;
    border-bottom: 0.3pt solid #ddd;
  }}
  td.r {{
    font-weight: 700;
    background: #faf8f4;
  }}
  td.clk {{
    color: #777;
    font-size: 7.5pt;
    font-style: italic;
  }}
  td.tof {{
    color: #555;
    font-size: 7.5pt;
  }}
  tr.pred td {{
    font-style: italic;
    color: #444;
  }}
  .pred-banner {{
    margin: 2pt 0 3pt;
    padding: 2pt 4pt;
    border: 0.8pt solid #000;
    background: #f3efe6;
    font-size: 6pt;
    font-weight: 700;
    text-align: center;
    line-height: 1.15;
  }}
  .footer {{
    margin-top: 3pt;
    font-size: 6pt;
    color: #888;
    text-align: center;
  }}
  /* Elevation + Wind = the values the shooter actually dials. Box them
     so they stand out from the reference columns (drop/vel/energy/tof). */
  .dial {{ background: #fff5e6; }}
  .dl {{ border-left: 1.7pt solid #000 !important; }}
  .dr {{ border-right: 1.7pt solid #000 !important; }}
  .dt {{ border-top: 1.7pt solid #000 !important; }}
  .db {{ border-bottom: 1.7pt solid #000 !important; }}
  {layout_css}
</style>
</head>
<body>
{body_inner}
</body>
</html>"""


def generate_pocket_card(workbook_path, open_after=True, layout="field"):
    """Generate a Pocket Range Card HTML from the workbook's Ballistics tab.
    Returns the path to the generated file. If `open_after`, opens it in
    the user's default browser so they can print it.

    layout defaults to "field" (US Letter, two true-size 4x6 cards,
    top-left, cut guides) so the app's Print path lands 1:1 on a normal
    home printer. "large" = one big card per Letter sheet (desk). "card"
    = bare 6x4 (used only to render the website/demo screenshot)."""
    data = _gather(workbook_path)
    html = _build_html(data, layout=layout)

    # Pick a writable output directory. Default = "Range Cards" next to
    # the workbook (works for normal user workbooks in their project
    # folder). Fallback = ~/Documents/Loadscope Range Cards/ when the
    # workbook lives in a read-only location like the bundled demo
    # workbook at /Applications/Loadscope.app/Contents/Resources/.
    # v0.14.1 fix: previously this would crash with PermissionError when
    # demo users clicked Print Pocket Range Card.
    project_dir = os.path.dirname(os.path.abspath(workbook_path))
    out_dir = os.path.join(project_dir, "Range Cards")
    try:
        os.makedirs(out_dir, exist_ok=True)
        # Also test writability — makedirs can succeed on a path that's
        # readable but not writable (e.g., something already exists).
        _probe = os.path.join(out_dir, ".loadscope_writable_probe")
        with open(_probe, "w") as _pf:
            _pf.write("")
        os.remove(_probe)
    except (PermissionError, OSError):
        # Fall back to user's Documents folder
        fallback = os.path.expanduser("~/Documents/Loadscope Range Cards")
        os.makedirs(fallback, exist_ok=True)
        out_dir = fallback

    workbook_base = os.path.splitext(os.path.basename(workbook_path))[0]
    stamp = datetime.now().strftime("%Y-%m-%d %H-%M")
    out_path = os.path.join(out_dir, f"{workbook_base} — Pocket Card {stamp}.html")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    if open_after:
        subprocess.run(["open", out_path])

    return out_path
