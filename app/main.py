"""
Loadscope — GUI app.

A small window with a drop zone. When CSVs are dropped:
    1. Each file is auto-classified (Garmin vs BallisticX) by sniffing its first lines.
    2. The file is copied into the correct import folder (Garmin Imports/ or BallisticX Imports/).
    3. The full import runs and the workbook opens in Excel filled out.

Run from the project folder:
    /usr/bin/python3 app/main.py
"""

import io
import os
import re
import shutil
import subprocess
import sys
import traceback
from contextlib import redirect_stdout
from pathlib import Path

try:
    from PyQt5.QtCore import QEvent, Qt, QTimer, QUrl, QPointF, QRectF
    from PyQt5.QtGui import (
        QBrush,
        QColor,
        QDesktopServices,
        QFont,
        QPainter,
        QPainterPath,
        QPen,
        QRadialGradient,
    )
    from PyQt5.QtWidgets import (
        QAction,
        QApplication,
        QComboBox,
        QFrame,
        QGraphicsDropShadowEffect,
        QHBoxLayout,
        QLabel,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
except ImportError:
    print("ERROR: PyQt5 is not installed.")
    print("Install with:  /usr/bin/python3 -m pip install --user PyQt5")
    sys.exit(1)

# Make the local package importable
HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
sys.path.insert(0, REPO_ROOT)

from parsers import detect_parser, chronograph_parsers, group_parsers
import import_data
import config as app_config
from setup_wizard import run_wizard
from updater import UpdateChecker, UpdateDownloader, DEFAULT_MANIFEST_URL, resolve_download_url
import installer
from version import APP_NAME, APP_VERSION
from disclaimer import needs_disclaimer, show_disclaimer, view_disclaimer
import license as app_license
from license_dialog import show_license_dialog
from settings_dialog import show_settings
from help_dialog import show_help
from new_cycle_dialog import show_new_cycle
from welcome_tutorial import needs_tutorial, show_tutorial
import theme


# Default window size — large enough that all UI elements have room to
# breathe and the user has space to drag CSVs in. Saved+restored across
# launches via QSettings (see _restore_window_state / _save_window_state).
WINDOW_W = 960
WINDOW_H = 760


def parse_loadscope_action(url):
    """Extract the action token from a loadscope:// QUrl.

    Returns the action as a lowercase string (e.g. 'reset-weights') or
    '' for an invalid / actionless URL. Accepts both `loadscope://action`
    (action in host) and `loadscope:///action` (action in path) shapes.

    Pure function — safe to unit-test without instantiating QApplication.
    """
    if url is None or not url.isValid() or url.scheme() != "loadscope":
        return ""
    return (url.host() or url.path().strip("/") or "").strip().lower()


class RifleLoadApp(QApplication):
    """Custom QApplication that intercepts macOS QFileOpenEvent so CSVs
    dragged onto the Dock icon (or right-clicked → Open With) are routed
    through to the main window's external-files handler.

    Files received before the main window is set are queued and flushed
    once `set_main_window` is called.

    Multiple file events arriving in rapid succession (drag of 3 files
    triggers 3 separate events on macOS) are batched via a 2-second debounce
    so they import as a single batch. Users are instructed to grab all the
    CSVs they want in a single Finder selection, then drag the whole batch
    onto the icon at once. The 2-second window also tolerates slightly
    delayed drops (e.g., from network drives or AirDrop completing).
    """

    BATCH_DEBOUNCE_MS = 2000

    def __init__(self, argv):
        super().__init__(argv)
        self._main_window = None
        self._pending_files = []
        # URLs (loadscope:// scheme) received before the main window is
        # ready get queued and dispatched once set_main_window fires.
        self._pending_urls = []
        self._batch_timer = QTimer(self)
        self._batch_timer.setSingleShot(True)
        self._batch_timer.timeout.connect(self._flush_pending)

    def event(self, e):
        if e.type() == QEvent.FileOpen:
            # QFileOpenEvent carries either a file path (CSV drag-on-icon)
            # or a URL (custom loadscope:// scheme clicked from Excel etc.).
            # Route accordingly.
            url = e.url()
            if url is not None and url.isValid() and url.scheme() == "loadscope":
                if self._main_window:
                    self._main_window.handle_loadscope_url(url)
                else:
                    self._pending_urls.append(url)
                return True
            path = e.file()
            if path:
                self._pending_files.append(path)
                self._batch_timer.start(self.BATCH_DEBOUNCE_MS)
            return True
        return super().event(e)

    def set_main_window(self, win):
        self._main_window = win
        # Flush any URLs that arrived before the window was ready.
        for url in self._pending_urls:
            win.handle_loadscope_url(url)
        self._pending_urls = []

    def _flush_pending(self):
        if not self._pending_files or not self._main_window:
            return
        files = self._pending_files
        self._pending_files = []
        self._main_window.handle_external_files(files)


class CarbonBackground(QWidget):
    """Central widget that paints a carbon-fiber twill pattern as its background.
    Inner widgets keep their own styled backgrounds; the carbon shows through
    in the gaps and around the edges."""

    def __init__(self):
        super().__init__()
        self._tile = theme.generate_carbon_tile()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawTiledPixmap(self.rect(), self._tile)
        painter.end()


class HeroCrosshair(QWidget):
    """Big precision-rifle crosshair painted as the dropzone hero icon.

    Concentric rings + full-diameter crosshair + mil-dots + accent center.
    Brightens to accent color on hover (set via set_hover()).
    """

    def __init__(self, parent=None, size=92):
        super().__init__(parent)
        self.setFixedSize(size, size)
        # Mouse events pass through to the parent QFrame so drag-over still hits
        self.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self._hover = False

    def set_hover(self, on):
        self._hover = on
        self.update()

    def paintEvent(self, _):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        s = self.width()
        cx = cy = s / 2
        fg = QColor(theme.ACCENT if self._hover else theme.TEXT_PRIMARY)
        accent = QColor(theme.ACCENT)

        # Outer ring (heavy)
        p.setPen(QPen(fg, 2.2))
        p.setBrush(Qt.NoBrush)
        p.drawEllipse(QPointF(cx, cy), s * 0.46, s * 0.46)

        # Two inner rings (lighter)
        ring = QColor(fg)
        ring.setAlpha(140)
        p.setPen(QPen(ring, 1.4))
        p.drawEllipse(QPointF(cx, cy), s * 0.32, s * 0.32)
        p.drawEllipse(QPointF(cx, cy), s * 0.18, s * 0.18)

        # Crosshair (full diameter, slim)
        p.setPen(QPen(fg, 1.6))
        tick = s * 0.50
        p.drawLine(QPointF(cx - tick, cy), QPointF(cx + tick, cy))
        p.drawLine(QPointF(cx, cy - tick), QPointF(cx, cy + tick))

        # Mil-dots along the crosshair arms
        p.setBrush(QBrush(fg))
        p.setPen(Qt.NoPen)
        for d in (s * 0.18, s * 0.32):
            for dx, dy in ((-d, 0), (d, 0), (0, -d), (0, d)):
                p.drawEllipse(QPointF(cx + dx, cy + dy), 1.6, 1.6)

        # Accent center dot — always at full accent for the "you are here" pop
        p.setBrush(QBrush(accent))
        p.drawEllipse(QPointF(cx, cy), 3.5, 3.5)
        p.end()


class DropZone(QFrame):
    """Refined-card drop zone with a hero precision crosshair.

    Layout: hero crosshair → "Drop your CSV files here" title → file-type
    chips (Garmin Xero · BallisticX) → footer hint. No dashed border —
    solid card surface with a soft drop shadow for depth.
    """

    def __init__(self, on_drop, on_chip_click=None, on_import_click=None,
                 chip_labels=None):
        super().__init__()
        self.on_drop = on_drop
        self.on_chip_click = on_chip_click   # opens import folder in Finder
        self.on_import_click = on_import_click  # runs the import
        # Chip labels are data-driven so adding a new parser (LabRadar,
        # MagnetoSpeed, etc.) auto-populates the dropzone with no UI edit.
        # Callers should pass [parser.NAME for parser in chronograph_parsers()
        #                      + group_parsers()].
        if chip_labels is None:
            chip_labels = ("Garmin Xero", "BallisticX")  # fallback for tests
        self._chip_labels = list(chip_labels)
        self._hovering = False
        self.setObjectName("DropZone")
        self.setAcceptDrops(True)
        self.setMinimumHeight(300)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 24, 20, 22)
        layout.setSpacing(10)
        layout.setAlignment(Qt.AlignCenter)

        # Hero crosshair
        self.icon = HeroCrosshair(self, size=92)
        layout.addWidget(self.icon, alignment=Qt.AlignCenter)

        # Title
        self.title = QLabel("Drop your CSV files here")
        self.title.setObjectName("DropZoneTitle")
        self.title.setAlignment(Qt.AlignCenter)
        self.title.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        layout.addWidget(self.title)

        # Chip row — file-type indicators. Clickable when on_chip_click
        # callback is supplied: clicking opens the import folder for that
        # device in Finder so users can quickly grab their CSVs.
        chip_row = QHBoxLayout()
        chip_row.setAlignment(Qt.AlignCenter)
        chip_row.setSpacing(8)
        self._chips = []
        for label_text in self._chip_labels:
            chip = QLabel(label_text)
            chip.setObjectName("DropZoneChip")
            chip.setAlignment(Qt.AlignCenter)
            if on_chip_click:
                chip.setCursor(Qt.PointingHandCursor)
                chip.setToolTip(
                    f"Open the {label_text} import folder in Finder"
                )
                # Bind label_text into the lambda's default-arg scope
                chip.mousePressEvent = (
                    lambda e, name=label_text: on_chip_click(name)
                )
            else:
                chip.setAttribute(Qt.WA_TransparentForMouseEvents, True)
            chip_row.addWidget(chip)
            self._chips.append(chip)
        layout.addLayout(chip_row)

        # In-card CTA button — visible only when files are staged.
        # Replaces the external Run Import button; clicking calls
        # on_import_click. Hidden by default.
        self.cta_button = QPushButton("Run Import →")
        self.cta_button.setObjectName("DropZoneCTA")
        self.cta_button.setCursor(Qt.PointingHandCursor)
        self.cta_button.setFixedHeight(40)
        self.cta_button.setMinimumWidth(180)
        self.cta_button.hide()
        if on_import_click:
            self.cta_button.clicked.connect(on_import_click)
        cta_row = QHBoxLayout()
        cta_row.setAlignment(Qt.AlignCenter)
        cta_row.addWidget(self.cta_button)
        layout.addLayout(cta_row)
        self._chip_row = chip_row  # remember so we can hide/show

        # Footer hint
        self.footer = QLabel("Drop multiple at once · format auto-detected")
        self.footer.setObjectName("DropZoneFooter")
        self.footer.setAlignment(Qt.AlignCenter)
        self.footer.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        layout.addWidget(self.footer)

        # Soft drop shadow for the card (gives the "lifted off the surface" feel)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(22)
        shadow.setOffset(0, 4)
        shadow.setColor(QColor(0, 0, 0, 90))
        self.setGraphicsEffect(shadow)

        self._set_idle_style()

    def _set_idle_style(self):
        self._hovering = False
        self.icon.set_hover(False)
        self.setStyleSheet(theme.dropzone_idle_stylesheet())

    def _set_hover_style(self):
        self._hovering = True
        self.icon.set_hover(True)
        self.setStyleSheet(theme.dropzone_hover_stylesheet())

    def set_staged_state(self, count, breakdown_text=""):
        """Update the title + footer to reflect staged-files state.

        count=0 → neutral state ('Drop your CSV files here' + format hint).
        count>0 → accent state ('✓ X files staged' + 'Click Run Import below ↓').
        """
        if count <= 0:
            # Neutral state: chips visible, CTA hidden, format hint footer.
            self.title.setText("Drop your CSV files here")
            self.title.setStyleSheet("")  # fall back to parent QSS rule
            self.footer.setText("Drop multiple at once · format auto-detected")
            for chip in self._chips:
                chip.show()
            self.cta_button.hide()
        else:
            # Staged state: chips hidden, CTA shown, accent title.
            suffix = "s" if count != 1 else ""
            fallback = f"{count} file{suffix} staged"
            text = breakdown_text or fallback
            self.title.setText(f"✓  {text}")
            # Inline stylesheet REPLACES parent QSS, so include font props.
            self.title.setStyleSheet(
                f"color: {theme.ACCENT}; font-size: 18pt; font-weight: 600;"
                " background: transparent; border: none; padding: 0;"
            )
            self.footer.setText("Ready to build your workbook")
            for chip in self._chips:
                chip.hide()
            self.cta_button.show()

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            self._set_hover_style()
            e.acceptProposedAction()

    def dragLeaveEvent(self, _):
        self._set_idle_style()

    def dropEvent(self, e):
        self._set_idle_style()
        paths = []
        for url in e.mimeData().urls():
            local = url.toLocalFile()
            if local:
                paths.append(local)
        if paths:
            self.on_drop(paths)


class MainWindow(QMainWindow):
    def __init__(self, project_folder):
        super().__init__()
        self.project = str(project_folder)
        self.setWindowTitle(f"{APP_NAME}  v{APP_VERSION}")

        # Restore window size+position from config if we have one saved,
        # otherwise default to a generous size.
        cfg_for_window = app_config.load_config()
        saved_geom = cfg_for_window.get("window_geometry")  # [x, y, w, h]
        if saved_geom and isinstance(saved_geom, list) and len(saved_geom) == 4:
            self.setGeometry(*saved_geom)
        else:
            self.resize(WINDOW_W, WINDOW_H)

        # Per-session staging counter — reset after each import
        self.staged_garmin = 0
        self.staged_bx = 0

        # Holds the last manifest from a successful update check
        self._last_manifest = None
        self._update_checker = None

        # In-app updater state. The full flow:
        #   _pending_app_update_url   — set when the manifest reports a newer
        #     version; used by the Install Update button and the manual fallback
        #   _pending_app_update_version — display string for the new version
        #   _update_downloader        — QThread streaming the .zip
        #   _downloaded_zip_path      — local path once download is complete;
        #     consumed by installer.launch_install_swap on Quit-and-Install
        self._pending_app_update_url = None
        self._pending_app_update_version = None
        self._pending_app_update_manifest = None  # full manifest for resolve_download_url
        self._pending_app_update_website = None   # fallback link to send users to
        self._update_downloader = None
        self._downloaded_zip_path = None

        # Carbon-fiber background is built (CarbonBackground class) but
        # disabled for now — Chad wants a flat graphite background while
        # we finish the rest. We'll re-enable carbon during final polish.
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(theme.PAD_WINDOW, theme.PAD_WINDOW,
                                  theme.PAD_WINDOW, theme.PAD_WINDOW)
        layout.setSpacing(theme.GAP_LARGE)

        # v0.14 demo-mode banner — shows ONLY when the app is running unlicensed
        # (no valid key). Persistent, never auto-dismisses. Has a "Purchase"
        # button that opens loadscope.app/.
        if app_license.is_demo_mode():
            demo_banner = self._build_demo_banner()
            layout.addWidget(demo_banner)

        # Tools-menu banner — points new users at the menu bar (which is in
        # macOS's system menu, easy to overlook). Dismissed by clicking the X;
        # acceptance saved to config so it doesn't keep popping back.
        cfg_for_banner = app_config.load_config()
        if not cfg_for_banner.get("tools_banner_dismissed"):
            self.tools_banner = QLabel(
                "↑  More options live in the <b>Tools</b> menu in the macOS menu bar "
                "above this window — How to Use, Settings, Generate Load Card, "
                "and more.   "
                "<a href='dismiss:tools-banner' style='color:#888;'>(dismiss)</a>"
            )
            self.tools_banner.setOpenExternalLinks(False)
            self.tools_banner.setWordWrap(True)
            self.tools_banner.setTextInteractionFlags(Qt.LinksAccessibleByMouse)
            self.tools_banner.linkActivated.connect(self._on_tools_banner_link)
            self.tools_banner.setStyleSheet(
                f"QLabel {{ "
                f"  background-color: {theme.BG_SURFACE}; "
                f"  border: 1px solid {theme.BORDER_SUBTLE}; "
                f"  border-radius: 6px; "
                f"  padding: 10px 14px; "
                f"  color: {theme.TEXT_PRIMARY}; "
                f"  font-size: {theme.FONT_SIZE_SMALL}px; "
                f"}}"
            )
            layout.addWidget(self.tools_banner)
        else:
            self.tools_banner = None

        # Update banner — hidden by default, shown when an update is found
        self.update_banner = QLabel()
        self.update_banner.setVisible(False)
        self.update_banner.setOpenExternalLinks(False)
        self.update_banner.setWordWrap(True)
        self.update_banner.setTextInteractionFlags(Qt.LinksAccessibleByMouse)
        self.update_banner.linkActivated.connect(self._on_update_link)
        self.update_banner.setStyleSheet(theme.update_banner_stylesheet())
        layout.addWidget(self.update_banner)

        # Project folder is intentionally hidden from the main window — friends
        # don't need to see the underlying file path. The path is still tracked
        # internally as self.project and used for all file operations.

        # Workbook picker — lets users with multiple cartridge workbooks choose
        # which one to import into. Hidden when only one workbook exists so it
        # doesn't add visual noise for single-workbook users.
        wb_row = QHBoxLayout()
        wb_row.setSpacing(8)
        self.wb_picker_label = QLabel("Importing into:")
        self.wb_picker_label.setStyleSheet(
            f"color: {theme.TEXT_TERTIARY}; "
            f"font-size: {theme.FONT_SIZE_TINY}px; "
            f"letter-spacing: 0.6px; text-transform: uppercase;"
        )
        wb_row.addWidget(self.wb_picker_label)
        self.wb_picker = QComboBox()
        self.wb_picker.setToolTip(
            "Pick which workbook to import data into. Useful if you're "
            "developing loads for multiple cartridges. The choice is "
            "remembered between launches."
        )
        self.wb_picker.setStyleSheet(
            f"QComboBox {{ "
            f"  background-color: {theme.BG_ELEVATED}; "
            f"  color: {theme.TEXT_PRIMARY}; "
            f"  border: 1px solid {theme.BORDER_MEDIUM}; "
            f"  padding: 4px 8px; border-radius: 6px; "
            f"  font-size: {theme.FONT_SIZE_BODY}px; "
            f"}}"
        )
        self.wb_picker.currentIndexChanged.connect(self._on_workbook_changed)
        wb_row.addWidget(self.wb_picker, stretch=1)
        # Refresh button to rescan the folder for newly-added workbooks
        self.wb_refresh_btn = QPushButton("Refresh")
        self.wb_refresh_btn.setMaximumWidth(90)
        self.wb_refresh_btn.setToolTip(
            "Re-scan the project folder for working .xlsx files. Click this "
            "after creating a new workbook from the template."
        )
        self.wb_refresh_btn.clicked.connect(self._refresh_workbooks)
        wb_row.addWidget(self.wb_refresh_btn)
        self.wb_row_widget = QWidget()
        self.wb_row_widget.setLayout(wb_row)
        layout.addWidget(self.wb_row_widget)
        # Defer initial population until the rest of __init__ is done
        # (after layout is ready)
        QTimer.singleShot(0, self._refresh_workbooks)

        # Tools menu — manual "Check for Updates", folder shortcuts, About.
        # Menu bar organized into four top-level menus instead of one giant
        # "Tools" menu. macOS-specific: avoid the literal name "Help" — the
        # OS auto-merges any menu by that name into its system Help menu and
        # can hide custom items. "Support" sidesteps that.
        mbar = self.menuBar()

        # v0.14: menu order is Settings → Workbook → Folders → Support
        # (Chad 2026-05-14: Settings should be first user menu, before
        # Folders). The addMenu calls below establish the bar order;
        # actions are added to each menu further down where each section
        # block lives, so the existing action-setup code stays put.
        settings_menu = mbar.addMenu("Settings")

        # ===== 1. WORKBOOK =====
        wb_menu = mbar.addMenu("Workbook")

        open_wb_action = QAction("Open Workbook in Excel", self)
        open_wb_action.setMenuRole(QAction.NoRole)
        open_wb_action.triggered.connect(self._open_workbook_in_excel)
        wb_menu.addAction(open_wb_action)
        self._open_wb_action = open_wb_action  # store so we can enable/disable

        print_wb_action = QAction("Print Workbook…", self)
        print_wb_action.setMenuRole(QAction.NoRole)
        print_wb_action.triggered.connect(self._print_workbook)
        wb_menu.addAction(print_wb_action)
        self._print_wb_action = print_wb_action

        wb_menu.addSeparator()

        new_cycle_action = QAction("Start New Cycle…", self)
        new_cycle_action.setMenuRole(QAction.NoRole)
        new_cycle_action.triggered.connect(self._start_new_cycle)
        wb_menu.addAction(new_cycle_action)

        wb_menu.addSeparator()

        run_now_action = QAction("Run Import Now (use existing folder contents)", self)
        run_now_action.setMenuRole(QAction.NoRole)
        run_now_action.triggered.connect(self._run_import_now_menu)
        wb_menu.addAction(run_now_action)

        restore_action = QAction("Restore From Backup…", self)
        restore_action.setMenuRole(QAction.NoRole)
        restore_action.triggered.connect(self._restore_from_backup)
        wb_menu.addAction(restore_action)

        wb_menu.addSeparator()

        load_card_action = QAction("Generate Load Card…", self)
        load_card_action.setMenuRole(QAction.NoRole)
        load_card_action.triggered.connect(self._generate_load_card)
        wb_menu.addAction(load_card_action)

        export_load_action = QAction("Export Suggested Load…", self)
        export_load_action.setMenuRole(QAction.NoRole)
        export_load_action.triggered.connect(self._export_load)
        wb_menu.addAction(export_load_action)

        import_load_action = QAction("Import Shared Load…", self)
        import_load_action.setMenuRole(QAction.NoRole)
        import_load_action.triggered.connect(self._import_load)
        wb_menu.addAction(import_load_action)

        reset_weights_action = QAction("Reset Composite Weights…", self)
        reset_weights_action.setMenuRole(QAction.NoRole)
        reset_weights_action.triggered.connect(self._reset_composite_weights)
        wb_menu.addAction(reset_weights_action)

        save_load_action = QAction("Save Suggested Load to Library…", self)
        save_load_action.setMenuRole(QAction.NoRole)
        save_load_action.triggered.connect(self._save_suggested_load)
        wb_menu.addAction(save_load_action)

        pocket_card_action = QAction("Print Pocket Range Card…", self)
        pocket_card_action.setMenuRole(QAction.NoRole)
        pocket_card_action.triggered.connect(self._print_pocket_card)
        wb_menu.addAction(pocket_card_action)

        wb_menu.addSeparator()

        demo_tour_action = QAction("Replay the Demo Tour…", self)
        demo_tour_action.setMenuRole(QAction.NoRole)
        demo_tour_action.triggered.connect(self._open_demo_tour)
        wb_menu.addAction(demo_tour_action)

        # ===== 2. FOLDERS =====
        folders_menu = mbar.addMenu("Folders")

        show_project_action = QAction("Show Project Folder", self)
        show_project_action.setMenuRole(QAction.NoRole)
        show_project_action.triggered.connect(
            lambda: self._reveal_in_finder(self.project)
        )
        folders_menu.addAction(show_project_action)

        show_garmin_action = QAction("Show Garmin Imports", self)
        show_garmin_action.setMenuRole(QAction.NoRole)
        show_garmin_action.triggered.connect(
            lambda: self._reveal_in_finder(os.path.join(self.project, "Garmin Imports"))
        )
        folders_menu.addAction(show_garmin_action)

        show_bx_action = QAction("Show BallisticX Imports", self)
        show_bx_action.setMenuRole(QAction.NoRole)
        show_bx_action.triggered.connect(
            lambda: self._reveal_in_finder(os.path.join(self.project, "BallisticX Imports"))
        )
        folders_menu.addAction(show_bx_action)

        show_backups_action = QAction("Show Backups", self)
        show_backups_action.setMenuRole(QAction.NoRole)
        show_backups_action.triggered.connect(
            lambda: self._reveal_in_finder(os.path.join(self.project, ".backups"))
        )
        folders_menu.addAction(show_backups_action)

        # ===== 3. SETTINGS ===== (menu created above; just adding actions here)
        settings_action = QAction("Settings…", self)
        settings_action.setMenuRole(QAction.NoRole)
        settings_action.triggered.connect(self._show_settings)
        settings_menu.addAction(settings_action)

        check_action = QAction("Check for Updates…", self)
        check_action.setMenuRole(QAction.NoRole)
        check_action.triggered.connect(lambda: self._start_update_check(manual=True))
        settings_menu.addAction(check_action)

        # ===== 4. SUPPORT =====
        support_menu = mbar.addMenu("Support")

        help_action = QAction(f"How to Use {APP_NAME}…", self)
        help_action.setMenuRole(QAction.NoRole)
        help_action.triggered.connect(self._show_help)
        support_menu.addAction(help_action)

        # v0.14: dedicated FAQ dialog (Chad 2026-05-14: commercial-product
        # polish — sits in Support menu next to How to Use).
        faq_action = QAction("Frequently Asked Questions…", self)
        faq_action.setMenuRole(QAction.NoRole)
        faq_action.triggered.connect(self._show_faq)
        support_menu.addAction(faq_action)

        support_menu.addSeparator()

        visit_website_action = QAction(f"Visit {APP_NAME} Website…", self)
        visit_website_action.setMenuRole(QAction.NoRole)
        visit_website_action.triggered.connect(self._visit_website)
        support_menu.addAction(visit_website_action)

        contact_support_action = QAction("Contact Support…", self)
        contact_support_action.setMenuRole(QAction.NoRole)
        contact_support_action.triggered.connect(self._send_feedback)
        support_menu.addAction(contact_support_action)

        support_menu.addSeparator()

        about_action = QAction("About Loadscope™", self)
        about_action.setMenuRole(QAction.NoRole)
        about_action.triggered.connect(self._show_about)
        support_menu.addAction(about_action)

        disclaimer_action = QAction("View Disclaimer…", self)
        disclaimer_action.setMenuRole(QAction.NoRole)
        disclaimer_action.triggered.connect(self._show_disclaimer)
        support_menu.addAction(disclaimer_action)

        # Wire up enable/disable for "Open Workbook in Excel" based on whether
        # there's actually a workbook to open. Updates whenever the picker
        # selection changes.
        self._update_open_workbook_action_state()

        # Drop zone
        # Chip labels come from the live parser registry so adding a new
        # parser (LabRadar, MagnetoSpeed, etc.) auto-extends the chips
        # with zero UI work.
        chip_names = [p.NAME for p in chronograph_parsers()] + \
                     [p.NAME for p in group_parsers()]
        self.drop = DropZone(self.handle_drops,
                             on_chip_click=self._open_import_folder_for_chip,
                             on_import_click=self.run_import_clicked,
                             chip_labels=chip_names)
        self.drop.setToolTip(
            "Drag your Garmin and BallisticX CSVs here. "
            "Loadscope detects the format automatically and routes each file "
            "to the right import folder. Click Run Import when you've dropped everything."
        )
        layout.addWidget(self.drop)

        # Supported-devices caption — pulled live from the parser registry so
        # adding a new parser automatically updates this label.
        all_names = [p.NAME for p in chronograph_parsers()] + \
                    [p.NAME for p in group_parsers()]
        if all_names:
            supports_text = (
                "Supports " + " · ".join(all_names) +
                "    Have a different chronograph? Email support@loadscope.app with a sample CSV "
                "and we'll add it in the next update."
            )
        else:
            supports_text = "(no parsers registered)"
        self.supports_label = QLabel(supports_text)
        self.supports_label.setAlignment(Qt.AlignCenter)
        self.supports_label.setStyleSheet(
            f"color: {theme.TEXT_TERTIARY}; "
            f"font-size: {theme.FONT_SIZE_TINY}px; "
            f"letter-spacing: 0.4px;"
        )
        self.supports_label.setWordWrap(True)
        layout.addWidget(self.supports_label)

        # Status + button row
        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        self.status_label = QLabel("0 files staged")
        self.status_label.setStyleSheet(theme.status_label_stylesheet())
        button_row.addWidget(self.status_label)
        button_row.addStretch(1)

        self.clear_button = QPushButton("Clear")
        self.clear_button.setEnabled(False)
        self.clear_button.setToolTip(
            "Reset the staged file count. Doesn't delete your CSVs — they "
            "stay in the import folders and will still be picked up next time."
        )
        self.clear_button.clicked.connect(self.clear_staged)
        button_row.addWidget(self.clear_button)

        # External Run Import button — kept as a hidden fallback for keyboard
        # default-button behavior (Enter key triggers it). The visible CTA
        # now lives INSIDE the DropZone card (Chad 2026-05-14: cleaner,
        # matches Stripe/Notion/Linear upload patterns).
        self.go_button = QPushButton("Run Import")
        self.go_button.setObjectName("primary")
        self.go_button.setEnabled(False)
        self.go_button.setDefault(True)
        self.go_button.setToolTip(
            "Read every CSV in your import folders, write the data into your "
            "active workbook, and open the workbook in Excel."
        )
        self.go_button.clicked.connect(self.run_import_clicked)
        self.go_button.hide()  # CTA moved into the DropZone card
        button_row.addWidget(self.go_button)

        layout.addLayout(button_row)

        # Log area
        log_label = QLabel("ACTIVITY")
        log_label.setStyleSheet(theme.section_label_stylesheet())
        layout.addWidget(log_label)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        mono = QFont()
        mono.setFamily("SF Mono")
        mono.setPointSize(11)
        self.log.setFont(mono)
        layout.addWidget(self.log, stretch=1)

        self._log("Ready. Drop CSV files into the box above.", color=theme.LOG_INFO)
        self._log("New here? Tools → How to Use Loadscope… walks you through "
                  "labels, workflow, and tips.", color=theme.LOG_DIM)

        # Show a snapshot of what's already in the workbook so the user knows
        # state at a glance. Done after a short delay so the rest of the UI
        # paints first.
        QTimer.singleShot(300, self._log_workbook_state)

        # Kick off a non-blocking update check 1 second after launch.
        QTimer.singleShot(1000, lambda: self._start_update_check(manual=False))

    # ---------- logging helpers ----------

    def _log(self, msg, color=None):
        if color:
            self.log.append(f'<span style="color: {color};">{msg}</span>')
        else:
            self.log.append(msg)
        self.log.ensureCursorVisible()

    def _log_block(self, text, color="#d4d4d4"):
        """Append a multi-line block, preserving newlines."""
        for line in text.splitlines():
            self._log(line, color=color)

    # ---------- the meaty handler ----------

    def _stage_file(self, src, dest_dir, label, color):
        """Copy `src` into `dest_dir` if not already there. Logs the result.
        Returns True if the file is now (or was already) in dest_dir, False on copy failure."""
        name = Path(src).name
        dest = os.path.join(dest_dir, name)
        try:
            already_there = os.path.exists(dest) and os.path.samefile(src, dest)
        except OSError:
            already_there = False

        if already_there:
            self._log(f"  {label}  {name}  (already in {os.path.basename(dest_dir)}/)", color=color)
            return True

        try:
            shutil.copy2(src, dest)
            self._log(f"  {label}  {name}  →  {os.path.basename(dest_dir)}/", color=color)
            return True
        except shutil.SameFileError:
            self._log(f"  {label}  {name}  (already in {os.path.basename(dest_dir)}/)", color=color)
            return True
        except Exception as e:
            self._log(f"  COPY FAILED for {name}: {e}", color=theme.LOG_ERROR)
            return False

    def _refresh_status(self):
        total = self.staged_garmin + self.staged_bx
        if total == 0:
            self.status_label.setText("0 files staged")
            breakdown = ""
        else:
            parts = []
            if self.staged_garmin:
                parts.append(f"{self.staged_garmin} Garmin")
            if self.staged_bx:
                parts.append(f"{self.staged_bx} BallisticX")
            breakdown = " · ".join(parts) + " staged"
            self.status_label.setText(breakdown + " — drop more or click Run Import")
        self.go_button.setEnabled(total > 0)
        self.clear_button.setEnabled(total > 0)
        # v0.14: update the dropzone card to show the staged state inline
        # (accent title + "click Run Import below ↓" footer). Visible
        # post-drop CTA Chad asked for 2026-05-14.
        if hasattr(self, "drop"):
            self.drop.set_staged_state(total, breakdown)

    def clear_staged(self):
        # Doesn't delete files from folders — just resets the on-screen staging count.
        self.staged_garmin = 0
        self.staged_bx = 0
        self._log("\nStaged list cleared. (Files remain in import folders.)", color=theme.LOG_WARNING)
        self._refresh_status()

    def _build_demo_banner(self):
        """Yellow/orange banner shown at the top of the main window when the
        app is in demo mode. Has the Purchase a License CTA. Non-dismissable —
        the banner is the constant "you're previewing" cue.
        """
        banner = QWidget()
        banner.setObjectName("demoBanner")
        banner.setStyleSheet(
            f"#demoBanner {{ "
            f"  background-color: {theme.ACCENT}; "
            f"  border-radius: 6px; "
            f"}}"
        )
        row = QHBoxLayout(banner)
        row.setContentsMargins(14, 10, 14, 10)
        row.setSpacing(12)

        msg = QLabel(
            "DEMO MODE — You're previewing Loadscope with sample data. "
            "Purchase a license to import your own CSVs."
        )
        msg_font = QFont()
        msg_font.setPointSize(12)
        msg_font.setWeight(QFont.DemiBold)
        msg.setFont(msg_font)
        msg.setStyleSheet("color: white;")
        msg.setWordWrap(True)
        row.addWidget(msg, stretch=1)

        purchase_btn = QPushButton("Purchase a License")
        purchase_btn.setMinimumHeight(34)
        pf = QFont()
        pf.setPointSize(12)
        pf.setWeight(QFont.DemiBold)
        purchase_btn.setFont(pf)
        purchase_btn.setStyleSheet(
            "QPushButton { background-color: white; color: " + theme.ACCENT + "; "
            "border: none; border-radius: 4px; padding: 6px 14px; }"
            "QPushButton:hover { background-color: #f4f4f4; }"
        )
        purchase_btn.clicked.connect(self._open_purchase_page)
        row.addWidget(purchase_btn)

        return banner

    def _open_purchase_page(self):
        """Open the loadscope.app purchase page in the default browser."""
        import webbrowser
        webbrowser.open(app_license.PURCHASE_URL)
        self._log("Opened purchase page in browser.", color=theme.LOG_SUCCESS)

    def _show_demo_upgrade_prompt(self, reason="general"):
        """Modal that gates a real action behind a license purchase.

        `reason` lets us tailor the message slightly:
            'csv_drop'  — user tried to drop real CSVs in demo mode
            'general'   — generic "this feature needs a license"
        """
        if reason == "csv_drop":
            title = "Import requires a license"
            body = (
                "You're in demo mode. Importing your own Garmin Xero and "
                "BallisticX CSVs requires a Loadscope license.\n\n"
                "Want to continue with your own data? Purchase a license — "
                "your demo workbook stays read-only, and a fresh empty "
                "workbook takes its place."
            )
        else:
            title = "Feature requires a license"
            body = (
                "This feature requires a Loadscope license. You're currently "
                "in demo mode."
            )

        box = QMessageBox(self)
        box.setIcon(QMessageBox.Information)
        box.setWindowTitle(title)
        box.setText(body)
        purchase = box.addButton("Purchase a License", QMessageBox.AcceptRole)
        enter_key = box.addButton("Enter License Key…", QMessageBox.ActionRole)
        cancel = box.addButton("Not now", QMessageBox.RejectRole)
        box.setDefaultButton(purchase)
        box.exec_()
        clicked = box.clickedButton()
        if clicked is purchase:
            self._open_purchase_page()
        elif clicked is enter_key:
            from license_dialog import show_license_dialog
            if show_license_dialog(parent=self):
                # Successfully entered key — refresh banner state by warning
                # the user to restart for full effect.
                QMessageBox.information(
                    self,
                    "License accepted",
                    "Your license key is saved. Restart Loadscope to "
                    "exit demo mode and use your own data.",
                )

    def handle_drops(self, paths):
        """Called when the user drops files into the drop zone. Stages but does NOT import.

        For each dropped CSV, walks the parser registry to find which parser
        handles it. The matching parser's IMPORT_FOLDER determines where the
        file is copied. New parsers added under app/parsers/ work automatically
        without changes here.

        v0.14 demo-mode gating: in demo mode, real CSV drops trigger an upgrade
        modal instead of staging. The demo tour uses bundled demo CSVs through
        a different code path that bypasses this gate.
        """
        if app_license.is_demo_mode():
            self._show_demo_upgrade_prompt(reason="csv_drop")
            return

        self._log("")
        self._log(f"Staging {len(paths)} dropped item(s)…", color=theme.LOG_SUCCESS)

        for p in paths:
            name = Path(p).name
            if not p.lower().endswith(".csv"):
                self._log(f"  SKIP  {name}  (not a .csv)", color=theme.LOG_DIM)
                continue

            parser = detect_parser(p)
            if parser is None:
                self._log(
                    f"  UNKNOWN     {name}  (no matching parser — not staged)",
                    color=theme.LOG_ERROR,
                )
                continue

            dest_dir = os.path.join(self.project, parser.IMPORT_FOLDER)
            os.makedirs(dest_dir, exist_ok=True)

            # Pick a log color based on parser kind
            color = (
                theme.LOG_GARMIN if parser.KIND == "chronograph"
                else theme.LOG_BALLISTICX
            )
            label = f"{parser.NAME:<11}"  # left-padded for column alignment

            if self._stage_file(p, dest_dir, label, color):
                if parser.KIND == "chronograph":
                    self.staged_garmin += 1
                else:
                    self.staged_bx += 1

        self._refresh_status()
        if self.staged_garmin or self.staged_bx:
            self._log(
                "Click Run Import when you've dropped everything.",
                color=theme.LOG_WARNING,
            )

    def handle_loadscope_url(self, url):
        """Dispatcher for loadscope:// URLs (clickable workbook hyperlinks).

        URL shapes recognized:
          loadscope://reset-weights   — run the Reset Composite Weights flow
                                        on the active workbook.

        Bring the window forward so the user sees the confirmation dialog.
        Unknown actions are logged but do not raise.
        """
        self.show()
        self.raise_()
        self.activateWindow()

        action = parse_loadscope_action(url)
        if not action:
            self._log("Received loadscope:// URL with no action — ignoring.",
                      color=theme.LOG_DIM)
            return
        if action == "reset-weights":
            self._reset_composite_weights()
        elif action == "save-to-library":
            self._save_suggested_load()
        elif action == "print-pocket-card":
            self._print_pocket_card()
        elif action == "print-workbook":
            self._print_workbook()
        else:
            self._log(f"Unknown loadscope:// action: {action!r}",
                      color=theme.LOG_DIM)

    def handle_external_files(self, paths):
        """Files arrived from outside the window — drag-on-Dock-icon, "Open With…",
        or sys.argv at launch.

        Different UX from a window drop: the user already expressed clear intent
        by dragging onto the icon, so we auto-trigger the import after staging.
        Bring the window forward so the user sees what's happening.
        """
        self.show()
        self.raise_()
        self.activateWindow()

        self._log("")
        self._log(
            f"Files received from outside the window ({len(paths)}). "
            f"Staging and importing automatically…",
            color=theme.LOG_SUCCESS,
        )

        self.handle_drops(paths)

        # Auto-import after a short delay so the user sees the staging log first
        if self.staged_garmin or self.staged_bx:
            QTimer.singleShot(200, self.run_import_clicked)

    def _preflight_check(self):
        """Quick scan over the import folders before running the actual import.
        Surfaces likely issues as warnings in the log so the user can fix
        them before the workbook gets touched."""
        from parsers._common import parse_label
        warnings = []

        for parser in chronograph_parsers() + group_parsers():
            folder = os.path.join(self.project, parser.IMPORT_FOLDER)
            if not os.path.isdir(folder):
                continue
            for fn in os.listdir(folder):
                if not fn.lower().endswith(".csv"):
                    continue
                full = os.path.join(folder, fn)

                # Check the parser actually claims this file
                try:
                    if not parser.detect(full):
                        # File is in this parser's folder but doesn't match the
                        # format. Could be a misplaced CSV.
                        warnings.append(
                            f"  {parser.IMPORT_FOLDER}/{fn} doesn't look like "
                            f"a {parser.NAME} CSV — will be skipped."
                        )
                        continue
                except Exception as e:
                    warnings.append(
                        f"  {parser.IMPORT_FOLDER}/{fn}: couldn't read this CSV "
                        f"({type(e).__name__}: {e}). Will try to import anyway."
                    )

                # For BallisticX, the filename is the label source. Check that
                # the filename parses to something usable.
                if parser.KEY == "ballisticx":
                    base = os.path.splitext(fn)[0]
                    tag, charge, _ = parse_label(base)
                    if not tag or charge is None:
                        warnings.append(
                            f"  {fn} (BallisticX): filename doesn't look like a load label "
                            f"(needs format like 'P1 45.5 H4350.csv'). The data will be "
                            f"imported but may not land in the right row."
                        )

        return warnings

    def run_import_clicked(self):
        """User clicked the green Run Import button. Run the import on everything in the folders."""
        self._log("")
        self._log("=" * 60, color=theme.LOG_SUCCESS)
        self._log("Running import…", color=theme.LOG_SUCCESS)

        # Preflight check — surface issues before touching the workbook
        preflight_warnings = self._preflight_check()
        if preflight_warnings:
            self._log(f"Preflight check found {len(preflight_warnings)} issue(s):",
                      color=theme.LOG_WARNING)
            for w in preflight_warnings:
                self._log(w, color=theme.LOG_WARNING)
            self._log("(Continuing the import — review the warnings above.)",
                      color=theme.LOG_DIM)

        # Disable buttons during import to prevent double-clicks
        self.go_button.setEnabled(False)
        self.clear_button.setEnabled(False)
        QApplication.processEvents()

        try:
            self._do_import()
        finally:
            # After import attempt, reset session staging counter
            self.staged_garmin = 0
            self.staged_bx = 0
            self._refresh_status()

    def _do_import(self):

        # Step 2 — find the workbook (using the picker's current selection if
        # multiple workbooks exist, otherwise auto-pick most-recent)
        workbook_path = self._selected_workbook()
        if not workbook_path:
            # First-time user: no .xlsx yet. Prompt for a name and create one
            # from the template before continuing.
            workbook_path = self._prompt_first_load_name()
            if not workbook_path:
                self._log("\nImport cancelled — no workbook to write to.",
                          color=theme.LOG_DIM)
                return
            self._refresh_workbooks()
        else:
            # Active workbook already exists. If it has data, confirm with the user
            # so they don't accidentally mix new-cycle data into an old workbook.
            if self._workbook_has_data(workbook_path):
                choice = self._confirm_continue_or_new_cycle(workbook_path)
                if choice == "cancel":
                    self._log("\nImport cancelled.", color=theme.LOG_DIM)
                    return
                elif choice == "new_cycle":
                    new_wb = self._start_new_cycle_inline()
                    if not new_wb:
                        self._log("\nImport cancelled — new cycle wasn't created.",
                                  color=theme.LOG_DIM)
                        return
                    workbook_path = new_wb
                    self._refresh_workbooks()
                # else "continue": just keep going

        # If picker shows multiple, log which one we're using so the user is sure
        all_workbooks = import_data.list_workbooks(project_dir=self.project)
        if len(all_workbooks) > 1:
            self._log(
                f"\nUsing workbook: {os.path.basename(workbook_path)}  "
                f"(of {len(all_workbooks)} in project folder)",
                color=theme.LOG_WARNING,
            )

        # Step 3 — run the import, capturing stdout into our log
        self._log(f"\nRunning import → {os.path.basename(workbook_path)}", color=theme.LOG_SUCCESS)

        buf = io.StringIO()
        try:
            with redirect_stdout(buf):
                result = import_data.run_import(workbook_path, project_dir=self.project)
        except Exception:
            output = buf.getvalue()
            if output:
                self._log_block(output)
            self._log("\nIMPORT CRASHED:", color=theme.LOG_ERROR)
            self._log_block(traceback.format_exc(), color=theme.LOG_ERROR)
            return

        # Render captured output
        output = buf.getvalue().strip()
        if output:
            self._log_block(output)

        if not result["ok"]:
            self._log(f"\nIMPORT FAILED: {result['error']}", color=theme.LOG_ERROR)
            return

        # Success summary
        self._log("", )
        if result.get("safety_stop"):
            self._log("Safety stop — workbook left unchanged.", color=theme.LOG_WARNING)
        else:
            n_g = result["garmin_rows"]
            n_b = result["ballisticx_rows"]
            self._log(
                f"SUCCESS — wrote {n_g} Garmin rows and {n_b} BallisticX rows.",
                color=theme.LOG_SUCCESS,
            )
            # v0.14: SD-only auto-prompt (Chad 2026-05-14). When user
            # imported seating-depth data without a powder ladder AND no
            # winning charge has been set yet, ask for the charge so we
            # can populate Seating Depth!B10 + Charts!B3 automatically.
            if result.get("needs_sd_charge"):
                self._prompt_sd_only_charge(workbook_path)
            self._log("Workbook opened in Excel.", color=theme.LOG_SUCCESS)

            # macOS notification — useful when the user has switched to Excel
            # or another app and wants to know when the import is complete.
            self._show_macos_notification(
                title=f"{APP_NAME} — import complete",
                message=f"Imported {n_g} chronograph rows and {n_b} group rows. "
                        "Workbook open in Excel.",
            )

    # ---------- update check ----------

    def _start_update_check(self, manual=False):
        """Kick off a background update check. If `manual`, surface the result
        even when there's nothing new (so the user knows we checked)."""
        cfg = app_config.load_config()

        # Respect the user's "auto-check on launch" toggle. Manual checks
        # always run regardless of the setting.
        if not manual and not cfg.get("update_check_enabled", True):
            return

        # Use override from config if set, else fall back to the baked-in
        # default URL. This way updates work out of the box for everyone.
        url = cfg.get("update_manifest_url") or DEFAULT_MANIFEST_URL
        if not url:
            if manual:
                QMessageBox.information(
                    self,
                    "No updates available",
                    f"You're running {APP_NAME} v{APP_VERSION}.\n\n"
                    "Auto-update isn't enabled in this build yet. "
                    "We'll let you know by email when a new version is ready.",
                )
            return

        # Avoid double-launch
        if self._update_checker and self._update_checker.isRunning():
            return

        self._update_checker = UpdateChecker(url, parent=self)
        self._update_checker.finished_with_result.connect(
            lambda r: self._on_update_result(r, manual=manual)
        )
        self._update_checker.start()

    def _on_update_result(self, result, manual=False):
        if not result["ok"]:
            if manual:
                QMessageBox.warning(
                    self,
                    "Couldn't check for updates",
                    f"The update check failed:\n\n{result['error']}",
                )
            return

        manifest = result["manifest"] or {}
        self._last_manifest = manifest

        if not result["app_update"] and not result["template_update"]:
            if manual:
                QMessageBox.information(
                    self,
                    "You're up to date",
                    f"{APP_NAME} v{APP_VERSION} is the latest.",
                )
            return

        # Stash app update state for the install flow's use. Template updates
        # still go through the browser (template files aren't bundled into a
        # .app, so there's nothing to swap — user picks them up via the
        # workbook the next time they Save As from a fresh template).
        if result["app_update"]:
            self._pending_app_update_manifest = manifest
            self._pending_app_update_url = manifest.get("app_download_url")  # legacy fallback
            self._pending_app_update_version = manifest.get("app_version")
            self._pending_app_update_website = manifest.get("app_website_url") or "https://loadscope.app/"
        else:
            self._pending_app_update_manifest = None
            self._pending_app_update_url = None
            self._pending_app_update_version = None
            self._pending_app_update_website = None

        # Render the initial "ready to install" banner state.
        self._render_update_banner(state="ready", manifest=manifest, result=result)
        self.update_banner.setVisible(True)

    def _render_update_banner(self, state, manifest=None, result=None,
                              progress_pct=None, error=None):
        """Render the update banner in one of several states:
            'ready'      — update available, primary action: Install Update
            'downloading' — download in flight, show progress
            'installing' — download complete, primary action: Quit and Install
            'error'      — install or download failed, show fallback link

        Keeping this in one method so the markup stays consistent and we
        don't accidentally drift between paths."""
        if manifest is None:
            manifest = self._last_manifest or {}
        if result is None:
            # Synthesize a minimal result if we're being called from a
            # progress/install path that doesn't have one handy.
            result = {
                "app_update": bool(self._pending_app_update_url or self._pending_app_update_manifest),
                "template_update": False,
            }

        parts = []

        if result.get("app_update"):
            new_v = self._pending_app_update_version or manifest.get("app_version", "?")

            if state == "ready":
                parts.append(
                    f"<b>App update:</b> v{new_v} is available "
                    f"(you have v{APP_VERSION})."
                )
                # For new-style (gated) manifests we have either an endpoint
                # or a direct URL; either way we can attempt the in-app install.
                has_app_update = bool(
                    (self._pending_app_update_manifest and self._pending_app_update_manifest.get("app_download_endpoint"))
                    or self._pending_app_update_url
                )
                website = self._pending_app_update_website or "https://loadscope.app/"
                if installer.can_self_install() and has_app_update:
                    parts.append(
                        '<a href="install:start"><b>Install Update</b></a>'
                        ' &nbsp;·&nbsp; '
                        f'<a href="app:{website}" '
                        'style="color:#aaa;">Or download from the website</a>'
                    )
                else:
                    # Dev mode or read-only filesystem — no self-install path.
                    parts.append(
                        f'<a href="app:{website}">Download new version from the website</a>'
                    )

            elif state == "downloading":
                pct_text = f"{progress_pct}%" if progress_pct is not None else "…"
                parts.append(
                    f"<b>Downloading v{new_v}</b> &nbsp; {pct_text} "
                    '&nbsp;·&nbsp; '
                    '<a href="install:cancel" style="color:#aaa;">Cancel</a>'
                )

            elif state == "installing":
                parts.append(
                    f"<b>Update v{new_v} ready.</b> "
                    'Loadscope will close, install the update, and reopen.'
                )
                parts.append(
                    '<a href="install:swap"><b>Quit and Install</b></a>'
                )

            elif state == "error":
                parts.append(
                    f"<b>Update v{new_v}:</b> couldn't install automatically."
                )
                if error:
                    parts.append(f"<i>{error}</i>")
                website = self._pending_app_update_website or "https://loadscope.app/"
                parts.append(
                    f'<a href="app:{website}">Download manually from the website</a>'
                )

        if result.get("template_update") and state == "ready":
            new_t = manifest.get("template_version", "?")
            parts.append(f"<b>Template update:</b> v{new_t} is available.")
            url = manifest.get("template_download_url")
            if url:
                parts.append(f'<a href="template:{url}">Download new template</a>')

        # Release notes on the ready/error states only — we don't want the
        # downloading state to be a wall of text.
        if state in ("ready", "error"):
            notes_app = manifest.get("app_release_notes", "").strip()
            notes_t = manifest.get("template_release_notes", "").strip()
            notes = " · ".join(n for n in (notes_app, notes_t) if n)
            if notes:
                parts.append(f"<i>{notes}</i>")

        self.update_banner.setText("<br>".join(parts))

    def _on_update_link(self, link):
        """Handle banner link clicks.

        Link prefixes:
            install:start  — kick off the in-app download
            install:cancel — cancel an in-flight download
            install:swap   — quit and let the helper script swap in the new app
            app:URL        — open URL in the browser (manual download fallback)
            template:URL   — open URL in the browser (templates are browser-only)
        """
        if link == "install:start":
            self._begin_update_download()
            return

        if link == "install:cancel":
            self._cancel_update_download()
            return

        if link == "install:swap":
            self._perform_install_swap()
            return

        if link.startswith("app:"):
            url = link[len("app:"):]
        elif link.startswith("template:"):
            url = link[len("template:"):]
        else:
            url = link
        QDesktopServices.openUrl(QUrl(url))

    def _begin_update_download(self):
        """Start downloading the update zip in the background.

        For gated (v0.11.0+) manifests, this first POSTs the user's saved
        license key to the Cloudflare Worker and gets back a short-lived
        signed URL. For legacy manifests, it uses the direct app_download_url.
        """
        if self._update_downloader and self._update_downloader.isRunning():
            return

        manifest = self._pending_app_update_manifest
        if manifest is not None:
            # Resolves either via the gated Worker (new) or returns the direct
            # URL (legacy). Returns None if the license-gated lookup fails.
            url = resolve_download_url(manifest)
        else:
            url = self._pending_app_update_url

        if not url:
            self._render_update_banner(
                state="error",
                error=(
                    "Couldn't authorize the update download. Your license key may "
                    "have been revoked. Download the latest version manually from the website."
                ),
            )
            return

        self._render_update_banner(state="downloading", progress_pct=0)

        self._update_downloader = UpdateDownloader(url, parent=self)
        self._update_downloader.progress.connect(self._on_download_progress)
        self._update_downloader.finished_with_result.connect(self._on_download_finished)
        self._update_downloader.start()

    def _cancel_update_download(self):
        """Cancel an in-flight download and revert the banner."""
        if self._update_downloader and self._update_downloader.isRunning():
            self._update_downloader.cancel()
        # Revert to ready state so the user can try again
        self._render_update_banner(state="ready")

    def _on_download_progress(self, downloaded, total):
        """Update the banner with current progress percentage."""
        if total > 0:
            pct = min(100, int(downloaded * 100 / total))
        else:
            pct = None  # indeterminate
        self._render_update_banner(state="downloading", progress_pct=pct)

    def _on_download_finished(self, result):
        """Either swap directly into "Quit and Install" state, or surface
        the error and fall back to the manual link."""
        if not result["ok"]:
            err = result.get("error") or "Download failed"
            if err == "Cancelled":
                # User-initiated; banner already reverted by cancel handler
                return
            self._render_update_banner(state="error", error=err)
            return

        self._downloaded_zip_path = result["file_path"]
        self._render_update_banner(state="installing")

    def _perform_install_swap(self):
        """User clicked Quit and Install. Spawn the helper script and quit."""
        zip_path = self._downloaded_zip_path
        if not zip_path:
            self._render_update_banner(
                state="error",
                error="Internal error: no downloaded zip path."
            )
            return

        ok = installer.launch_install_swap(zip_path)
        if not ok:
            self._render_update_banner(
                state="error",
                error=(
                    "Loadscope couldn't swap in the new version automatically. "
                    "Use the download link to install manually."
                ),
            )
            return

        # Helper is running detached. Quit cleanly so the helper can do its
        # work without macOS holding the bundle open.
        QApplication.quit()

    def _on_tools_banner_link(self, link):
        """Dismiss the Tools-menu banner permanently."""
        if link == "dismiss:tools-banner":
            cfg = app_config.load_config()
            cfg["tools_banner_dismissed"] = True
            app_config.save_config(cfg)
            if self.tools_banner is not None:
                self.tools_banner.setVisible(False)

    def _show_macos_notification(self, title, message):
        """Show a macOS Notification Center notification using osascript.
        Silent failure if osascript isn't available or the notification is
        suppressed by macOS — never block the import flow on this."""
        try:
            # Escape double quotes for AppleScript safety
            esc_title = title.replace('"', '\\"')
            esc_message = message.replace('"', '\\"')
            script = f'display notification "{esc_message}" with title "{esc_title}"'
            subprocess.run(
                ["osascript", "-e", script],
                timeout=3,
                capture_output=True,
            )
        except Exception:
            pass  # never let a notification failure break anything

    def closeEvent(self, event):
        """Save window geometry on quit. Also confirm if files are staged but
        not yet imported — gentle nudge so the user doesn't accidentally lose
        their click-Run-Import intent by closing the window."""
        # If staged but not imported, ask before closing
        if self.staged_garmin or self.staged_bx:
            reply = QMessageBox.question(
                self,
                "Files staged",
                f"You have {self.staged_garmin + self.staged_bx} file(s) staged "
                "but haven't run the import yet.\n\n"
                "Quit anyway? Staged files stay in the import folders and "
                "will be picked up next time you click Run Import.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                event.ignore()
                return

        # Save window geometry so next launch opens at the same size+position
        try:
            geom = self.geometry()
            cfg = app_config.load_config()
            cfg["window_geometry"] = [geom.x(), geom.y(), geom.width(), geom.height()]
            app_config.save_config(cfg)
        except Exception:
            pass  # never block quit on a config save error

        super().closeEvent(event)

    def _show_disclaimer(self):
        """Show the full disclaimer in a read-only, scrollable dialog (for
        users who want to review it again after the first-launch acceptance).
        Uses DisclaimerViewer rather than QMessageBox so the long text scrolls
        properly on smaller screens."""
        view_disclaimer(parent=self)

    def _refresh_workbooks(self):
        """Re-scan the project folder for working .xlsx files and update the
        picker. Hides the picker entirely if only one workbook exists, so
        single-workbook users don't see UI clutter."""
        workbooks = import_data.list_workbooks(project_dir=self.project)

        # Remember the currently-selected path so we can restore it after rebuild
        prior_selection = self.wb_picker.currentData() if self.wb_picker.count() else None

        # Saved last-pick from config (preserves selection across launches)
        cfg = app_config.load_config()
        saved = cfg.get("last_selected_workbook")

        self.wb_picker.blockSignals(True)
        self.wb_picker.clear()
        for wb_path in workbooks:
            self.wb_picker.addItem(os.path.basename(wb_path), wb_path)
        self.wb_picker.blockSignals(False)

        # Restore: prior runtime selection > saved config selection > most-recent
        target = None
        if prior_selection in workbooks:
            target = prior_selection
        elif saved in workbooks:
            target = saved
        elif workbooks:
            target = workbooks[0]
        if target is not None:
            idx = self.wb_picker.findData(target)
            if idx >= 0:
                self.wb_picker.setCurrentIndex(idx)

        # Hide the whole row if only zero or one workbook exists — no need
        # for a picker when there's nothing to pick from.
        self.wb_row_widget.setVisible(len(workbooks) > 1)

        # Refresh the enable state of the "Open Workbook in Excel" menu item
        # so it greys out when no workbook exists yet.
        self._update_open_workbook_action_state()

    def _on_workbook_changed(self, _index):
        """User picked a different workbook in the combo box. Persist their
        choice to config so the next launch defaults to it."""
        path = self.wb_picker.currentData()
        if not path:
            return
        cfg = app_config.load_config()
        cfg["last_selected_workbook"] = path
        app_config.save_config(cfg)

    def _selected_workbook(self):
        """Return the path of the workbook currently selected in the picker,
        or None if no workbooks exist in the project folder."""
        path = self.wb_picker.currentData()
        if path and os.path.isfile(path):
            return path
        # Fallback: rescan and use most-recent
        workbooks = import_data.list_workbooks(project_dir=self.project)
        return workbooks[0] if workbooks else None

    def _prompt_sd_only_charge(self, workbook_path):
        """v0.14: when user imported seating-depth data without a powder
        ladder, ask them what charge weight they used. Write it to both
        Seating Depth!B10 (the visible 'Charge:' cell) and Charts!B3
        (the source-of-truth winner cell that all other formulas reference).
        Chad 2026-05-14: "I also like the ability for the auto prompt."
        """
        from PyQt5.QtWidgets import QInputDialog
        charge, ok = QInputDialog.getDouble(
            self,
            "What charge weight did you use?",
            "Loadscope didn't see a powder ladder in this import — only "
            "seating depth data. To finish setting up the workbook, type the "
            "powder charge (in grains) you used for your seating depth tests:",
            value=42.0,        # sensible default for 6.5 CM-class loads
            min=1.0, max=200.0, decimals=2,
        )
        if not ok:
            self._log(
                "Skipped seating-depth charge prompt — you can type it "
                "manually into Seating Depth row 10.",
                color=theme.LOG_DIM,
            )
            # Even if user cancelled, open Excel so they see their data
            subprocess.run(["open", workbook_path], check=False)
            self._minimize_chrome_after_excel_loads()
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook_path, data_only=False)
            if "Seating Depth" in wb.sheetnames:
                wb["Seating Depth"]["B10"].value = charge
            if "Charts" in wb.sheetnames:
                wb["Charts"]["B3"].value = charge
            wb.save(workbook_path)
            self._log(
                f"Set seating-depth charge to {charge} gr.",
                color=theme.LOG_SUCCESS,
            )
        except Exception as e:
            self._log(
                f"Couldn't save the charge: {e}. "
                "Open the workbook and type it into Seating Depth row 10.",
                color=theme.LOG_ERROR,
            )
        # Open Excel after we've written (or attempted to write) the charge.
        subprocess.run(["open", workbook_path], check=False)
        self._minimize_chrome_after_excel_loads()

    def _open_import_folder_for_chip(self, chip_label):
        """DropZone chip click handler — opens the import folder for the
        named device in Finder. Looks up the IMPORT_FOLDER from the live
        parser registry so adding a new parser (LabRadar, etc.) means the
        chip just works without touching this method.
        """
        # Build chip-name → IMPORT_FOLDER map dynamically from registry
        rel = None
        for p in chronograph_parsers() + group_parsers():
            if p.NAME == chip_label:
                rel = p.IMPORT_FOLDER
                break
        if not rel:
            return
        path = os.path.join(self.project, rel)
        if not os.path.isdir(path):
            try:
                os.makedirs(path, exist_ok=True)
            except OSError as e:
                QMessageBox.warning(
                    self, "Couldn't open folder",
                    f"Couldn't create or open {rel}: {e}",
                )
                return
        subprocess.run(["open", path], check=False)

    def _open_workbook_in_excel(self):
        """Open the currently-active workbook in Excel (or whatever app is
        registered to open .xlsx files). Useful when Excel ends up behind
        the Loadscope window and the user can't find their workbook."""
        wb_path = self._selected_workbook()
        if not wb_path or not os.path.isfile(wb_path):
            self._log("No workbook to open yet. Drop CSVs and click Run Import to create one.",
                      color=theme.LOG_DIM)
            return
        try:
            subprocess.run(["open", wb_path], check=False)
            self._minimize_chrome_after_excel_loads()
        except Exception as e:
            self._log(f"Couldn't open workbook: {e}", color=theme.LOG_ERROR)

    def _minimize_chrome_after_excel_loads(self):
        """After triggering Excel to open a workbook, give it ~1.5s to launch
        and render the workbook, then hide the Excel chrome (ribbon, formula
        bar, status bar, row/column headings). Best-effort and deferred via
        QTimer so it doesn't block the UI thread.
        See app/excel_chrome.py for the full keep/hide matrix.
        """
        from PyQt5.QtCore import QTimer
        from excel_chrome import minimize_excel_chrome
        QTimer.singleShot(1500, minimize_excel_chrome)

    def _update_open_workbook_action_state(self):
        """Enable the 'Open Workbook in Excel' menu item only when there's
        actually a workbook to open. Called on startup and whenever the
        workbook picker selection changes."""
        wb = self._selected_workbook()
        enabled = bool(wb and os.path.isfile(wb))
        for attr in ("_open_wb_action", "_print_wb_action"):
            action = getattr(self, attr, None)
            if action is not None:
                action.setEnabled(enabled)

    def _reset_composite_weights(self):
        """Restore the workbook's composite-score weights to Loadscope's
        defaults on both the Charts and Seating Depth sheets."""
        wb_path = self._selected_workbook()
        if not wb_path or not os.path.isfile(wb_path):
            QMessageBox.information(
                self,
                "No workbook",
                "There's no workbook to reset yet. Drop CSVs and click Run "
                "Import to create one first.",
            )
            return
        reply = QMessageBox.question(
            self,
            "Reset Composite Weights",
            (
                f"Reset the composite-score weights on '{os.path.basename(wb_path)}' "
                f"to Loadscope defaults?\n\n"
                f"Charts (powder ladder):   Vel 0.30  •  SD 0.20  •  MR 0.20  •  SD-Vert 0.30\n"
                f"Seating Depth:            Vel 0.15  •  SD 0.25  •  MR 0.25  •  SD-Vert 0.35\n\n"
                f"Any custom weights you've typed will be overwritten."
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            changes = import_data.reset_composite_weights(wb_path)
        except PermissionError:
            QMessageBox.warning(
                self,
                "Workbook is open",
                "Close the workbook in Excel and try again. Loadscope can't "
                "write to a workbook that Excel has open.",
            )
            return
        except Exception as e:
            QMessageBox.critical(
                self,
                "Reset failed",
                f"Couldn't reset weights:\n\n{e}",
            )
            return
        self._log(
            f"Reset composite weights on {os.path.basename(wb_path)} "
            f"({len(changes)} cell(s) updated).",
            color=theme.LOG_SUCCESS,
        )

    def _save_suggested_load(self):
        """Append a new row to Load Library capturing the current
        suggested load (winning charge from Charts, winning jump from
        Seating Depth, components from Load Log header, performance
        metrics from the matching winner rows)."""
        wb_path = self._selected_workbook()
        if not wb_path or not os.path.isfile(wb_path):
            QMessageBox.information(
                self,
                "No workbook",
                "There's no workbook to save from yet. Drop CSVs and click "
                "Run Import to create one first.",
            )
            return
        # Gather the suggested load data so the confirmation dialog has
        # something concrete to show. Surfaces "no winner yet" errors
        # early before bothering the user with a save dialog.
        try:
            data = import_data.gather_suggested_load(wb_path)
        except ValueError as e:
            QMessageBox.warning(self, "Nothing to save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Couldn't read workbook", str(e))
            return

        # Build a human-readable summary for the confirmation dialog.
        def fmt(val, suffix=""):
            if val in (None, ""):
                return "—"
            if isinstance(val, float):
                return f"{val:g}{suffix}"
            return f"{val}{suffix}"

        summary_lines = [
            f"Date Added:  {fmt(data.get('date_added'))}",
            f"Load Name:   {fmt(data.get('load_name'))}",
            f"Rifle:       {fmt(data.get('rifle'))}",
            f"Bullet:      {fmt(data.get('bullet'))}  ({fmt(data.get('bullet_wt'), ' gr')})",
            f"Powder:      {fmt(data.get('powder'))}  ({fmt(data.get('charge'), ' gr')})",
            f"Primer:      {fmt(data.get('primer'))}",
            f"Brass:       {fmt(data.get('brass'))}",
            f"CBTO:        {fmt(data.get('cbto'), ' in')}",
            f"Jump:        {fmt(data.get('jump'), ' in')}",
            f"Avg Vel:     {fmt(data.get('avg_vel'), ' fps')}",
            f"SD:          {fmt(data.get('sd_fps'), ' fps')}",
            f"Group:       {fmt(data.get('group_moa'), ' MOA')}",
            f"Mean Radius: {fmt(data.get('mr_moa'), ' MOA')}",
        ]
        if data.get("notes"):
            summary_lines.append(f"Notes:       {data.get('notes')[:80]}")
        prompt = (
            f"Save this load to Load Library on "
            f"'{os.path.basename(wb_path)}'?\n\n"
            + "\n".join(summary_lines)
            + "\n\nYou can edit any of these cells in Load Library after "
            "the row is added."
        )

        reply = QMessageBox.question(
            self,
            "Save Suggested Load to Library",
            prompt,
            QMessageBox.Save | QMessageBox.Cancel,
            QMessageBox.Save,
        )
        if reply != QMessageBox.Save:
            return
        try:
            row, written = import_data.save_suggested_load_to_library(wb_path, data=data)
        except PermissionError:
            QMessageBox.warning(
                self,
                "Workbook is open",
                "Close the workbook in Excel and try again. Loadscope can't "
                "write to a workbook that Excel has open.",
            )
            return
        except ValueError as e:
            QMessageBox.warning(self, "Couldn't save", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Save failed", f"Couldn't save load:\n\n{e}")
            return
        self._log(
            f"Saved suggested load to Load Library!A{row} on "
            f"{os.path.basename(wb_path)}.",
            color=theme.LOG_SUCCESS,
        )

    def _print_pocket_card(self):
        """Generate + open a polished printable Pocket Range Card.

        Renders a 4×6 landscape HTML card from the Ballistics tab data and
        opens it in the user's default browser — user prints to cardstock
        or saves as PDF. The 'Range Card' worksheet (v0.14+) shows the
        in-workbook preview, but this menu action produces the polished
        artifact with proper typography."""
        wb_path = self._selected_workbook()
        # v0.14: in demo mode, the user is reviewing the bundled demo
        # workbook (which lives outside the project folder, so the picker
        # doesn't see it). Force the bundled demo workbook so the demo's
        # Print Pocket Card button always works, regardless of what
        # _selected_workbook() returns.
        if app_license.is_demo_mode():
            try:
                from demo_tour import get_bundled_demo_workbook_path
                bundled = get_bundled_demo_workbook_path()
                if bundled and os.path.isfile(bundled):
                    wb_path = bundled
            except ImportError:
                pass
        if not wb_path or not os.path.isfile(wb_path):
            QMessageBox.information(
                self,
                "No workbook",
                "There's no workbook to generate a card from yet.",
            )
            return
        try:
            from pocket_card import generate_pocket_card
            out_path = generate_pocket_card(wb_path, open_after=True)
        except ValueError as e:
            QMessageBox.warning(self, "Couldn't generate card", str(e))
            return
        except PermissionError:
            QMessageBox.warning(
                self,
                "Workbook is open",
                "Close the workbook in Excel and try again. Loadscope needs "
                "to read the latest cached values to build the card.",
            )
            return
        except Exception as e:
            QMessageBox.critical(self, "Generation failed",
                                 f"Couldn't generate Pocket Range Card:\n\n{e}")
            return
        self._log(
            f"Pocket Range Card generated: {os.path.basename(out_path)}",
            color=theme.LOG_SUCCESS,
        )

    def _open_demo_tour(self):
        """Open the guided demo tour panel beside Excel.

        Walks the user through Load Log, Charts, Seating Depth, Ballistics,
        Pocket Range Card, and Load Library with timed narration. Used both
        from the Workbook menu (any user can replay) and from the first-launch
        trial flow (auto-fires once).
        """
        # v0.14: ALWAYS prefer the bundled demo workbook in demo mode.
        # Chad caught this 2026-05-14: in demo mode the tour was opening
        # legacy workbooks (e.g., 6.xlsx) from his project folder via
        # _selected_workbook(), which triggered Excel "found a problem with
        # content" recovery dialogs. The demo tour is supposed to walk the
        # CURATED demo workbook, never random project files.
        wb_path = None
        try:
            from demo_tour import get_bundled_demo_workbook_path
            bundled = get_bundled_demo_workbook_path()
            if app_license.is_demo_mode():
                # Demo users → always the bundled workbook
                wb_path = bundled if (bundled and os.path.isfile(bundled)) else None
            else:
                # Licensed users replaying the tour → use their selected
                # workbook if they have one, fall back to bundled otherwise
                wb_path = self._selected_workbook()
                if (not wb_path or not os.path.isfile(wb_path)) and bundled:
                    wb_path = bundled
        except ImportError:
            wb_path = self._selected_workbook()
        if not wb_path or not os.path.isfile(wb_path):
            QMessageBox.information(
                self,
                "No workbook",
                "Pick a workbook first so the tour has something to walk you through.",
            )
            return
        try:
            from demo_tour import DemoTourPanel
        except ImportError as e:
            QMessageBox.critical(
                self,
                "Couldn't load demo tour",
                f"demo_tour module is missing or broken:\n\n{e}",
            )
            return

        def _on_purchase():
            # Placeholder until v0.14 commerce ships — opens the marketing site.
            import webbrowser
            webbrowser.open("https://loadscope.app/")
            self._log("Opened purchase page in browser.", color=theme.LOG_SUCCESS)

        self._demo_tour_panel = DemoTourPanel(
            wb_path,
            on_purchase=_on_purchase,
            parent=None,
        )
        # v0.14.2: hide the main Loadscope window during the tour so it
        # doesn't compete with the tour panel for screen real estate at
        # the top of the screen. Tour panel + Excel are the two visible
        # windows during the tour. Main window comes back when the tour
        # panel is destroyed.
        self.hide()
        # Defensive re-show: when the user quits Loadscope while the tour
        # is open, the tour panel's destroyed signal fires AFTER the
        # MainWindow's C++ object is already gone. Catch the
        # "wrapped C/C++ object has been deleted" RuntimeError so the
        # crash reporter doesn't fire a misleading bug report.
        def _safe_reshow_main(*_args):
            try:
                self.show()
            except RuntimeError:
                pass  # MainWindow was destroyed during app shutdown
        self._demo_tour_panel.destroyed.connect(_safe_reshow_main)
        self._demo_tour_panel.show()
        self._demo_tour_panel.start()
        self._log("Demo tour started.", color=theme.LOG_SUCCESS)

    def _print_workbook(self):
        """Open the active workbook in Excel and trigger the Print dialog.
        Uses AppleScript to send Excel a print command after opening the file.
        Each user-facing sheet is preconfigured to fit on one landscape page,
        so the user just confirms and prints."""
        wb_path = self._selected_workbook()
        if not wb_path or not os.path.isfile(wb_path):
            self._log("No workbook to print yet. Drop CSVs and click Run Import to create one.",
                      color=theme.LOG_DIM)
            return
        # AppleScript: open in Excel + show print dialog
        osa = (
            'tell application "Microsoft Excel"\n'
            '    activate\n'
            f'    open POSIX file "{wb_path}"\n'
            '    delay 1\n'
            '    tell active workbook to print out without print dialog -- show=true makes it interactive\n'
            'end tell'
        )
        # NOTE: "print out without print dialog" actually sends straight to default
        # printer. To SHOW the dialog (so the user can pick printer, copies, page
        # range), use "print out" without that clause. We use the latter.
        osa = (
            'tell application "Microsoft Excel"\n'
            '    activate\n'
            f'    open POSIX file "{wb_path}"\n'
            '    delay 1\n'
            '    print active workbook\n'
            'end tell'
        )
        try:
            subprocess.run(["osascript", "-e", osa], check=False)
            self._minimize_chrome_after_excel_loads()
        except Exception as e:
            self._log(f"Couldn't trigger print: {e}", color=theme.LOG_ERROR)

    def _workbook_has_data(self, workbook_path):
        """Return True if Load Log row 16+ has any imported data (any charge value
        in column B). Used to decide whether to prompt for new-cycle confirmation."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook_path, data_only=True, keep_vba=False)
            if "Load Log" not in wb.sheetnames:
                return False
            ws = wb["Load Log"]
            for r in range(16, 26):
                v = ws.cell(r, 2).value
                if v not in (None, ""):
                    return True
            return False
        except Exception:
            return False  # err on the safe side — don't prompt if we can't read

    def _confirm_continue_or_new_cycle(self, workbook_path):
        """Ask the user whether to continue with the current workbook or start a
        new cycle. Returns 'continue', 'new_cycle', or 'cancel'."""
        from PyQt5.QtWidgets import QMessageBox
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("Continue this load, or start a new one?")
        box.setText(
            f"You're about to import into:\n\n"
            f"     {os.path.basename(workbook_path)}\n\n"
            f"Is this more data for the same load, or are you starting a new one "
            f"(different powder, bullet, or cartridge)?"
        )
        continue_btn = box.addButton("Continue this load", QMessageBox.AcceptRole)
        new_btn = box.addButton("Start a new load…", QMessageBox.ActionRole)
        cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(continue_btn)
        box.exec_()
        clicked = box.clickedButton()
        if clicked is continue_btn: return "continue"
        if clicked is new_btn:      return "new_cycle"
        return "cancel"

    def _start_new_cycle_inline(self):
        """Launch the New Cycle dialog and return the path of the new workbook
        (or None if the user cancelled)."""
        from new_cycle_dialog import show_new_cycle
        return show_new_cycle(self.project, self._selected_workbook(), parent=self)

    def _prompt_first_load_name(self):
        """Brand-new user has no .xlsx in their project folder yet. Prompt for
        a name for their first load and create the workbook from the template.
        Returns the new workbook's path, or None if the user cancelled."""
        from PyQt5.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(
            self,
            "Name your first load",
            "Looks like this is your first load. What would you like to call it?\n\n"
            "Pick something descriptive — for example:\n"
            "  6.5 Creedmoor 140 ELD-M H4350\n"
            "  300 PRC 225 ELD-M H1000",
            text="",
        )
        if not ok:
            return None
        name = name.strip()
        if not name:
            return None
        # Sanitize for filename safety
        safe = name.replace("/", "-").replace(":", "-")
        new_path = os.path.join(self.project, f"{safe}.xlsx")
        if os.path.exists(new_path):
            self._log(f"\nA workbook named '{safe}.xlsx' already exists. Pick a different name.",
                      color=theme.LOG_ERROR)
            return None
        # Find the bundled template
        try:
            from setup_wizard import find_bundled_template
            template = find_bundled_template()
        except Exception:
            template = None
        # Also check the project folder itself
        if not template:
            for f in os.listdir(self.project):
                if f.lower().endswith(".xltx") and "template" in f.lower():
                    template = os.path.join(self.project, f)
                    break
        if not template or not os.path.exists(str(template)):
            self._log("\nCouldn't find the workbook template — please contact support@loadscope.app.",
                      color=theme.LOG_ERROR)
            return None
        # Copy template → .xlsx (re-save through openpyxl so it's a proper workbook,
        # not a template — same trick the New Cycle dialog uses). Also stamp
        # the load name onto each user-facing sheet so the user knows which
        # load they're viewing.
        try:
            from openpyxl import load_workbook
            import import_data
            wb = load_workbook(str(template), keep_vba=False)
            wb.template = False
            import_data.stamp_load_name(wb, name)
            inherited = import_data.inherit_rifle_setup(wb, self.project, exclude_path=new_path)
            wb.save(new_path)
        except Exception as e:
            self._log(f"\nCouldn't create the new workbook: {e}", color=theme.LOG_ERROR)
            return None
        self._log(f"\nCreated workbook: {safe}.xlsx", color=theme.LOG_SUCCESS)
        if inherited:
            self._log(f"  Pre-filled from previous workbook: {', '.join(inherited)}",
                      color=theme.LOG_DIM)
        return new_path

    def _log_workbook_state(self):
        """Show a one-line summary of the current workbook contents in the
        activity log. Called on startup so the user sees what's already loaded."""
        wb_path = self._selected_workbook()
        if not wb_path:
            self._log("\nNo working .xlsx in the project folder yet. Save a "
                      "copy from the .xltx template to start.", color=theme.LOG_WARNING)
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(wb_path, data_only=True, keep_vba=False)
        except Exception:
            return  # silent fail — not worth alarming the user

        def _count_distinct_tags(sheet_name, pattern):
            """Count distinct tags in column A whose uppercase form matches the
            given regex pattern (anchored)."""
            if sheet_name not in wb.sheetnames:
                return 0
            sht = wb[sheet_name]
            tags = set()
            for row in sht.iter_rows(min_row=2, max_col=1, values_only=True):
                tag = (row[0] or "")
                if not isinstance(tag, str):
                    continue
                tag_u = tag.upper()
                if pattern.match(tag_u):
                    tags.add(tag_u)
            return len(tags)

        # Powder ladder: P followed by a digit (P1, P2, ...). S-tags need a digit
        # too so they don't match user-typed labels like "SHOT-3" or "SUNDAY".
        n_powder = _count_distinct_tags("GarminSessions", re.compile(r"^P\d"))
        n_seating = _count_distinct_tags("GarminSessions", re.compile(r"^S\d"))
        n_confirm = _count_distinct_tags("GarminSessions", re.compile(r"^CONFIRM"))

        # Read suggested charge / jump from Charts
        try:
            suggested_charge = wb["Charts"]["B3"].value
        except Exception:
            suggested_charge = None
        try:
            suggested_jump = wb["Seating Depth"]["D2"].value
        except Exception:
            suggested_jump = None

        # Build the summary line
        parts = []
        if n_powder:
            parts.append(f"{n_powder} powder ladder load{'s' if n_powder != 1 else ''}")
        if n_seating:
            parts.append(f"{n_seating} seating depth test{'s' if n_seating != 1 else ''}")
        if n_confirm:
            parts.append(f"{n_confirm} confirmation group{'s' if n_confirm != 1 else ''}")

        # Only show the workbook filename in the log when there are multiple
        # workbooks in the project folder — otherwise it's redundant noise
        # (the user already knows which workbook they're working with).
        all_workbooks = import_data.list_workbooks(project_dir=self.project)
        if len(all_workbooks) > 1:
            wb_name_prefix = f"{os.path.basename(wb_path)}: "
        else:
            wb_name_prefix = ""

        if not parts:
            self._log(f"\n{wb_name_prefix}Workbook is empty so far. Drop CSVs to import.",
                      color=theme.LOG_INFO)
            return

        self._log(f"\n{wb_name_prefix}Currently has " + " · ".join(parts) + ".",
                  color=theme.LOG_INFO)

        # Report suggested winners if they exist. Use `is not None` rather than
        # truthiness so a legitimate result of 0.0 (e.g., a seating jump of 0.0
        # = touching the lands) doesn't get silently dropped.
        winner_parts = []
        if isinstance(suggested_charge, (int, float)) and suggested_charge is not None:
            winner_parts.append(f"suggested charge {suggested_charge:g} gr")
        if isinstance(suggested_jump, (int, float)) and suggested_jump is not None:
            winner_parts.append(f"suggested jump {suggested_jump:g}\"")
        if winner_parts:
            self._log("  Current winner: " + ", ".join(winner_parts) + ".",
                      color=theme.LOG_SUCCESS)
        elif n_powder < 3 or n_seating < 3:
            self._log("  (Need at least 3 loads in either ladder before "
                      "Loadscope picks a suggested winner.)",
                      color=theme.LOG_DIM)

    def _show_help(self):
        """Tools → How to Use Loadscope…"""
        show_help(parent=self)

    def _show_faq(self):
        """Support → Frequently Asked Questions… — opens the categorized
        FAQ dialog with search."""
        from faq_dialog import show_faq_dialog
        show_faq_dialog(parent=self)

    def _show_settings(self):
        """Tools → Settings…"""
        show_settings(self.project, parent=self)

    def _run_import_now_menu(self):
        """Tools → Run Import Now — runs import on whatever's already in the
        Garmin/BallisticX folders, no dropping required. Useful for re-running
        after fixing a CSV name or adding files outside the GUI."""
        # Same flow as run_import_clicked but doesn't require staged files
        self.run_import_clicked()

    def _start_new_cycle(self):
        """Tools → Start New Cycle… — wraps up the current cycle and sets up
        a fresh workbook in one step."""
        current = self._selected_workbook()
        new_path = show_new_cycle(self.project, current, parent=self)
        if new_path:
            # Refresh the picker so the new workbook becomes selectable, and
            # auto-select it
            self._refresh_workbooks()
            idx = self.wb_picker.findData(new_path)
            if idx >= 0:
                self.wb_picker.setCurrentIndex(idx)
            self._log(f"\nNew cycle started: {os.path.basename(new_path)}",
                      color=theme.LOG_SUCCESS)
            self._log("Drop CSVs from your next range trip to begin.",
                      color=theme.LOG_INFO)

    def _restore_from_backup(self):
        """Tools → Restore From Backup… — opens the .backups/ folder, lets
        the user pick a backup .xlsx, copies it over the current workbook
        (after backing up the CURRENT current workbook first, so this
        operation itself is undoable)."""
        backups_dir = os.path.join(self.project, ".backups")
        if not os.path.isdir(backups_dir) or not os.listdir(backups_dir):
            QMessageBox.information(
                self,
                "No backups yet",
                "Loadscope hasn't created any backups yet. Backups are saved "
                "to the .backups/ folder before each import.",
            )
            return

        from PyQt5.QtWidgets import QFileDialog
        backup_path, _ = QFileDialog.getOpenFileName(
            self,
            "Pick a backup to restore",
            backups_dir,
            "Excel workbooks (*.xlsx);;All files (*)",
        )
        if not backup_path:
            return

        target_path = self._selected_workbook()
        if not target_path:
            QMessageBox.warning(
                self, "No workbook",
                "There's no active workbook to restore over. Save a copy from "
                "the .xltx template first.",
            )
            return

        backup_name = os.path.basename(backup_path)
        target_name = os.path.basename(target_path)
        reply = QMessageBox.question(
            self,
            "Confirm restore",
            f"Replace '{target_name}' with '{backup_name}'?\n\n"
            "Your current workbook will itself be backed up first, so you "
            "can undo this if needed.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        # Back up the CURRENT workbook before overwriting (so this is undoable).
        # Honors the user's backup_keep_count setting; if backups are turned off,
        # we still take an unconditional safety copy alongside the workbook so
        # the restore is undoable even with retention disabled.
        backup_keep = import_data._read_backup_keep_setting(default=5)
        try:
            if backup_keep > 0:
                import_data._rotate_workbook_backups(target_path, keep=backup_keep)
            else:
                # Backups disabled — drop a one-shot safety copy next to the workbook
                from datetime import datetime
                stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                safety_path = f"{target_path}.before-restore-{stamp}.xlsx"
                shutil.copy2(target_path, safety_path)
        except Exception as e:
            QMessageBox.warning(
                self, "Couldn't create safety backup",
                f"Stopping the restore to be safe.\n\nError: {e}",
            )
            return

        # Replace the target with the chosen backup
        try:
            shutil.copy2(backup_path, target_path)
        except Exception as e:
            QMessageBox.warning(
                self, "Restore failed",
                f"Couldn't copy backup over the current workbook.\n\nError: {e}",
            )
            return

        self._log(f"\nRestored '{target_name}' from backup '{backup_name}'.",
                  color=theme.LOG_SUCCESS)
        self._log("Your previous version was backed up first. "
                  "Use Restore From Backup again to revert if needed.",
                  color=theme.LOG_INFO)

        # Open the restored workbook so the user can verify
        subprocess.run(["open", target_path])

    def _generate_load_card(self):
        """Tools → Generate Load Card…
        Picks the most-recent working .xlsx in the project folder, builds an
        HTML load card, and opens it in the user's browser to print or save as PDF."""
        try:
            from load_card import generate_load_card
        except ImportError as e:
            QMessageBox.warning(
                self,
                "Couldn't generate load card",
                f"Missing module: {e}",
            )
            return

        workbook_path = self._selected_workbook()
        if not workbook_path:
            QMessageBox.warning(
                self,
                "No workbook found",
                "There's no working .xlsx in the project folder. "
                "Save a copy of the template first, then come back.",
            )
            return

        try:
            out_path = generate_load_card(workbook_path, open_after=True)
            self._log(f"\nLoad card generated: {os.path.basename(out_path)}",
                      color=theme.LOG_SUCCESS)
            self._log("Opened in your browser. Use File → Print → Save as PDF "
                      "to save a copy.", color=theme.LOG_INFO)
        except Exception as e:
            QMessageBox.warning(
                self,
                "Couldn't generate load card",
                f"An error occurred:\n\n{e}",
            )

    def _export_load(self):
        """Tools → Export Suggested Load… — write the current workbook's
        suggested load to a .loadscope file in the Shared Loads/ subfolder."""
        try:
            from load_sharing import export_load
        except ImportError as e:
            QMessageBox.warning(self, "Couldn't export", f"Missing module: {e}")
            return

        workbook_path = self._selected_workbook()
        if not workbook_path:
            QMessageBox.warning(self, "No workbook", "No working .xlsx in the project folder.")
            return

        try:
            out_path = export_load(workbook_path)
        except Exception as e:
            QMessageBox.warning(self, "Export failed", f"An error occurred:\n\n{e}")
            return

        self._log(f"\nExported load to: {os.path.basename(out_path)}",
                  color=theme.LOG_SUCCESS)
        self._log("Find it in: Shared Loads/  (Tools → Show Project Folder)",
                  color=theme.LOG_INFO)

        # Reveal it in Finder so the user can email/AirDrop right away
        subprocess.run(["open", "-R", out_path])

    def _import_load(self):
        """Tools → Import Shared Load… — open a .loadscope file in a file picker
        and show its contents in a read-only dialog. Doesn't write to the user's
        workbook (intentional — user must consciously decide to try the load)."""
        try:
            from load_sharing import import_load, format_load_for_display, LoadFileError
        except ImportError as e:
            QMessageBox.warning(self, "Couldn't import", f"Missing module: {e}")
            return

        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Shared Load",
            os.path.join(self.project, "Shared Loads"),
            "Loadscope loads (*.loadscope);;All files (*)",
        )
        if not path:
            return

        try:
            data = import_load(path)
        except LoadFileError as e:
            QMessageBox.warning(self, "Couldn't read load", str(e))
            return

        text = format_load_for_display(data)
        # Plain-text dialog — no fancy formatting needed
        msg = QMessageBox(self)
        msg.setWindowTitle(f"Shared Load — {os.path.basename(path)}")
        msg.setIcon(QMessageBox.Information)
        msg.setText("A load shared by another Loadscope user:")
        msg.setInformativeText(text)
        # Force monospace by setting the detailed-text font
        font = msg.font()
        font.setFamily("SF Mono")
        msg.setFont(font)
        msg.exec_()

    def _reveal_in_finder(self, path):
        """Open the given folder in macOS Finder. Creates the folder first if
        it doesn't exist (e.g., .backups before the first import)."""
        if not os.path.isdir(path):
            try:
                os.makedirs(path, exist_ok=True)
            except OSError as e:
                QMessageBox.warning(
                    self,
                    "Couldn't open folder",
                    f"This folder doesn't exist and couldn't be created:\n\n{path}\n\n{e}",
                )
                return
        subprocess.run(["open", path])

    def _visit_website(self):
        """Tools → Visit Loadscope Website — opens the marketing site in the
        user's default browser. Useful for sharing the app, checking for
        documentation, or pointing friends at the access-code request form."""
        from PyQt5.QtCore import QUrl
        from PyQt5.QtGui import QDesktopServices
        QDesktopServices.openUrl(QUrl("https://loadscope.app/"))

    def _send_feedback(self):
        """Support → Contact Support — opens the user's default email client
        with a pre-filled message to support@loadscope.app. Covers bugs,
        feature requests, device support requests, questions — anything.
        Renamed from "Send Feedback" 2026-05-14 (Chad: broader scope)."""
        from PyQt5.QtCore import QUrl
        from PyQt5.QtGui import QDesktopServices
        subject = f"{APP_NAME} v{APP_VERSION}"
        body = (
            "Hi Loadscope team,%0A%0A"
            "(Tell us what you need — bug reports, feature requests, "
            "questions about how something works, requests to support a "
            "new chronograph or target app, anything is welcome.)%0A%0A"
            "Thanks!"
        )
        display_name = f"{APP_NAME} Support"
        recipient = "support@loadscope.app"
        encoded_recipient = (
            f"%22{display_name.replace(' ', '%20')}%22%20%3C{recipient}%3E"
        )
        mailto = f"mailto:{encoded_recipient}?subject={subject.replace(' ', '%20')}&body={body}"
        QDesktopServices.openUrl(QUrl(mailto))

    def _show_about(self):
        # Pull supported-device lists live from the parser registry so this
        # stays in sync automatically when a new parser is added.
        chrono_names = ", ".join(p.NAME for p in chronograph_parsers()) or "(none)"
        group_names = ", ".join(p.NAME for p in group_parsers()) or "(none)"

        # Contact link — mailto opens the user's default email client with
        # subject/body pre-filled. The address itself isn't shown as plain
        # text; users see "click here" and only a hover-tooltip reveals the URL.
        # NOTE on display name: macOS Mail.app substitutes the recipient's
        # contact-card name if the email is in the user's address book. The
        # support address (support@loadscope.app) shouldn't be in anyone's
        # contacts by default, so the "Loadscope Support" display name we
        # provide here will show through cleanly for both Chad and friends.
        contact_subject = f"{APP_NAME} v{APP_VERSION} — feedback"
        contact_body = (
            "Hi,%0A%0A"
            "I'd like to request support for the following chronograph or "
            "analysis app:%0A%0A"
            "(your device/app)%0A%0A"
            "I've attached a sample CSV from my device.%0A%0A"
            "Thanks!"
        )
        # Format: mailto:"Display Name" <email>?subject=...&body=...
        # The %22 are URL-encoded double quotes around the display name; %20 is
        # space; %3C and %3E are < and >.
        display_name = f"{APP_NAME} Support"
        recipient = "support@loadscope.app"
        encoded_recipient = (
            f"%22{display_name.replace(' ', '%20')}%22%20%3C{recipient}%3E"
        )
        mailto = (
            f"mailto:{encoded_recipient}"
            f"?subject={contact_subject}"
            f"&body={contact_body}"
        )

        msg = QMessageBox(self)
        msg.setWindowTitle("About Loadscope™")
        msg.setIcon(QMessageBox.Information)
        msg.setTextFormat(Qt.RichText)
        msg.setTextInteractionFlags(Qt.TextBrowserInteraction)
        msg.setText(
            f"<b>Loadscope™</b><br>"
            f"Version {APP_VERSION}<br><br>"
            f"<b>Supported chronographs:</b><br>{chrono_names}<br><br>"
            f"<b>Supported target-group apps:</b><br>{group_names}<br><br>"
            f"<i>More devices and apps will be added in future updates. "
            f"If you'd like support for a chronograph or analysis app that's "
            f"not listed, "
            f'<a href="{mailto}">click here to contact the developers</a> '
            f"and send a sample CSV from your device.</i><br><br>"
            f"<b>Disclaimer:</b> Loadscope is a data-analysis tool. It does "
            f"not provide load data and is not a substitute for safe "
            f"handloading practice. See Tools → View Disclaimer for the full text."
        )
        msg.exec_()


def main():
    # RifleLoadApp = QApplication subclass that intercepts macOS QFileOpenEvent
    # so CSVs dragged onto the Dock icon (or right-click → Open With) reach
    # the window's handle_external_files() handler.
    app = RifleLoadApp(sys.argv)
    app.setStyleSheet(theme.application_stylesheet())

    # v0.14.2: Loadscope collapses Excel's ribbon (a GLOBAL, persistent
    # Excel preference) while a workbook is open. Restore the user's prior
    # ribbon state when Loadscope quits so we never leave a customer's
    # Excel permanently altered. The demo tour also restores on its own
    # close; both paths are idempotent + no-op if nothing was minimized.
    try:
        from excel_chrome import restore_excel_chrome
        app.aboutToQuit.connect(restore_excel_chrome)
    except Exception:
        pass  # never block startup on the chrome-restore wiring

    # Install the opt-in crash reporter so unhandled exceptions show a friendly
    # dialog with copy/email options instead of silent-crash-to-Console-app.
    try:
        import crash_reporter
        crash_reporter.install()
    except ImportError:
        pass  # safe to skip if module isn't available for some reason

    # License gate — v0.14 changed from a hard wall to a soft splash. On the
    # very first launch (no valid key, splash never dismissed) we offer the
    # user three choices: try the demo, enter a license key, or purchase. On
    # subsequent launches the user just lands directly in whichever mode they
    # have — licensed users skip the splash entirely.
    state = app_license.license_state()
    splash_choice = None  # set if the splash fired, used to auto-launch tour
    if state != "valid" and app_license.should_show_first_launch_splash():
        from splash_dialog import (
            FirstLaunchSplash,
            CHOICE_DEMO,
            CHOICE_LICENSE,
            CHOICE_PURCHASE,
            CHOICE_CANCEL,
        )
        splash = FirstLaunchSplash()
        splash.exec_()
        splash_choice = splash.choice
        if splash_choice == CHOICE_LICENSE:
            # User wants to enter a key — open the existing license dialog.
            # If they enter a valid key, license_state() becomes 'valid' on
            # the next call. If they Cancel out, fall through to demo mode.
            show_license_dialog(revoked=(state == "invalid"))
        # Whatever they picked (or didn't), mark the splash dismissed so
        # they don't see it every launch.
        app_license.mark_first_launch_splash_seen()
    # No hard quit gate anymore — the app always proceeds, either in
    # licensed mode or demo mode.

    # First-launch disclaimer — must be accepted before the user can use the app.
    # Tracked in config; user only sees it once unless DISCLAIMER_VERSION bumps.
    if needs_disclaimer():
        if not show_disclaimer():
            # User clicked Quit — don't proceed
            return 0

    # Resolve which project folder to use — config first, wizard if missing.
    project_folder = app_config.get_project_folder()

    if project_folder is None:
        # First-time run (or config got deleted). Walk the user through setup.
        chosen = run_wizard()
        if chosen is None:
            return 0
        app_config.set_project_folder(chosen)
        project_folder = chosen

    win = MainWindow(project_folder)
    app.set_main_window(win)
    win.show()
    win.raise_()
    win.activateWindow()

    # If the previous launch attempted a self-install and the helper script
    # bailed out, surface that so the user knows what happened. The error
    # log is consumed (deleted) on read so we don't keep nagging.
    last_install_err = installer.consume_last_install_error()
    if last_install_err:
        QTimer.singleShot(
            500,
            lambda: QMessageBox.warning(
                win,
                "Last update didn't install",
                "Loadscope tried to install an update but couldn't finish:\n\n"
                f"{last_install_err}\n\n"
                "You're still on the previous version. Use the update banner "
                "to try again, or download the new version manually from GitHub."
            ),
        )

    # First-launch tutorial — only shows if the user hasn't seen this
    # tutorial version yet. Tracked via config; users see it once.
    if needs_tutorial():
        QTimer.singleShot(300, lambda: show_tutorial(parent=win))

    # v0.14.2: splash → tour auto-launch. When the user picks "Try the
    # Free Demo" on the first-launch splash, automatically open the demo
    # tour after the main window finishes painting. Without this the
    # splash captures the choice but nothing visibly happens — user
    # lands on the drop-zone in demo mode wondering where the tour is.
    if splash_choice is not None:
        try:
            from splash_dialog import CHOICE_DEMO as _CHOICE_DEMO
            if splash_choice == _CHOICE_DEMO:
                # 800ms delay so the main window paints first
                QTimer.singleShot(800, lambda: win._open_demo_tour())
        except Exception:
            pass  # never block the app on a tour-launch failure

    # CSVs passed on the command line (e.g., dragged onto the .app icon while
    # it wasn't running — py2app forwards those via sys.argv on launch).
    # Drag-on-icon while the app IS running goes through QFileOpenEvent instead.
    cli_files = [a for a in sys.argv[1:] if a.lower().endswith(".csv")]
    if cli_files:
        QTimer.singleShot(150, lambda: win.handle_external_files(cli_files))

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
