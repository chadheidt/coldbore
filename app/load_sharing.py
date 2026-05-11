"""
Load sharing — export and import individual loads as .coldbore files.

A .coldbore file is a small JSON document containing the suggested winning
load from a workbook (cartridge, components, charge, jump, performance stats).
Friends can email/AirDrop these to each other.

Tools menu in the app:
  - Export Suggested Load…  → writes a .coldbore file next to the workbook
  - Import Shared Load…     → opens a .coldbore file and shows its contents
                              in a read-only dialog the user can copy from

This v1 keeps it simple — imports SHOW the load rather than writing into the
recipient's workbook. Less risk of clobbering data, and the recipient has to
consciously decide whether to try the load (which is the right interaction
for safety-adjacent data).
"""

import json
import os
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from version import APP_NAME, APP_VERSION
from load_card import LOAD_LOG_FIELDS, _extract_bullet_weight  # reuse


COLDBORE_FORMAT = "coldbore.load"
COLDBORE_VERSION = 1
FILE_EXT = ".coldbore"


# ============================================================
# Export
# ============================================================

def export_load(workbook_path):
    """Read the suggested winning load from the workbook and write it to a
    .coldbore file next to the workbook. Returns the path to the written file."""
    wb = load_workbook(workbook_path, data_only=True, keep_vba=False)

    def cell(sheet, coord):
        if sheet not in wb.sheetnames:
            return None
        return wb[sheet][coord].value

    # Pull the same fields the load card uses
    bullet_str = cell(*LOAD_LOG_FIELDS["bullet"])
    components = {
        "bullet": bullet_str,
        "bullet_weight_gr": _extract_bullet_weight(bullet_str),
        "powder": cell(*LOAD_LOG_FIELDS["powder"]),
        "primer": cell(*LOAD_LOG_FIELDS["primer"]),
        "brass": cell(*LOAD_LOG_FIELDS["brass"]),
        "cbto": cell(*LOAD_LOG_FIELDS["cbto"]),
        "off_lands": cell(*LOAD_LOG_FIELDS["off_lands"]),
    }
    rifle = {
        "cartridge": cell(*LOAD_LOG_FIELDS["cartridge"]),
        "rifle": cell(*LOAD_LOG_FIELDS["rifle"]),
        "barrel": cell(*LOAD_LOG_FIELDS["barrel"]),
        "optic": cell(*LOAD_LOG_FIELDS["optic"]),
        "distance_yd": cell(*LOAD_LOG_FIELDS["distance"]),
    }
    load_data = {
        "charge_gr": cell("Charts", "B3"),     # suggested charge
        "jump_in": cell("Seating Depth", "D2"),  # suggested jump
    }
    performance = {
        "avg_velocity_fps": cell("Charts", "E5"),
        "sd_fps": cell("Charts", "G3"),
        "group_in": cell("Charts", "G4"),
        "mean_radius_in": cell("Seating Depth", "L2"),
    }

    payload = {
        "format": COLDBORE_FORMAT,
        "format_version": COLDBORE_VERSION,
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "exported_by": f"{APP_NAME} v{APP_VERSION}",
        "rifle": rifle,
        "components": components,
        "load_data": load_data,
        "performance": performance,
        "notes": "",
    }

    # Filename: "<Cartridge> <Bullet wt><Powder> <charge>gr.coldbore"
    parts = []
    if rifle["cartridge"]:
        parts.append(str(rifle["cartridge"]))
    if components["bullet_weight_gr"]:
        parts.append(f"{components['bullet_weight_gr']:g}gr")
    if components["powder"]:
        parts.append(str(components["powder"]))
    if load_data["charge_gr"]:
        parts.append(f"@{load_data['charge_gr']:g}gr")
    base = " ".join(parts) if parts else "shared load"
    base = base.replace("/", "-")  # avoid path separators in name

    project_dir = os.path.dirname(os.path.abspath(workbook_path))
    out_dir = os.path.join(project_dir, "Shared Loads")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{base}{FILE_EXT}")

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, default=str)

    return out_path


# ============================================================
# Import
# ============================================================

class LoadFileError(Exception):
    """Raised when a file isn't a valid .coldbore load."""


def import_load(file_path):
    """Read and validate a .coldbore file. Returns the parsed dict.
    Raises LoadFileError if the file isn't valid."""
    try:
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        raise LoadFileError(f"Couldn't read file: {e}") from e

    if not isinstance(data, dict):
        raise LoadFileError("File doesn't contain a load object.")

    fmt = data.get("format")
    if fmt != COLDBORE_FORMAT:
        raise LoadFileError(
            f"This file isn't a True Zero shared load "
            f"(format={fmt!r}, expected {COLDBORE_FORMAT!r})."
        )

    fmt_v = data.get("format_version")
    if isinstance(fmt_v, int) and fmt_v > COLDBORE_VERSION:
        # Newer format version — try to read it anyway, but warn
        # (intentionally not raising; forward-compat best effort)
        pass

    return data


def format_load_for_display(data):
    """Format an imported load as human-readable text for the dialog."""
    rifle = data.get("rifle") or {}
    comp = data.get("components") or {}
    ld = data.get("load_data") or {}
    perf = data.get("performance") or {}

    def line(label, value, unit=""):
        if value is None or value == "":
            return None
        return f"  {label:<22} {value}{(' ' + unit) if unit else ''}"

    lines = []
    lines.append(f"SHARED LOAD — {rifle.get('cartridge') or '(unspecified cartridge)'}")
    lines.append(f"Exported by {data.get('exported_by', '?')} on {data.get('exported_at', '?')}")
    lines.append("")

    lines.append("RIFLE & SETUP")
    for label, key, unit in [
        ("Cartridge", "cartridge", ""),
        ("Rifle", "rifle", ""),
        ("Barrel", "barrel", ""),
        ("Optic", "optic", ""),
        ("Distance tested", "distance_yd", "yd"),
    ]:
        out = line(label, rifle.get(key), unit)
        if out:
            lines.append(out)

    lines.append("")
    lines.append("COMPONENTS")
    for label, key, unit in [
        ("Bullet", "bullet", ""),
        ("Bullet weight", "bullet_weight_gr", "gr"),
        ("Powder", "powder", ""),
        ("Primer", "primer", ""),
        ("Brass", "brass", ""),
        ("CBTO", "cbto", ""),
        ("Distance to lands", "off_lands", ""),
    ]:
        out = line(label, comp.get(key), unit)
        if out:
            lines.append(out)

    lines.append("")
    lines.append("LOAD DATA")
    for label, key, unit in [
        ("Powder charge", "charge_gr", "gr"),
        ("Bullet jump", "jump_in", "in"),
    ]:
        out = line(label, ld.get(key), unit)
        if out:
            lines.append(out)

    lines.append("")
    lines.append("PERFORMANCE")
    for label, key, unit in [
        ("Avg velocity", "avg_velocity_fps", "fps"),
        ("SD", "sd_fps", "fps"),
        ("Group", "group_in", "in"),
        ("Mean radius", "mean_radius_in", "in"),
    ]:
        out = line(label, perf.get(key), unit)
        if out:
            lines.append(out)

    if data.get("notes"):
        lines.append("")
        lines.append("NOTES")
        lines.append(f"  {data['notes']}")

    lines.append("")
    lines.append("⚠  Always cross-reference loads against published reloading manuals.")
    lines.append("   Start below this charge and work up. Watch for pressure signs.")

    return "\n".join(lines)
